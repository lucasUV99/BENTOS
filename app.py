"""
BENTOS - Sistema de Gestión de Bitácoras MSC
Interfaz Gráfica Principal
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import os
import sys
import json
import hashlib
import socket
import atexit
import tempfile
import platform
import threading
from datetime import datetime, timedelta
from pathlib import Path
from CTkMessagebox import CTkMessagebox
from tkcalendar import DateEntry

# Intentar importar tkinterdnd2 para drag & drop
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DRAG_DROP_AVAILABLE = True
except ImportError:
    DRAG_DROP_AVAILABLE = False
    print("⚠️ tkinterdnd2 no disponible. Drag & drop deshabilitado.")

# Configurar path para importar módulos backend
# En modo PyInstaller, los archivos están en sys._MEIPASS
if getattr(sys, 'frozen', False):
    _base_path = sys._MEIPASS
else:
    _base_path = os.path.dirname(os.path.abspath(__file__))

sys.path.insert(0, os.path.join(_base_path, 'backend'))

from firebase_manager import FirebaseManager
from pdf_parser_v2 import BitacoraParser
from updater import UpdateManager, APP_VERSION, aplicar_actualizacion_pendiente


# Configuración de tema
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class SplashScreen:
    """Pantalla de inicio con animación"""
    
    def __init__(self, parent_callback):
        self.root = ctk.CTk()
        self.root.title("BENTOS")
        self.parent_callback = parent_callback
        
        # Configurar ventana
        width = 600
        height = 400
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.overrideredirect(True)  # Sin bordes
        
        # Frame principal con degradado
        self.main_frame = ctk.CTkFrame(
            self.root,
            fg_color=("#0A4D68", "#0A4D68"),
            corner_radius=20
        )
        self.main_frame.pack(fill="both", expand=True, padx=2, pady=2)
        
        # Logo/Título
        self.title_label = ctk.CTkLabel(
            self.main_frame,
            text="BENTOS",
            font=ctk.CTkFont(family="Segoe UI", size=72, weight="bold"),
            text_color="#05BFDB"
        )
        self.title_label.place(relx=0.5, rely=0.35, anchor="center")
        
        # Subtítulo
        self.subtitle = ctk.CTkLabel(
            self.main_frame,
            text="Sistema de Gestión de Bitácoras Electrónicas",
            font=ctk.CTkFont(size=16),
            text_color="#88D4E8"
        )
        self.subtitle.place(relx=0.5, rely=0.5, anchor="center")
        
        # Descripción
        self.desc = ctk.CTkLabel(
            self.main_frame,
            text="Certificación MSC • Pesquera Quintero S.A.",
            font=ctk.CTkFont(size=12),
            text_color="#B8E4F0"
        )
        self.desc.place(relx=0.5, rely=0.6, anchor="center")
        
        # Barra de progreso
        self.progress = ctk.CTkProgressBar(
            self.main_frame,
            width=400,
            height=8,
            corner_radius=4,
            fg_color="#1A5F7A",
            progress_color="#05BFDB"
        )
        self.progress.place(relx=0.5, rely=0.75, anchor="center")
        self.progress.set(0)
        
        # Status
        self.status_label = ctk.CTkLabel(
            self.main_frame,
            text="Iniciando sistema...",
            font=ctk.CTkFont(size=11),
            text_color="#88D4E8"
        )
        self.status_label.place(relx=0.5, rely=0.85, anchor="center")
        
        # Iniciar animación
        self.progress_value = 0
        self.animate_progress()
        
    def animate_progress(self):
        """Anima la barra de progreso"""
        if self.progress_value < 1.0:
            self.progress_value += 0.02
            self.progress.set(self.progress_value)
            
            # Actualizar status
            if self.progress_value < 0.3:
                self.status_label.configure(text="Cargando módulos...")
            elif self.progress_value < 0.6:
                self.status_label.configure(text="Conectando con la nube...")
            elif self.progress_value < 0.9:
                self.status_label.configure(text="Inicializando interfaz...")
            else:
                self.status_label.configure(text="¡Listo!")
            
            self.root.after(50, self.animate_progress)
        else:
            # Cerrar splash y abrir app principal
            self.root.after(500, self.close_splash)
    
    def close_splash(self):
        """Cierra el splash y abre la app principal"""
        self.root.destroy()
        self.parent_callback()
    
    def show(self):
        """Muestra el splash screen"""
        self.root.mainloop()


class LoginScreen:
    """Pantalla de inicio de sesión con protección contra intentos fallidos"""
    
    MAX_INTENTOS = 4
    TIEMPO_BLOQUEO = 30  # minutos
    
    # Lockout file: en exe usa carpeta junto al ejecutable, en dev usa config/
    if getattr(sys, 'frozen', False):
        _lockout_dir = os.path.join(os.path.dirname(sys.executable), "config")
    else:
        _lockout_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
    LOCKOUT_FILE = os.path.join(_lockout_dir, "lockout.json")
    
    def __init__(self, on_success):
        self.on_success = on_success
        self.firebase = FirebaseManager()
        
        # Cargar estado de bloqueo
        self._cargar_estado_bloqueo()
        
        # Crear ventana
        self.root = ctk.CTk()
        self.root.title("BENTOS - Inicio de Sesión")
        
        # Centrar ventana
        width, height = 480, 560
        x = (self.root.winfo_screenwidth() - width) // 2
        y = (self.root.winfo_screenheight() - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.resizable(False, False)
        
        # Inicializar credenciales en Firebase si no existen
        if self.firebase.db:
            self.firebase.inicializar_credenciales()
        
        self._crear_ui()
        
        # Cerrar = salir de la aplicación
        self.root.protocol("WM_DELETE_WINDOW", self._cerrar_app)
    
    def _cargar_estado_bloqueo(self):
        """Carga el estado de bloqueo desde archivo local"""
        self.intentos_fallidos = 0
        self.bloqueado_hasta = None
        
        try:
            if os.path.exists(self.LOCKOUT_FILE):
                with open(self.LOCKOUT_FILE, 'r') as f:
                    data = json.load(f)
                    self.intentos_fallidos = data.get('intentos_fallidos', 0)
                    lockout = data.get('bloqueado_hasta')
                    if lockout:
                        self.bloqueado_hasta = datetime.fromisoformat(lockout)
                        # Si ya pasó el bloqueo, resetear
                        if datetime.now() >= self.bloqueado_hasta:
                            self.intentos_fallidos = 0
                            self.bloqueado_hasta = None
                            self._guardar_estado_bloqueo()
        except Exception:
            pass
    
    def _guardar_estado_bloqueo(self):
        """Guarda estado de bloqueo en archivo local"""
        try:
            os.makedirs(os.path.dirname(self.LOCKOUT_FILE), exist_ok=True)
            data = {
                'intentos_fallidos': self.intentos_fallidos,
                'bloqueado_hasta': self.bloqueado_hasta.isoformat() if self.bloqueado_hasta else None
            }
            with open(self.LOCKOUT_FILE, 'w') as f:
                json.dump(data, f)
        except Exception:
            pass
    
    def _crear_ui(self):
        """Crea la interfaz de login"""
        # Frame principal
        main = ctk.CTkFrame(self.root, fg_color=("#0A4D68", "#0A4D68"), corner_radius=0)
        main.pack(fill="both", expand=True)
        
        # Logo
        ctk.CTkLabel(
            main, text="🐟",
            font=ctk.CTkFont(size=56)
        ).pack(pady=(45, 0))
        
        ctk.CTkLabel(
            main, text="BENTOS",
            font=ctk.CTkFont(size=38, weight="bold"),
            text_color="#05BFDB"
        ).pack(pady=(5, 0))
        
        ctk.CTkLabel(
            main, text="Gestión de Bitácoras Electrónicas",
            font=ctk.CTkFont(size=13),
            text_color="#88D4E8"
        ).pack(pady=(2, 25))
        
        # Card de login
        card = ctk.CTkFrame(main, fg_color="#0D1B2A", corner_radius=15)
        card.pack(padx=55, fill="x")
        
        # Título del card
        ctk.CTkLabel(
            card, text="🔐  Inicio de Sesión",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#6BCFE0"
        ).pack(pady=(22, 18), padx=30, anchor="w")
        
        # Campo ID
        ctk.CTkLabel(
            card, text="Identificador",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#5A8A9A"
        ).pack(padx=30, anchor="w")
        
        self.id_entry = ctk.CTkEntry(
            card, height=42,
            placeholder_text="Ingresa el ID de acceso",
            border_color="#1B3A4B",
            fg_color="#111D2E",
            corner_radius=8,
            font=ctk.CTkFont(size=13)
        )
        self.id_entry.pack(padx=30, fill="x", pady=(4, 12))
        
        # Campo Contraseña
        ctk.CTkLabel(
            card, text="Contraseña",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#5A8A9A"
        ).pack(padx=30, anchor="w")
        
        self.clave_entry = ctk.CTkEntry(
            card, height=42,
            placeholder_text="Ingresa la contraseña",
            show="●",
            border_color="#1B3A4B",
            fg_color="#111D2E",
            corner_radius=8,
            font=ctk.CTkFont(size=13)
        )
        self.clave_entry.pack(padx=30, fill="x", pady=(4, 18))
        
        # Botón de login
        self.login_btn = ctk.CTkButton(
            card, text="Acceder",
            height=44,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#05BFDB",
            hover_color="#057A8F",
            corner_radius=8,
            command=self._intentar_login
        )
        self.login_btn.pack(pady=(0, 22), padx=30, fill="x")
        
        # Mensaje de estado
        self.msg_label = ctk.CTkLabel(
            main, text="",
            font=ctk.CTkFont(size=12),
            text_color="#FF5252",
            wraplength=380
        )
        self.msg_label.pack(pady=(15, 0))
        
        # Indicador de conexión
        db_icon = "🟢" if self.firebase.db else "🔴"
        db_text = "Conectado a la nube" if self.firebase.db else "Sin conexión a la nube"
        db_color = "#4CAF50" if self.firebase.db else "#FF5252"
        
        ctk.CTkLabel(
            main, text=f"{db_icon}  {db_text}",
            font=ctk.CTkFont(size=10),
            text_color=db_color
        ).pack(side="bottom", pady=15)
        
        # Empresa
        ctk.CTkLabel(
            main, text="Pesquera Quintero S.A.",
            font=ctk.CTkFont(size=10),
            text_color="#3A6A7A"
        ).pack(side="bottom", pady=(0, 2))
        
        # Bind Enter
        self.root.bind('<Return>', lambda e: self._intentar_login())
        
        # Verificar si está bloqueado al iniciar
        if self.bloqueado_hasta and datetime.now() < self.bloqueado_hasta:
            self._mostrar_bloqueo()
    
    def _intentar_login(self):
        """Intenta iniciar sesión validando contra Firebase"""
        # Verificar bloqueo
        if self.bloqueado_hasta and datetime.now() < self.bloqueado_hasta:
            self._mostrar_bloqueo()
            return
        
        id_input = self.id_entry.get().strip()
        clave_input = self.clave_entry.get().strip()
        
        if not id_input or not clave_input:
            self.msg_label.configure(
                text="⚠️  Completa ambos campos",
                text_color="#FF9800"
            )
            return
        
        # Verificar conexión a Firebase
        if not self.firebase.db:
            self.msg_label.configure(
                text="❌  Sin conexión a la base de datos.\nNo es posible verificar las credenciales.",
                text_color="#FF5252"
            )
            return
        
        # Validar credenciales
        if self.firebase.verificar_credenciales(id_input, clave_input):
            # Éxito
            self.intentos_fallidos = 0
            self.bloqueado_hasta = None
            self._guardar_estado_bloqueo()
            self.root.quit()
            self.root.destroy()
            self.on_success()
        else:
            # Fallo
            self.intentos_fallidos += 1
            restantes = self.MAX_INTENTOS - self.intentos_fallidos
            
            if restantes <= 0:
                # Bloquear
                self.bloqueado_hasta = datetime.now() + timedelta(minutes=self.TIEMPO_BLOQUEO)
                self._guardar_estado_bloqueo()
                self._mostrar_bloqueo()
            else:
                self._guardar_estado_bloqueo()
                self.msg_label.configure(
                    text=f"❌  Credenciales incorrectas — {restantes} intento(s) restante(s)",
                    text_color="#FF5252"
                )
                # Limpiar contraseña
                self.clave_entry.delete(0, "end")
    
    def _mostrar_bloqueo(self):
        """Muestra cuenta regresiva de bloqueo y deshabilita campos"""
        self.login_btn.configure(state="disabled", fg_color="#333333", text="Bloqueado")
        self.id_entry.configure(state="disabled")
        self.clave_entry.configure(state="disabled")
        self._actualizar_countdown()
    
    def _actualizar_countdown(self):
        """Actualiza la cuenta regresiva del bloqueo"""
        if not self.bloqueado_hasta:
            return
        
        try:
            restante = self.bloqueado_hasta - datetime.now()
            if restante.total_seconds() <= 0:
                # Desbloquear
                self.intentos_fallidos = 0
                self.bloqueado_hasta = None
                self._guardar_estado_bloqueo()
                self.login_btn.configure(state="normal", fg_color="#05BFDB", text="Acceder")
                self.id_entry.configure(state="normal")
                self.clave_entry.configure(state="normal")
                self.msg_label.configure(
                    text="✅  Bloqueo finalizado. Puedes intentar de nuevo.",
                    text_color="#4CAF50"
                )
            else:
                minutos = int(restante.total_seconds() // 60)
                segundos = int(restante.total_seconds() % 60)
                self.msg_label.configure(
                    text=f"🔒  Acceso bloqueado por demasiados intentos fallidos\nEspera {minutos:02d}:{segundos:02d} para volver a intentar",
                    text_color="#FF5252"
                )
                self.root.after(1000, self._actualizar_countdown)
        except (tk.TclError, Exception):
            pass
    
    def _cerrar_app(self):
        """Cierra la aplicación si se cierra la ventana de login"""
        self.root.destroy()
    
    def show(self):
        """Muestra la pantalla de login"""
        self.root.mainloop()


class BentosApp:
    """Aplicación principal BENTOS"""
    
    def __init__(self):
        # Crear root con soporte para drag & drop si está disponible
        if DRAG_DROP_AVAILABLE:
            self.root = TkinterDnD.Tk()
            self.root.withdraw()  # Ocultar temporalmente
            self.root.deiconify()  # Mostrar de nuevo (fix para TkinterDnD)
        else:
            self.root = ctk.CTk()
        
        self.root.title("BENTOS - Sistema de Gestión de Bitácoras MSC")
        
        # Configurar ventana maximizada (NO pantalla completa)
        self.root.state('zoomed')
        
        # Fondo de la ventana raíz — elimina la línea blanca entre sidebar y contenido
        try:
            self.root.configure(fg_color="#0D1B2A")
        except Exception:
            self.root.configure(bg="#0D1B2A")
        
        # Firebase manager
        self.firebase = FirebaseManager()
        
        # Archivo de estado de la aplicación (para notificaciones perdidas)
        if getattr(sys, 'frozen', False):
            _state_dir = os.path.join(os.path.dirname(sys.executable), "config")
        else:
            _state_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
        self.APP_STATE_FILE = os.path.join(_state_dir, "app_state.json")
        
        # Variables
        self.current_section = "subir"  # subir o buscar
        self.loading_animation_running = False
        self.notificaciones = []  # Lista de notificaciones
        self.archivos_pendientes = []  # Archivos para procesar
        self.resultados_parseados = {}  # Cache de resultados parseados {archivo: resultado}
        self.internet_conectado = True  # Estado de conexión a internet
        self._app_closing = False  # Flag para detener ciclos after al cerrar
        self._temp_files = []  # Archivos temporales a limpiar al cerrar
        self._viajes_conocidos = set()  # IDs de viajes ya conocidos (para detectar nuevos)
        self._viajes_inicializados = False  # Flag para saber si ya se cargó el set inicial
        self._notif_timer_id = None  # ID del timer de notificaciones (evitar duplicados)
        
        # Cargar estado previo de viajes (para detectar cambios mientras estuvo apagado)
        self._cargar_estado_app()
        
        # Crear UI
        self.create_ui()
        
        # Registrar atajos de teclado
        self._registrar_atajos_teclado()
        
        # Protocolo de cierre con confirmación
        self.root.protocol("WM_DELETE_WINDOW", self._confirmar_cierre)
        
        # Verificar conexión a internet al inicio
        self.verificar_conexion_internet_inicio()
        
        # Monitoreo periódico de conexión
        self.monitorear_conexion()
        
        # Verificar notificaciones periódicamente
        self.verificar_notificaciones()
        
        # Actualizar reloj de la barra de estado
        self._actualizar_reloj()
        
        # Verificar actualizaciones disponibles
        self._verificar_actualizacion()
        
    def create_ui(self):
        """Crea la interfaz de usuario"""
        
        # ===== HEADER PROFESIONAL =====
        self.header_frame = ctk.CTkFrame(
            self.root,
            height=70,
            corner_radius=0,
            fg_color=("#0A4D68", "#0A4D68")
        )
        self.header_frame.pack(fill="x", side="top")
        self.header_frame.pack_propagate(False)
        
        # Logo y título
        title_frame = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        title_frame.pack(side="left", padx=25, pady=10)
        
        ctk.CTkLabel(
            title_frame,
            text="🐟",
            font=ctk.CTkFont(size=32)
        ).pack(side="left", padx=(0, 8))
        
        title_text_frame = ctk.CTkFrame(title_frame, fg_color="transparent")
        title_text_frame.pack(side="left")
        
        ctk.CTkLabel(
            title_text_frame,
            text="BENTOS",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#05BFDB"
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            title_text_frame,
            text="Gestión de Bitácoras Electrónicas",
            font=ctk.CTkFont(size=11),
            text_color="#6BCFE0"
        ).pack(anchor="w")
        
        # Controles del header (derecha)
        controls_frame = ctk.CTkFrame(self.header_frame, fg_color="transparent")
        controls_frame.pack(side="right", padx=25)
        
        # Indicadores de estado (fila superior)
        status_row = ctk.CTkFrame(controls_frame, fg_color="transparent")
        status_row.pack(pady=(0, 5))
        
        # Indicador de base de datos
        db_status = "🟢" if self.firebase.db else "🔴"
        self.db_status_label = ctk.CTkLabel(
            status_row,
            text=f"{db_status} Base de datos",
            font=ctk.CTkFont(size=10),
            text_color="#4CAF50" if self.firebase.db else "#FF5252"
        )
        self.db_status_label.pack(side="left", padx=(0, 12))
        
        # Separador visual
        ctk.CTkLabel(
            status_row,
            text="│",
            font=ctk.CTkFont(size=10),
            text_color="#2E8B9E"
        ).pack(side="left", padx=(0, 12))
        
        # Indicador de internet
        self.internet_status_label = ctk.CTkLabel(
            status_row,
            text="🌐 Verificando...",
            font=ctk.CTkFont(size=10),
            text_color="#B8E4F0"
        )
        self.internet_status_label.pack(side="left")
        
        # Botones de acción (fila inferior)
        buttons_row = ctk.CTkFrame(controls_frame, fg_color="transparent")
        buttons_row.pack()
        
        # Contenedor para botón de notificaciones con badge
        notif_container = ctk.CTkFrame(buttons_row, fg_color="transparent")
        notif_container.pack(side="left", padx=4)
        
        self.btn_notificaciones = ctk.CTkButton(
            notif_container,
            text="🔔",
            font=ctk.CTkFont(size=16),
            width=40,
            height=32,
            corner_radius=8,
            fg_color="#2E8B9E",
            hover_color="#1A5F7A",
            command=self.mostrar_notificaciones
        )
        self.btn_notificaciones.pack()
        self._crear_tooltip(self.btn_notificaciones, "Centro de notificaciones")
        
        # Badge de notificaciones
        self.notif_badge = ctk.CTkLabel(
            notif_container,
            text="",
            font=ctk.CTkFont(size=8, weight="bold"),
            text_color="white",
            fg_color="#FF3B30",
            corner_radius=8,
            width=16,
            height=16
        )
        
        # Botón de reporte de bugs
        self.btn_bug = ctk.CTkButton(
            buttons_row,
            text="🐛",
            font=ctk.CTkFont(size=16),
            width=40,
            height=32,
            corner_radius=8,
            fg_color="#2E8B9E",
            hover_color="#1A5F7A",
            command=self.abrir_reporte_bug
        )
        self.btn_bug.pack(side="left", padx=4)
        self._crear_tooltip(self.btn_bug, "Reportar un problema")
        
        # ===== BARRA DE ESTADO (se empaqueta ANTES del sidebar para todo el ancho) =====
        self._crear_barra_estado()
        
        # ===== SIDEBAR PROFESIONAL =====
        self.nav_frame = ctk.CTkFrame(
            self.root,
            width=260,
            corner_radius=0,
            fg_color=("#0D1B2A", "#0D1B2A")
        )
        self.nav_frame.pack(fill="y", side="left")
        self.nav_frame.pack_propagate(False)
        
        # Título del menú
        ctk.CTkLabel(
            self.nav_frame,
            text="MENÚ",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#5A7A8A"
        ).pack(pady=(25, 12), padx=20, anchor="w")
        
        # Botón: Subir Bitácoras
        self.btn_subir = ctk.CTkButton(
            self.nav_frame,
            text="  📤  Subir Bitácoras",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=52,
            corner_radius=10,
            anchor="w",
            fg_color="#05BFDB",
            hover_color="#057A8F",
            border_width=0,
            command=lambda: self.show_section_animated("subir")
        )
        self.btn_subir.pack(pady=(0, 6), padx=16, fill="x")
        self._crear_tooltip(self.btn_subir, "Subir PDFs de bitácoras (F1)")
        
        # Botón: Buscar Datos
        self.btn_buscar = ctk.CTkButton(
            self.nav_frame,
            text="  🔍  Buscar Datos",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=52,
            corner_radius=10,
            anchor="w",
            fg_color="#1B3A4B",
            hover_color="#15303F",
            border_width=0,
            command=lambda: self.show_section_animated("buscar")
        )
        self.btn_buscar.pack(pady=(0, 6), padx=16, fill="x")
        self._crear_tooltip(self.btn_buscar, "Buscar y analizar datos MSC (F2)")
        
        # Botón: Administrar Datos
        self.btn_admin = ctk.CTkButton(
            self.nav_frame,
            text="  ⚙️  Administrar Datos",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=52,
            corner_radius=10,
            anchor="w",
            fg_color="#1B3A4B",
            hover_color="#15303F",
            border_width=0,
            command=self.abrir_panel_administrador
        )
        self.btn_admin.pack(pady=(0, 6), padx=16, fill="x")
        self._crear_tooltip(self.btn_admin, "Administrar y eliminar bitácoras")
        
        # Separador
        ctk.CTkFrame(
            self.nav_frame,
            height=1,
            fg_color="#1B3A4B"
        ).pack(fill="x", pady=18, padx=20)
        
        # Sección de estadísticas
        ctk.CTkLabel(
            self.nav_frame,
            text="RESUMEN",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#5A7A8A"
        ).pack(pady=(0, 10), padx=20, anchor="w")
        
        self.stats_frame = ctk.CTkFrame(
            self.nav_frame,
            fg_color="#111D2E",
            corner_radius=10
        )
        self.stats_frame.pack(pady=(0, 10), padx=16, fill="x")
        
        self.update_stats()
        
        # Espaciador flexible para empujar el pie al fondo
        spacer = ctk.CTkFrame(self.nav_frame, fg_color="transparent")
        spacer.pack(fill="both", expand=True)
        
        # Pie del sidebar
        footer_frame = ctk.CTkFrame(self.nav_frame, fg_color="#111D2E", corner_radius=8)
        footer_frame.pack(padx=16, pady=(0, 12), fill="x")
        
        ctk.CTkLabel(
            footer_frame,
            text="👥 Uso multiusuario habilitado",
            font=ctk.CTkFont(size=10),
            text_color="#4CAF50"
        ).pack(pady=(8, 2))
        
        ctk.CTkLabel(
            footer_frame,
            text="v2.1 — Pesquera Quintero S.A.",
            font=ctk.CTkFont(size=9),
            text_color="#4A6A7A"
        ).pack(pady=(0, 8))
        
        # ===== CONTENT AREA =====
        self.content_frame = ctk.CTkFrame(
            self.root,
            corner_radius=0,
            fg_color=("#EAEFF2", "#1A1A2E")
        )
        self.content_frame.pack(fill="both", expand=True, side="right")
        
        # Crear secciones
        self.create_upload_section()
        self.create_search_section()
        
        # Mostrar sección inicial
        self.show_section("subir")
    
    def update_stats(self):
        """Actualiza estadísticas en el sidebar (no bloqueante)"""
        def _consultar():
            try:
                viajes = self.firebase.listar_viajes(limite=100) if self.firebase.db else []
                count = len(viajes)
            except Exception:
                count = "?"
            # Actualizar UI en el hilo principal
            if not self._app_closing:
                self.root.after(0, lambda: self._render_stats(count))
        
        threading.Thread(target=_consultar, daemon=True).start()
    
    def _render_stats(self, count):
        """Renderiza las estadísticas en el sidebar (llamar desde hilo principal)"""
        try:
            for widget in self.stats_frame.winfo_children():
                widget.destroy()
        except (tk.TclError, Exception):
            return
        
        stats = [
            ("🚢  Bitácoras en la nube", str(count), "#05BFDB"),
            ("📅  Última actualización", datetime.now().strftime("%d/%m/%Y %H:%M"), "#6BCFE0")
        ]
        
        for label, value, color in stats:
            row = ctk.CTkFrame(self.stats_frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=6)
            
            ctk.CTkLabel(
                row,
                text=label,
                font=ctk.CTkFont(size=10),
                text_color="#6A8A9A"
            ).pack(anchor="w")
            
            ctk.CTkLabel(
                row,
                text=value,
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=color
            ).pack(anchor="w", pady=(1, 0))
    
    def create_upload_section(self):
        """Crea la sección de subida de PDFs"""
        self.upload_section = ctk.CTkFrame(
            self.content_frame,
            fg_color="transparent"
        )
        
        # Título
        title_frame = ctk.CTkFrame(self.upload_section, fg_color="transparent")
        title_frame.pack(pady=(25, 10), padx=40, fill="x")
        
        ctk.CTkLabel(
            title_frame,
            text="📤 Subir Bitácoras Electrónicas",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            title_frame,
            text="Procesa PDFs de bitácoras de Sernapesca y almacena los datos en la nube",
            font=ctk.CTkFont(size=14),
            text_color=("#666666", "#AAAAAA")
        ).pack(anchor="w", pady=(5, 0))
        
        # Área de drop/upload
        self.drop_frame = ctk.CTkFrame(
            self.upload_section,
            corner_radius=15,
            fg_color=("#E8F4F8", "#1A3A4A"),
            border_width=3,
            border_color=("#05BFDB", "#05BFDB")
        )
        self.drop_frame.pack(pady=20, padx=40, fill="both", expand=True)
        
        # Efecto hover al pasar el mouse sobre el área de drop
        def on_drop_enter(e):
            try:
                self.drop_frame.configure(border_color="#0492A8", border_width=4)
            except (tk.TclError, Exception):
                pass
        
        def on_drop_leave(e):
            try:
                self.drop_frame.configure(border_color="#05BFDB", border_width=3)
            except (tk.TclError, Exception):
                pass
        
        self.drop_frame.bind("<Enter>", on_drop_enter)
        self.drop_frame.bind("<Leave>", on_drop_leave)
        
        # Contenido del drop area
        drop_content = ctk.CTkFrame(self.drop_frame, fg_color="transparent")
        drop_content.place(relx=0.5, rely=0.5, anchor="center")
        
        # Drag & drop
        if DRAG_DROP_AVAILABLE:
            try:
                self.drop_frame.drop_target_register(DND_FILES)
                self.drop_frame.dnd_bind('<<Drop>>', self.drop_files)
            except Exception:
                self.dnd_target = tk.Frame(self.drop_frame, bg="")
                self.dnd_target.place(relx=0, rely=0, relwidth=1, relheight=1)
                self.dnd_target.drop_target_register(DND_FILES)
                self.dnd_target.dnd_bind('<<Drop>>', self.drop_files)
                self.dnd_target.lift()
                self.dnd_target.bind('<Button-1>', lambda e: self.select_multiple_pdfs())
        
        ctk.CTkLabel(
            drop_content,
            text="📄",
            font=ctk.CTkFont(size=70)
        ).pack()
        
        dnd_status = "Selecciona PDFs o arrástralos aquí" if DRAG_DROP_AVAILABLE else "Selecciona archivos PDF"
        ctk.CTkLabel(
            drop_content,
            text=dnd_status,
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(pady=8)
        
        ctk.CTkLabel(
            drop_content,
            text="Formato: Bitácora Electrónica de Sernapesca • Hasta 5 archivos a la vez",
            font=ctk.CTkFont(size=12),
            text_color=("#666666", "#AAAAAA")
        ).pack()
        
        # Botón de selección
        self.upload_btn = ctk.CTkButton(
            drop_content,
            text="📁 Seleccionar Archivos PDF",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=50,
            width=280,
            corner_radius=10,
            fg_color="#05BFDB",
            hover_color="#057A8F",
            command=self.select_multiple_pdfs
        )
        self.upload_btn.pack(pady=18)
        
        # Lista de archivos seleccionados
        self.files_list_frame = ctk.CTkScrollableFrame(
            drop_content,
            width=600,
            height=150,
            fg_color=("#E0E8ED", "#1A2535"),
            corner_radius=10
        )
        
        # Progress bar (oculto)
        self.upload_progress = ctk.CTkProgressBar(drop_content, width=400, height=10, corner_radius=5)
        self.upload_status = ctk.CTkLabel(drop_content, text="", font=ctk.CTkFont(size=12),
                                          text_color=("#666666", "#AAAAAA"))
    
    def create_search_section(self):
        """Crea la sección de búsqueda de datos con calendarios y estadísticas"""
        self.search_section = ctk.CTkFrame(
            self.content_frame,
            fg_color="transparent"
        )
        
        # Título
        title_frame = ctk.CTkFrame(self.search_section, fg_color="transparent")
        title_frame.pack(pady=30, padx=40, fill="x")
        
        ctk.CTkLabel(
            title_frame,
            text="🔍 Buscar y Analizar Datos MSC",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            title_frame,
            text="Filtra bitácoras y calcula totales de captura para reportes de sostenibilidad",
            font=ctk.CTkFont(size=14),
            text_color=("#666666", "#AAAAAA")
        ).pack(anchor="w", pady=(5, 0))
        
        # Panel de filtros
        filter_frame = ctk.CTkFrame(
            self.search_section,
            corner_radius=15,
            fg_color=("#FFFFFF", "#1E1E1E")
        )
        filter_frame.pack(pady=20, padx=40, fill="x")
        
        # Título de filtros
        ctk.CTkLabel(
            filter_frame,
            text="📋 Filtros de Búsqueda",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(pady=15, padx=20, anchor="w")
        
        # Grid de filtros
        filters_grid = ctk.CTkFrame(filter_frame, fg_color="transparent")
        filters_grid.pack(pady=10, padx=20, fill="x")
        
        # ===== SECCIÓN 1: IDENTIFICACIÓN =====
        id_section = ctk.CTkFrame(filters_grid, fg_color="#F0F8FF", corner_radius=10, border_width=1, border_color="#05BFDB")
        id_section.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            id_section,
            text="🔍 IDENTIFICACIÓN",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#0A4D68"
        ).pack(anchor="w", padx=15, pady=(10, 5))
        
        id_row = ctk.CTkFrame(id_section, fg_color="transparent")
        id_row.pack(fill="x", padx=15, pady=(0, 10))
        
        # Embarcación
        nave_frame = ctk.CTkFrame(id_row, fg_color="transparent")
        nave_frame.pack(side="left", padx=(0, 10), expand=True, fill="x")
        
        ctk.CTkLabel(
            nave_frame,
            text="Embarcación:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.nave_combo = ctk.CTkComboBox(
            nave_frame,
            values=["Todas", "RAUTEN"],
            height=35,
            border_color="#05BFDB"
        )
        self.nave_combo.pack(fill="x", pady=3)
        self.nave_combo.set("Todas")
        
        # Capitán
        capitan_frame = ctk.CTkFrame(id_row, fg_color="transparent")
        capitan_frame.pack(side="left", padx=(10, 0), expand=True, fill="x")
        
        ctk.CTkLabel(
            capitan_frame,
            text="Capitán:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.capitan_entry = ctk.CTkEntry(
            capitan_frame,
            placeholder_text="Nombre del capitán...",
            height=35,
            border_color="#05BFDB"
        )
        self.capitan_entry.pack(fill="x", pady=3)
        
        # ===== SECCIÓN 2: RANGO TEMPORAL =====
        fecha_section = ctk.CTkFrame(filters_grid, fg_color="#FFF8F0", corner_radius=10, border_width=1, border_color="#FF9800")
        fecha_section.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            fecha_section,
            text="📅 PERÍODO DE OPERACIÓN",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#0A4D68"
        ).pack(anchor="w", padx=15, pady=(10, 5))
        
        fecha_row = ctk.CTkFrame(fecha_section, fg_color="transparent")
        fecha_row.pack(fill="x", padx=15, pady=(0, 10))
        
        # Fecha desde
        fecha_desde_frame = ctk.CTkFrame(fecha_row, fg_color="transparent")
        fecha_desde_frame.pack(side="left", padx=(0, 10), expand=True, fill="x")
        
        ctk.CTkLabel(
            fecha_desde_frame,
            text="Fecha Zarpe Desde:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.fecha_desde = DateEntry(
            fecha_desde_frame,
            width=25,
            background='#FF9800',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            font=('Arial', 10)
        )
        self.fecha_desde.pack(fill="x", pady=3)
        
        # Fecha hasta
        fecha_hasta_frame = ctk.CTkFrame(fecha_row, fg_color="transparent")
        fecha_hasta_frame.pack(side="left", padx=10, expand=True, fill="x")
        
        ctk.CTkLabel(
            fecha_hasta_frame,
            text="Fecha Zarpe Hasta:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.fecha_hasta = DateEntry(
            fecha_hasta_frame,
            width=25,
            background='#FF9800',
            foreground='white',
            borderwidth=2,
            date_pattern='dd/mm/yyyy',
            font=('Arial', 10)
        )
        self.fecha_hasta.pack(fill="x", pady=3)
        
        # Botones rápidos
        period_btns = ctk.CTkFrame(fecha_row, fg_color="transparent")
        period_btns.pack(side="left", padx=(10, 0))
        
        ctk.CTkLabel(
            period_btns,
            text="Atajos:",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color="#666666"
        ).pack(anchor="w")
        
        btn_row = ctk.CTkFrame(period_btns, fg_color="transparent")
        btn_row.pack()
        
        ctk.CTkButton(
            btn_row,
            text="Últimos 7 días",
            width=100,
            height=28,
            command=self.set_last_7_days,
            fg_color="#FF9800",
            hover_color="#F57C00",
            font=ctk.CTkFont(size=10)
        ).pack(side="left", padx=2)
        
        ctk.CTkButton(
            btn_row,
            text="Mes actual",
            width=90,
            height=28,
            command=self.set_current_month,
            fg_color="#FF9800",
            hover_color="#F57C00",
            font=ctk.CTkFont(size=10)
        ).pack(side="left", padx=2)
        
        ctk.CTkButton(
            btn_row,
            text="Año actual",
            width=90,
            height=28,
            command=self.set_current_year,
            fg_color="#FF9800",
            hover_color="#F57C00",
            font=ctk.CTkFont(size=10)
        ).pack(side="left", padx=2)
        
        # ===== SECCIÓN 3: FILTROS DE CAPTURA =====
        captura_section = ctk.CTkFrame(filters_grid, fg_color="#F0FFF0", corner_radius=10, border_width=1, border_color="#4CAF50")
        captura_section.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            captura_section,
            text="🐟 FILTRO POR ESPECIES",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#0A4D68"
        ).pack(anchor="w", padx=15, pady=(10, 5))
        
        captura_row = ctk.CTkFrame(captura_section, fg_color="transparent")
        captura_row.pack(fill="x", padx=15, pady=(0, 10))
        
        # Seleccionar especie objetivo
        especie_frame = ctk.CTkFrame(captura_row, fg_color="transparent")
        especie_frame.pack(side="left", padx=(0, 10), expand=True, fill="x")
        
        ctk.CTkLabel(
            especie_frame,
            text="Especie Objetivo:",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.especie_combo = ctk.CTkComboBox(
            especie_frame,
            values=["Todas", "Camarón nailon", "Langostino amarillo", "Langostino colorado", "Gamba"],
            height=35,
            border_color="#4CAF50",
            command=self.on_especie_change
        )
        self.especie_combo.pack(fill="x", pady=3)
        self.especie_combo.set("Todas")
        
        # Frame para captura mínima (inicialmente oculto)
        self.captura_min_frame = ctk.CTkFrame(captura_row, fg_color="transparent")
        
        ctk.CTkLabel(
            self.captura_min_frame,
            text="Captura Mín. (TON):",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.captura_min_entry = ctk.CTkEntry(
            self.captura_min_frame,
            placeholder_text="0",
            height=35,
            border_color="#4CAF50"
        )
        self.captura_min_entry.pack(fill="x", pady=3)
        
        # Frame para captura máxima (inicialmente oculto)
        self.captura_max_frame = ctk.CTkFrame(captura_row, fg_color="transparent")
        
        ctk.CTkLabel(
            self.captura_max_frame,
            text="Captura Máx. (TON):",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        self.captura_max_entry = ctk.CTkEntry(
            self.captura_max_frame,
            placeholder_text="Sin límite",
            height=35,
            border_color="#4CAF50"
        )
        self.captura_max_entry.pack(fill="x", pady=3)
        
        # Botones de acción
        action_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        action_frame.pack(pady=15, padx=20, fill="x")
        
        self.search_btn = ctk.CTkButton(
            action_frame,
            text="🔍 Buscar y Analizar",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            width=220,
            corner_radius=10,
            fg_color="#05BFDB",
            hover_color="#057A8F",
            command=self.search_data
        )
        self.search_btn.pack(side="left", padx=5)
        self._crear_tooltip(self.search_btn, "Buscar datos según filtros aplicados")
        
        self.clear_btn = ctk.CTkButton(
            action_frame,
            text="🔄 Limpiar",
            font=ctk.CTkFont(size=14),
            height=45,
            width=120,
            corner_radius=10,
            fg_color="#666666",
            hover_color="#444444",
            command=self.clear_filters
        )
        self.clear_btn.pack(side="left", padx=5)
        self._crear_tooltip(self.clear_btn, "Restablecer todos los filtros")
        
        # Panel de estadísticas (oculto inicialmente)
        self.stats_panel = ctk.CTkFrame(
            self.search_section,
            corner_radius=15,
            fg_color=("#E8F4F8", "#1A3A4A")
        )
        
        # Frame para loading/status de búsqueda (oculto por defecto)
        self.search_status_frame = ctk.CTkFrame(
            self.search_section,
            fg_color="transparent"
        )
        
        # Label de conteo de resultados (se muestra en los filtros)
        self.results_count_label = ctk.CTkLabel(
            action_frame,
            text="",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=("#05BFDB", "#05BFDB")
        )
        self.results_count_label.pack(side="left", padx=15)
    
    # ===== BARRA DE ESTADO Y ATAJOS =====
    
    def _crear_barra_estado(self):
        """Crea una barra de estado profesional en la parte inferior"""
        self.status_bar = ctk.CTkFrame(
            self.root,
            height=28,
            corner_radius=0,
            fg_color=("#0A4D68", "#0A4D68")
        )
        self.status_bar.pack(fill="x", side="bottom")
        self.status_bar.pack_propagate(False)
        
        # Lado izquierdo: estado de la app
        self.status_label = ctk.CTkLabel(
            self.status_bar,
            text="✅ Sistema listo",
            font=ctk.CTkFont(size=10),
            text_color="#88D4E8"
        )
        self.status_label.pack(side="left", padx=15)
        
        # Separador
        ctk.CTkFrame(
            self.status_bar,
            width=1,
            fg_color="#1A5F7A"
        ).pack(side="left", fill="y", pady=4)
        
        # Centro: atajos de teclado
        ctk.CTkLabel(
            self.status_bar,
            text="F1: Subir  |  F2: Buscar  |  F5: Actualizar",
            font=ctk.CTkFont(size=9),
            text_color="#5A9AAE"
        ).pack(side="left", padx=15)
        
        # Lado derecho: reloj
        self.clock_label = ctk.CTkLabel(
            self.status_bar,
            text="",
            font=ctk.CTkFont(size=10),
            text_color="#88D4E8"
        )
        self.clock_label.pack(side="right", padx=15)
        
        # Separador derecho
        ctk.CTkFrame(
            self.status_bar,
            width=1,
            fg_color="#1A5F7A"
        ).pack(side="right", fill="y", pady=4)
        
        # Versión
        ctk.CTkLabel(
            self.status_bar,
            text=f"BENTOS v{APP_VERSION}",
            font=ctk.CTkFont(size=9, weight="bold"),
            text_color="#5A9AAE"
        ).pack(side="right", padx=10)
    
    def _actualizar_reloj(self):
        """Actualiza el reloj de la barra de estado"""
        if self._app_closing:
            return
        try:
            ahora = datetime.now().strftime("%d/%m/%Y  %H:%M:%S")
            self.clock_label.configure(text=f"🕐 {ahora}")
        except (tk.TclError, Exception):
            return
        self.root.after(1000, self._actualizar_reloj)
    
    def _actualizar_estado(self, mensaje, icono="ℹ️"):
        """Actualiza el mensaje de la barra de estado"""
        try:
            self.status_label.configure(text=f"{icono} {mensaje}")
        except (tk.TclError, Exception):
            pass
    
    def _registrar_atajos_teclado(self):
        """Registra atajos de teclado globales"""
        self.root.bind('<F1>', lambda e: self.show_section_animated("subir"))
        self.root.bind('<F2>', lambda e: self.show_section_animated("buscar"))
        self.root.bind('<F5>', lambda e: self._refrescar_aplicacion())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
    
    def _refrescar_aplicacion(self):
        """Refresca estadísticas y estado de conexión"""
        self._actualizar_estado("Actualizando...", "🔄")
        self.update_stats()
        self.internet_conectado = self.verificar_conexion_internet()
        self._actualizar_indicador_internet()
        self.root.after(1500, lambda: self._actualizar_estado("Sistema listo", "✅"))
    
    def _confirmar_cierre(self):
        """Confirma el cierre de la aplicación"""
        if self.archivos_pendientes:
            respuesta = CTkMessagebox(
                title="Confirmar Cierre",
                message=f"Tienes {len(self.archivos_pendientes)} archivo(s) pendiente(s) sin subir.\n\n¿Estás seguro de que quieres salir?",
                icon="warning",
                option_1="Salir",
                option_2="Cancelar"
            )
            if respuesta.get() != "Salir":
                return
        # Detener ciclos de after pendientes
        self._app_closing = True
        # Guardar estado actual de viajes conocidos
        self._guardar_estado_app()
        # Limpiar archivos temporales
        for f in self._temp_files:
            try:
                if os.path.exists(f):
                    os.unlink(f)
            except Exception:
                pass
        self.root.destroy()
    
    def _cargar_estado_app(self):
        """Carga el estado previo de viajes conocidos desde archivo local"""
        try:
            if os.path.exists(self.APP_STATE_FILE):
                with open(self.APP_STATE_FILE, 'r') as f:
                    data = json.load(f)
                    viajes_previos = set(data.get('viajes_conocidos', []))
                    if viajes_previos:
                        self._viajes_previos = viajes_previos
                        self._tiene_estado_previo = True
                        return
        except Exception as e:
            print(f"⚠️ No se pudo cargar estado previo: {e}")
        
        self._viajes_previos = set()
        self._tiene_estado_previo = False
    
    def _guardar_estado_app(self):
        """Guarda el estado actual de viajes conocidos en archivo local"""
        try:
            os.makedirs(os.path.dirname(self.APP_STATE_FILE), exist_ok=True)
            data = {
                'viajes_conocidos': list(self._viajes_conocidos),
                'ultima_actualizacion': datetime.now().isoformat()
            }
            with open(self.APP_STATE_FILE, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"⚠️ No se pudo guardar estado: {e}")
    
    # ===== TOOLTIPS =====
    
    def _crear_tooltip(self, widget, texto):
        """Crea un tooltip simple para un widget"""
        tooltip = None
        
        def mostrar(event):
            nonlocal tooltip
            x = event.x_root + 15
            y = event.y_root + 10
            tooltip = tk.Toplevel(widget)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(
                tooltip,
                text=texto,
                background="#333333",
                foreground="white",
                relief="flat",
                borderwidth=1,
                font=("Segoe UI", 9),
                padx=8,
                pady=4
            )
            label.pack()
        
        def ocultar(event):
            nonlocal tooltip
            if tooltip:
                tooltip.destroy()
                tooltip = None
        
        widget.bind("<Enter>", mostrar, add="+")
        widget.bind("<Leave>", ocultar, add="+")
    
    # ===== CONEXIÓN A INTERNET =====
    
    def verificar_conexion_internet(self):
        """Verifica si hay conexión a internet intentando conectar a DNS de Google"""
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(3)
            s.connect(("8.8.8.8", 53))
            s.close()
            return True
        except (socket.timeout, OSError):
            return False
    
    def verificar_conexion_internet_inicio(self):
        """Verifica conexión al iniciar la app y muestra alerta si no hay"""
        self.internet_conectado = self.verificar_conexion_internet()
        self._actualizar_indicador_internet()
        
        if not self.internet_conectado:
            self._mostrar_banner_sin_internet()
    
    def _verificar_actualizacion(self):
        """Verifica si hay una actualización disponible al iniciar"""
        if not self.internet_conectado or not self.firebase.db:
            return
        
        try:
            updater = UpdateManager(self.firebase)
            hay_update, info = updater.hay_actualizacion()
            
            if not hay_update or not info:
                return
            
            version_nueva = info.get('version', '?')
            notas = info.get('notas', '')
            obligatoria = info.get('obligatoria', False)
            url = updater.obtener_url_descarga(info)
            
            # Construir mensaje
            mensaje = f"Hay una nueva versión disponible: v{version_nueva}\n"
            mensaje += f"Versión actual: v{APP_VERSION}\n"
            if notas:
                mensaje += f"\nCambios:\n{notas}\n"
            if obligatoria:
                mensaje += "\n⚠️ Esta actualización es obligatoria."
            
            if url:
                mensaje += "\n\n¿Deseas descargar e instalar la actualización ahora?"
                
                respuesta = CTkMessagebox(
                    title="🔄 Actualización Disponible",
                    message=mensaje,
                    icon="info",
                    option_1="Actualizar",
                    option_2="Después" if not obligatoria else None,
                    fade_in_duration=200
                )
                
                if respuesta.get() == "Actualizar":
                    self._ejecutar_actualizacion(updater, info, url)
                elif obligatoria:
                    # Si es obligatoria y no quiere actualizar, cerrar la app
                    CTkMessagebox(
                        title="Actualización Requerida",
                        message="Esta actualización es obligatoria.\nLa aplicación se cerrará.",
                        icon="warning",
                        option_1="Aceptar"
                    )
                    self.root.destroy()
                    sys.exit(0)
            else:
                # Sin URL, solo notificar
                CTkMessagebox(
                    title="🔄 Actualización Disponible",
                    message=mensaje + "\n\nContacta al administrador para obtener la nueva versión.",
                    icon="info",
                    option_1="Aceptar"
                )
                
        except Exception as e:
            print(f"⚠️ Error verificando actualización: {e}")
    
    def _ejecutar_actualizacion(self, updater, info, url):
        """Descarga y aplica la actualización"""
        import tempfile
        
        # Mostrar progreso
        progreso_win = ctk.CTkToplevel(self.root)
        progreso_win.title("Actualizando BENTOS...")
        progreso_win.geometry("400x150")
        progreso_win.resizable(False, False)
        progreso_win.transient(self.root)
        progreso_win.grab_set()
        
        # Centrar
        progreso_win.update_idletasks()
        x = (progreso_win.winfo_screenwidth() - 400) // 2
        y = (progreso_win.winfo_screenheight() - 150) // 2
        progreso_win.geometry(f"+{x}+{y}")
        
        ctk.CTkLabel(
            progreso_win,
            text="📥 Descargando actualización...",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(20, 10))
        
        barra = ctk.CTkProgressBar(progreso_win, width=350, mode="indeterminate")
        barra.pack(pady=10)
        barra.start()
        
        estado_label = ctk.CTkLabel(progreso_win, text="Preparando descarga...")
        estado_label.pack(pady=5)
        
        progreso_win.update()
        
        # Descargar
        ext = ".exe" if platform.system() == "Windows" else ".app"
        temp_path = os.path.join(tempfile.gettempdir(), f"bentos_update{ext}")
        hash_esperado = updater.obtener_hash_esperado(info)
        
        estado_label.configure(text="Descargando...")
        progreso_win.update()
        
        exito = updater.descargar_actualizacion(url, temp_path, hash_esperado)
        
        if exito:
            estado_label.configure(text="Aplicando actualización...")
            progreso_win.update()
            
            if updater.aplicar_actualizacion(temp_path):
                barra.stop()
                progreso_win.grab_release()
                progreso_win.destroy()
                
                CTkMessagebox(
                    title="Actualización Exitosa",
                    message="La actualización se instalará automáticamente.\nLa aplicación se cerrará y reabrirá sola.",
                    icon="check",
                    option_1="Aceptar"
                )
                self._app_closing = True
                try:
                    self.root.destroy()
                except Exception:
                    pass
                # Forzar cierre inmediato para liberar el .exe
                os._exit(0)
            else:
                barra.stop()
                progreso_win.grab_release()
                progreso_win.destroy()
                CTkMessagebox(
                    title="Error",
                    message="No se pudo aplicar la actualización.\nIntenta de nuevo más tarde.",
                    icon="cancel",
                    option_1="Aceptar"
                )
        else:
            barra.stop()
            progreso_win.grab_release()
            progreso_win.destroy()
            CTkMessagebox(
                title="Error de Descarga",
                message="No se pudo descargar la actualización.\n\n"
                        "Posibles causas:\n"
                        "• Sin conexión a internet\n"
                        "• El repositorio de GitHub no es público\n"
                        "• La URL de descarga es incorrecta\n\n"
                        "Contacta al administrador.",
                icon="cancel",
                option_1="Aceptar"
            )
    
    def monitorear_conexion(self):
        """Monitorea la conexión a internet periódicamente"""
        if self._app_closing:
            return
        estado_anterior = self.internet_conectado
        self.internet_conectado = self.verificar_conexion_internet()
        try:
            self._actualizar_indicador_internet()
        except (tk.TclError, Exception):
            return
        
        # Notificar cambios de estado
        if estado_anterior and not self.internet_conectado:
            self._mostrar_banner_sin_internet()
        elif not estado_anterior and self.internet_conectado:
            self._mostrar_banner_internet_restaurado()
        
        # Verificar cada 30 segundos
        self.root.after(30000, self.monitorear_conexion)
    
    def _actualizar_indicador_internet(self):
        """Actualiza el indicador visual de conexión en el header"""
        if self.internet_conectado:
            self.internet_status_label.configure(
                text="🌐 En línea",
                text_color="#4CAF50"
            )
        else:
            self.internet_status_label.configure(
                text="🌐 Sin conexión",
                text_color="#FF5252"
            )
    
    def _mostrar_banner_sin_internet(self):
        """Muestra un banner de advertencia cuando no hay internet"""
        # Evitar banners duplicados (verificar que exista Y esté visible)
        if hasattr(self, '_banner_internet') and self._banner_internet.winfo_exists() and self._banner_internet.winfo_ismapped():
            return
        # Destruir banner anterior oculto si existe
        if hasattr(self, '_banner_internet') and self._banner_internet.winfo_exists():
            self._banner_internet.destroy()
        
        self._banner_internet = ctk.CTkFrame(
            self.root,
            height=45,
            corner_radius=0,
            fg_color="#FF5252"
        )
        self._banner_internet.pack(fill="x", side="top", before=self.nav_frame)
        self._banner_internet.pack_propagate(False)
        
        banner_content = ctk.CTkFrame(self._banner_internet, fg_color="transparent")
        banner_content.pack(expand=True)
        
        ctk.CTkLabel(
            banner_content,
            text="⚠️  Sin conexión a internet — No es posible buscar ni subir información",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="white"
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            banner_content,
            text="✕",
            width=30,
            height=25,
            corner_radius=5,
            fg_color="transparent",
            hover_color="#E04040",
            text_color="white",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=lambda: self._banner_internet.pack_forget()
        ).pack(side="right", padx=10)
    
    def _mostrar_banner_internet_restaurado(self):
        """Muestra un banner temporal cuando se restaura la conexión"""
        # Quitar banner de sin conexión si existe
        if hasattr(self, '_banner_internet') and self._banner_internet.winfo_exists():
            self._banner_internet.pack_forget()
        
        banner_ok = ctk.CTkFrame(
            self.root,
            height=40,
            corner_radius=0,
            fg_color="#4CAF50"
        )
        banner_ok.pack(fill="x", side="top", before=self.nav_frame)
        banner_ok.pack_propagate(False)
        
        ctk.CTkLabel(
            banner_ok,
            text="✅  Conexión a internet restaurada",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="white"
        ).pack(expand=True)
        
        # Auto-cerrar en 4 segundos con fade
        self.root.after(4000, lambda: self._fade_out_widget(banner_ok))
    
    def _sin_internet_alerta(self, accion="realizar esta acción"):
        """Muestra alerta cuando se intenta una acción sin internet. Retorna True si NO hay internet."""
        if not self.internet_conectado:
            CTkMessagebox(
                title="⚠️ Sin Conexión a Internet",
                message=f"No es posible {accion} sin conexión a internet.\n\n"
                        "Verifica tu conexión e inténtalo de nuevo.",
                icon="warning",
                option_1="Entendido"
            )
            return True
        return False
    
    # ===== ANIMACIONES Y EFECTOS SMOOTH =====
    
    def _fade_in_widget(self, widget, alpha=0.0):
        """Efecto fade-in para widgets toplevel"""
        if alpha < 1.0:
            try:
                widget.attributes('-alpha', alpha)
                self.root.after(20, lambda: self._fade_in_widget(widget, alpha + 0.05))
            except (tk.TclError, Exception):
                pass
        else:
            try:
                widget.attributes('-alpha', 1.0)
            except (tk.TclError, Exception):
                pass
    
    def _fade_out_widget(self, widget):
        """Efecto fade-out y destrucción de un widget frame"""
        try:
            if widget.winfo_exists():
                widget.pack_forget()
                widget.destroy()
        except (tk.TclError, Exception):
            pass
    
    def _efecto_click_boton(self, boton, color_original, color_press="#057A8F"):
        """Efecto visual de click en un botón (flash de color)"""
        try:
            boton.configure(fg_color=color_press)
            self.root.after(120, lambda: boton.configure(fg_color=color_original))
        except (tk.TclError, Exception):
            pass
    
    def _efecto_pulso_widget(self, widget, iteracion=0):
        """Efecto de pulso sutil para llamar atención sobre un widget"""
        colores = ["#05BFDB", "#0492A8", "#05BFDB"]
        if iteracion < len(colores):
            try:
                widget.configure(border_color=colores[iteracion])
                self.root.after(200, lambda: self._efecto_pulso_widget(widget, iteracion + 1))
            except (tk.TclError, Exception):
                pass
    
    def _animar_progreso_suave(self, progress_bar, valor_actual, valor_objetivo, paso=0.02):
        """Anima una barra de progreso suavemente hacia el valor objetivo"""
        if abs(valor_actual - valor_objetivo) > paso:
            nuevo_valor = valor_actual + paso if valor_actual < valor_objetivo else valor_actual - paso
            try:
                progress_bar.set(nuevo_valor)
                self.root.after(15, lambda: self._animar_progreso_suave(
                    progress_bar, nuevo_valor, valor_objetivo, paso))
            except (tk.TclError, Exception):
                pass
        else:
            try:
                progress_bar.set(valor_objetivo)
            except (tk.TclError, Exception):
                pass
    
    def show_section(self, section):
        """Muestra una sección específica con indicador visual activo"""
        self.current_section = section
        
        # Ocultar todas las secciones
        self.upload_section.pack_forget()
        self.search_section.pack_forget()
        
        # Actualizar estilos de botones — activo vs inactivo
        if section == "subir":
            self.btn_subir.configure(fg_color="#05BFDB", hover_color="#057A8F")
            self.btn_buscar.configure(fg_color="#1B3A4B", hover_color="#15303F")
            self.upload_section.pack(fill="both", expand=True)
        else:
            self.btn_subir.configure(fg_color="#1B3A4B", hover_color="#15303F")
            self.btn_buscar.configure(fg_color="#05BFDB", hover_color="#057A8F")
            self.search_section.pack(fill="both", expand=True)
    
    def show_section_animated(self, section):
        """Muestra una sección con efecto de transición suave y feedback visual"""
        if section == self.current_section:
            return  # Ya estamos en esa sección
        
        # Flash de color en botón (efecto click)
        if section == "subir":
            self._efecto_click_boton(self.btn_subir, "#05BFDB", "#03A0B8")
        else:
            self._efecto_click_boton(self.btn_buscar, "#05BFDB", "#1A5F7A")
        
        # Obtener sección actual y nueva
        seccion_actual = self.upload_section if self.current_section == "subir" else self.search_section
        seccion_nueva = self.upload_section if section == "subir" else self.search_section
        
        # Efecto de transición: reducir opacidad simulada con color de fondo
        self._transicion_seccion(seccion_actual, seccion_nueva, section)
    
    def _transicion_seccion(self, seccion_saliente, seccion_entrante, section_name):
        """Realiza una transición suave entre secciones"""
        # Fase 1: Ocultar sección actual
        seccion_saliente.pack_forget()
        
        # Fase 2: Mostrar nueva sección
        self.show_section(section_name)
        
        # Fase 3: Efecto sutil de transición
        try:
            self.content_frame.configure(fg_color="#1D1D32")
            self.root.after(80, lambda: self.content_frame.configure(fg_color=("#EAEFF2", "#1A1A2E")))
        except (tk.TclError, Exception):
            pass
    
    def start_loading_animation(self):
        """Inicia una animación de loading para mejorar la UX"""
        self.loading_animation_running = True
        self.animate_loading()
    
    def stop_loading_animation(self):
        """Detiene la animación de loading"""
        self.loading_animation_running = False
    
    def animate_loading(self):
        """Anima el texto de loading con puntos"""
        if self.loading_animation_running:
            current_text = self.upload_status.cget("text")
            if current_text and not current_text.startswith("✅") and not current_text.startswith("❌"):
                # Contar puntos actuales
                base_text = current_text.rstrip(".")
                dots = current_text.count(".")
                new_dots = (dots + 1) % 4  # Ciclo de 0 a 3 puntos
                new_text = base_text + "." * new_dots
                self.upload_status.configure(text=new_text)
            
            # Continuar animación
            self.root.after(500, self.animate_loading)
    
    def on_especie_change(self, choice):
        """Muestra u oculta los campos de captura min/max según la especie seleccionada"""
        if choice == "Todas":
            # Ocultar campos de captura
            self.captura_min_frame.pack_forget()
            self.captura_max_frame.pack_forget()
        else:
            # Mostrar campos de captura
            self.captura_min_frame.pack(side="left", padx=10, expand=True, fill="x")
            self.captura_max_frame.pack(side="left", padx=(10, 0), expand=True, fill="x")
    
    def set_last_7_days(self):
        """Establece el rango de fechas a los últimos 7 días"""
        today = datetime.now()
        last_week = today - timedelta(days=7)
        
        self.fecha_desde.set_date(last_week)
        self.fecha_hasta.set_date(today)
    
    def set_current_month(self):
        """Establece el rango de fechas al mes actual"""
        today = datetime.now()
        first_day = today.replace(day=1)
        
        # Último día del mes
        if today.month == 12:
            last_day = today.replace(day=31)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
            last_day = next_month - timedelta(days=1)
        
        self.fecha_desde.set_date(first_day)
        self.fecha_hasta.set_date(last_day)
    
    def set_current_year(self):
        """Establece el rango de fechas al año actual"""
        today = datetime.now()
        first_day = today.replace(month=1, day=1)
        last_day = today.replace(month=12, day=31)
        
        self.fecha_desde.set_date(first_day)
        self.fecha_hasta.set_date(last_day)
    
    def search_data(self):
        """Busca datos según los filtros y calcula estadísticas"""
        # Verificar conexión a internet
        if self._sin_internet_alerta("buscar información en la nube"):
            return
        
        # Efecto visual en botón de búsqueda
        self._efecto_click_boton(self.search_btn, "#05BFDB", "#03A0B8")
        self._actualizar_estado("Buscando datos en la nube...", "🔍")
        
        print("\n" + "="*60)
        print("🔍 INICIANDO BÚSQUEDA DE DATOS...")
        print("="*60)
        
        # Limpiar status anterior
        for widget in self.search_status_frame.winfo_children():
            widget.destroy()
        
        # Mostrar indicador de carga
        self.search_status_frame.pack(pady=20, padx=40, fill="x")
        loading_frame = ctk.CTkFrame(self.search_status_frame, fg_color="transparent")
        loading_frame.pack(pady=30)
        
        loading_label = ctk.CTkLabel(
            loading_frame,
            text="🔍 Buscando datos",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#05BFDB"
        )
        loading_label.pack(pady=20)
        
        progress_bar = ctk.CTkProgressBar(
            loading_frame,
            width=400,
            height=10,
            mode="indeterminate"
        )
        progress_bar.pack(pady=10)
        progress_bar.start()
        
        # Actualizar UI para mostrar el loading
        self.root.update()
        
        if self.stats_panel.winfo_ismapped():
            self.stats_panel.pack_forget()
            print("✓ Panel de estadísticas limpiado")
        
        if not self.firebase.db:
            print("❌ Firebase no está conectado")
            loading_frame.destroy()
            CTkMessagebox(
                title="⚠️ Advertencia",
                message="No hay conexión con la base de datos.\n"
                        "Verifica tu conexión a internet e inténtalo de nuevo.",
                icon="warning",
                option_1="OK"
            )
            return
        
        print("✓ Firebase conectado")
        
        # Obtener todos los viajes
        print("📥 Obteniendo viajes de Firebase...")
        loading_label.configure(text="📥 Obteniendo datos de la nube")
        self.root.update()
        
        viajes = self.firebase.listar_viajes(limite=500)
        print(f"✓ Obtenidos {len(viajes) if viajes else 0} viajes")
        
        if not viajes:
            print("❌ No se encontraron viajes en Firebase")
            loading_frame.destroy()
            self.search_status_frame.pack_forget()
            CTkMessagebox(
                title="Sin Datos",
                message="❌ No se encontraron viajes en la base de datos",
                icon="warning",
                option_1="OK"
            )
            return
        
        # Aplicar filtros
        print("\n🔧 Aplicando filtros...")
        loading_label.configure(text="🔧 Aplicando filtros")
        self.root.update()
        
        viajes_filtrados = viajes
        
        # Filtro por fechas
        try:
            fecha_desde = self.fecha_desde.get_date()
            fecha_hasta = self.fecha_hasta.get_date()
            print(f"  - Rango de fechas: {fecha_desde} a {fecha_hasta}")
            
            viajes_filtrados = [
                v for v in viajes_filtrados 
                if v.get('fecha_zarpe') and 
                fecha_desde <= datetime.fromisoformat(v['fecha_zarpe']).date() <= fecha_hasta
            ]
        except Exception as e:
            print(f"  ⚠️ Error en filtro de fechas: {e}")
            pass
        
        # Filtro por nave
        nave_filtro = self.nave_combo.get()
        if nave_filtro != "Todas":
            print(f"  - Nave: {nave_filtro}")
            viajes_filtrados = [v for v in viajes_filtrados if v.get('nave_nombre') == nave_filtro]
        
        # Filtro por capitán (búsqueda parcial case-insensitive)
        capitan_filtro = self.capitan_entry.get().strip()
        if capitan_filtro:
            print(f"  - Capitán contiene: {capitan_filtro}")
            viajes_filtrados = [
                v for v in viajes_filtrados 
                if capitan_filtro.upper() in v.get('capitan', '').upper()
            ]
        
        # Filtro por especie objetivo con rango de captura
        especie_filtro = self.especie_combo.get()
        if especie_filtro != "Todas":
            print(f"  - Especie objetivo: {especie_filtro}")
            captura_min = self.captura_min_entry.get().strip()
            captura_max = self.captura_max_entry.get().strip()
            
            viajes_con_especie = []
            for viaje in viajes_filtrados:
                lances = self.firebase.obtener_lances_viaje(viaje.get('id_viaje'))
                total_especie = 0.0
                
                # Usar CAPTURA TOTAL (lance 0) como fuente de verdad
                lance_ct = next((l for l in lances if l.get('numero_lance') == 0 or l.get('es_captura_total')), None)
                if lance_ct:
                    for especie in lance_ct.get('especies', []):
                        if especie_filtro.lower() in especie.get('nombre', '').lower():
                            total_especie += especie.get('cantidad_ton', 0)
                else:
                    # Fallback: sumar de lances individuales
                    for lance in lances:
                        if lance.get('numero_lance', -1) == 0:
                            continue
                        for especie in lance.get('especies', []):
                            if especie_filtro.lower() in especie.get('nombre', '').lower():
                                total_especie += especie.get('cantidad_ton', 0)
                
                # Si encontró la especie, aplicar filtros de rango
                if total_especie > 0:
                    # Aplicar filtro de mínimo
                    if captura_min:
                        try:
                            min_val = float(captura_min)
                            if total_especie < min_val:
                                continue
                        except ValueError:
                            pass
                    
                    # Aplicar filtro de máximo
                    if captura_max:
                        try:
                            max_val = float(captura_max)
                            if total_especie > max_val:
                                continue
                        except ValueError:
                            pass
                    
                    viajes_con_especie.append(viaje)
            
            viajes_filtrados = viajes_con_especie
            if captura_min and captura_max:
                print(f"    - Captura de {especie_filtro} entre {captura_min} y {captura_max} TON")
            elif captura_min:
                print(f"    - Captura de {especie_filtro} >= {captura_min} TON")
            elif captura_max:
                print(f"    - Captura de {especie_filtro} <= {captura_max} TON")
        
        print(f"\n✓ Viajes después de filtros: {len(viajes_filtrados)}")
        
        # Limpiar indicador de carga
        loading_frame.destroy()
        self.search_status_frame.pack_forget()
        
        # Mostrar resultados
        if not viajes_filtrados:
            CTkMessagebox(
                title="Sin Resultados",
                message="❌ No se encontraron resultados con los filtros seleccionados",
                icon="warning",
                option_1="OK"
            )
            self.results_count_label.configure(text="")
        else:
            # Actualizar contador
            self.results_count_label.configure(
                text=f"📊 {len(viajes_filtrados)} viaje(s) encontrado(s)"
            )
            
            # MOSTRAR RESULTADOS EN VENTANA EMERGENTE
            print(f"✓ Abriendo ventana de resultados con {len(viajes_filtrados)} viajes...")
            self.mostrar_ventana_resultados(viajes_filtrados)
            self._actualizar_estado(f"Búsqueda completada — {len(viajes_filtrados)} bitácora(s) encontrada(s)", "✅")
            
            print(f"✓ Ventana de resultados creada")
            print("="*60)
    
    def clear_filters(self):
        """Limpia todos los filtros y resultados"""
        self.nave_combo.set("Todas")
        self.capitan_entry.delete(0, 'end')
        self.especie_combo.set("Todas")
        self.captura_min_entry.delete(0, 'end')
        self.captura_max_entry.delete(0, 'end')
        
        # Ocultar campos de captura
        self.captura_min_frame.pack_forget()
        self.captura_max_frame.pack_forget()
        
        # Restablecer fechas al mes actual
        self.set_current_month()
        
        # Limpiar status
        self.search_status_frame.pack_forget()
        self.results_count_label.configure(text="")
        
        if self.stats_panel.winfo_ismapped():
            self.stats_panel.pack_forget()
    
    def mostrar_ventana_resultados(self, viajes):
        """Muestra los resultados en una ventana emergente con dos modos de visualización"""
        # Crear ventana emergente
        self.resultado_window = ctk.CTkToplevel(self.root)
        self.resultado_window.title(f"📊 Resultados de Búsqueda - {len(viajes)} Bitácora(s)")
        
        # Fade-in al abrir
        self.resultado_window.attributes('-alpha', 0.0)
        
        # Adaptar tamaño a la pantalla
        screen_w = self.resultado_window.winfo_screenwidth()
        screen_h = self.resultado_window.winfo_screenheight()
        win_w = min(1200, screen_w - 40)
        win_h = min(750, screen_h - 100)
        self.resultado_window.geometry(f"{win_w}x{win_h}")
        
        # Hacer que aparezca siempre encima
        self.resultado_window.transient(self.root)
        self.resultado_window.grab_set()
        self.resultado_window.focus_force()
        
        # Liberar grab al cerrar ventana
        def _on_close_results():
            try:
                self.resultado_window.grab_release()
            except Exception:
                pass
            self.resultado_window.destroy()
        self.resultado_window.protocol("WM_DELETE_WINDOW", _on_close_results)
        
        # Centrar la ventana
        self.resultado_window.update_idletasks()
        x = max(10, (screen_w - win_w) // 2)
        y = max(10, (screen_h - win_h) // 2 - 30)
        self.resultado_window.geometry(f"{win_w}x{win_h}+{x}+{y}")
        
        # Iniciar fade-in
        self._fade_in_widget(self.resultado_window)
        
        # Header con contador
        header = ctk.CTkFrame(self.resultado_window, fg_color="#05BFDB", height=70)
        header.pack(fill="x", padx=0, pady=0)
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text=f"📋 {len(viajes)} Bitácora(s) Encontrada(s)",
            font=ctk.CTkFont(size=26, weight="bold"),
            text_color="white"
        ).pack(side="left", padx=30, pady=20)
        
        # Panel de control - Selector de vista
        control_panel = ctk.CTkFrame(self.resultado_window, fg_color="#F0F0F0", height=80)
        control_panel.pack(fill="x", padx=20, pady=10)
        control_panel.pack_propagate(False)
        
        ctk.CTkLabel(
            control_panel,
            text="Modo de Visualización:",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#0A4D68"
        ).pack(side="left", padx=20)
        
        # Botones de modo
        self.btn_resumen = ctk.CTkButton(
            control_panel,
            text="📊 RESUMEN TOTAL DE CAPTURAS",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=50,
            width=300,
            fg_color="#00A86B",
            hover_color="#008B5A",
            command=lambda: self.cambiar_vista_resultados("resumen", viajes)
        )
        self.btn_resumen.pack(side="left", padx=10)
        
        self.btn_individual = ctk.CTkButton(
            control_panel,
            text="📑 BITÁCORAS INDIVIDUALES",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=50,
            width=300,
            fg_color="#2196F3",
            hover_color="#1976D2",
            command=lambda: self.cambiar_vista_resultados("individual", viajes)
        )
        self.btn_individual.pack(side="left", padx=10)
        
        # Contenedor para las vistas (se intercambian)
        self.contenedor_vistas = ctk.CTkFrame(self.resultado_window, fg_color="transparent")
        self.contenedor_vistas.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Guardar viajes para cambios de vista
        self.viajes_actuales = viajes
        
        # Mostrar vista de resumen por defecto
        self.cambiar_vista_resultados("resumen", viajes)
    
    def cambiar_vista_resultados(self, modo, viajes):
        """Cambia entre vista de resumen y vista individual"""
        print(f"\n🔄 Cambiando vista a: {modo}")
        print(f"📊 Viajes a mostrar: {len(viajes)}")
        
        # Limpiar contenedor
        for widget in self.contenedor_vistas.winfo_children():
            widget.destroy()
        
        # Actualizar colores de botones
        if modo == "resumen":
            self.btn_resumen.configure(fg_color="#00A86B", text_color="white")
            self.btn_individual.configure(fg_color="#CCCCCC", text_color="#666666")
            print("📈 Mostrando resumen total...")
            self.mostrar_resumen_total(viajes)
        else:
            self.btn_resumen.configure(fg_color="#CCCCCC", text_color="#666666")
            self.btn_individual.configure(fg_color="#2196F3", text_color="white")
            # Reiniciar variables de paginación al cambiar a vista individual
            self.pagina_actual = 1
            self.viajes_completos = viajes
            self.viajes_filtrados = viajes
            print("📑 Mostrando bitácoras individuales...")
            self.mostrar_bitacoras_individuales(viajes)
        
        print(f"✅ Vista de {modo} cargada correctamente")
    
    def mostrar_resumen_total(self, viajes):
        """Muestra el resumen total de todas las capturas agregadas"""
        # Calcular totales por especie SEPARANDO retenidas, descartadas e incidentales
        especies_retenidas = {}
        especies_descartadas = {}
        especies_incidentales = {}  # {nombre: unidades}
        total_lances = 0
        total_retenidas_ton = 0
        total_descartadas_ton = 0
        total_incidentales_unidades = 0
        
        for viaje in viajes:
            viaje_id = viaje.get('id_viaje')
            lances = self.firebase.obtener_lances_viaje(viaje_id)
            total_lances += viaje.get('total_lances_declarados', 0)
            
            # FILTRAR SOLO EL LANCE CAPTURA TOTAL (lance 0)
            lance_captura_total = None
            for lance in lances:
                if lance.get('es_captura_total', False):
                    lance_captura_total = lance
                    break
            
            # Si no hay lance de captura total, usar el primero como fallback
            if not lance_captura_total and lances:
                lance_captura_total = lances[0]
            
            # Procesar SOLO las especies del lance CAPTURA TOTAL
            if lance_captura_total:
                for especie_data in lance_captura_total.get('especies', []):
                    nombre = especie_data.get('nombre')
                    cantidad_ton = especie_data.get('cantidad_ton', 0)
                    cantidad_unidades = especie_data.get('cantidad_unidades', 0)
                    tipo_captura = especie_data.get('tipo_captura', 'retenida')
                    
                    # Separar por tipo de captura
                    if tipo_captura == 'retenida':
                        if cantidad_ton > 0:
                            if nombre not in especies_retenidas:
                                especies_retenidas[nombre] = 0
                            especies_retenidas[nombre] += cantidad_ton
                            total_retenidas_ton += cantidad_ton
                    elif tipo_captura == 'descartada':
                        if cantidad_ton > 0 or cantidad_unidades > 0:
                            if nombre not in especies_descartadas:
                                especies_descartadas[nombre] = {'ton': 0, 'unidades': 0}
                            especies_descartadas[nombre]['ton'] += cantidad_ton
                            especies_descartadas[nombre]['unidades'] += cantidad_unidades
                            total_descartadas_ton += cantidad_ton
                    elif tipo_captura == 'incidental':
                        if cantidad_unidades > 0:
                            if nombre not in especies_incidentales:
                                especies_incidentales[nombre] = 0
                            especies_incidentales[nombre] += cantidad_unidades
                            total_incidentales_unidades += cantidad_unidades
        
        # Calcular total general
        total_general = total_retenidas_ton + total_descartadas_ton
        
        # Calcular porcentajes
        porc_retenidas = (total_retenidas_ton / total_general * 100) if total_general > 0 else 0
        porc_descartadas = (total_descartadas_ton / total_general * 100) if total_general > 0 else 0
        
        # Contar especies únicas
        total_especies = len(especies_retenidas) + len(especies_descartadas) + len(especies_incidentales)
        
        # Crear frame principal
        resumen_frame = ctk.CTkFrame(self.contenedor_vistas, fg_color="white", corner_radius=15)
        resumen_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título con botón de gráfico
        titulo_frame = ctk.CTkFrame(resumen_frame, fg_color="#0A4D68", corner_radius=10)
        titulo_frame.pack(fill="x", padx=15, pady=15)
        
        titulo_content = ctk.CTkFrame(titulo_frame, fg_color="transparent")
        titulo_content.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(
            titulo_content,
            text="📊 RESUMEN TOTAL DE CAPTURAS",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="white"
        ).pack(side="left")
        
        # Frame para botones
        botones_resumen_frame = ctk.CTkFrame(titulo_content, fg_color="transparent")
        botones_resumen_frame.pack(side="right")
        
        # Botón para exportar a Excel
        ctk.CTkButton(
            botones_resumen_frame,
            text="📥 Exportar Excel",
            command=lambda: self.exportar_resumen_excel(
                viajes, especies_retenidas, especies_descartadas,
                total_retenidas_ton, total_descartadas_ton, total_general,
                total_lances, total_especies,
                porc_retenidas, porc_descartadas,
                especies_incidentales, total_incidentales_unidades
            ),
            fg_color="#2E8B9E",
            hover_color="#1A5F7A",
            height=35,
            width=160,
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="right", padx=(5, 0))
        
        # Botón para generar mapa de calor general
        ctk.CTkButton(
            botones_resumen_frame,
            text="🗺️ Mapa de Calor",
            command=lambda: self.generar_mapa_calor(viajes),
            fg_color="#FF9800",
            hover_color="#F57C00",
            height=35,
            width=160,
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="right", padx=(5, 0))
        
        # Botón para generar gráfico del resumen total
        ctk.CTkButton(
            botones_resumen_frame,
            text="📊 Gráfico",
            command=lambda: self.generar_grafico_resumen_total(viajes),
            fg_color="#4CAF50",
            hover_color="#388E3C",
            height=35,
            width=130,
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="right")
        
        # Info general con mejor diseño
        info_frame = ctk.CTkFrame(resumen_frame, fg_color="#E8F5E9", corner_radius=10, border_width=2, border_color="#4CAF50")
        info_frame.pack(fill="x", padx=15, pady=(5, 3))
        
        info_grid = ctk.CTkFrame(info_frame, fg_color="transparent")
        info_grid.pack(pady=8, padx=15)
        
        stats = [
            ("📋 Bitácoras", len(viajes), "#2196F3"),
            ("🎣 Lances Totales", total_lances, "#FF9800"),
            ("🐟 Especies Diferentes", total_especies, "#4CAF50"),
            ("⚖️ Captura Total", f"{total_general:.2f} TON", "#9C27B0")
        ]
        
        for i, (label, valor, color) in enumerate(stats):
            box = ctk.CTkFrame(info_grid, fg_color="white", corner_radius=8, 
                             border_width=2, border_color=color, width=150, height=60)
            box.pack(side="left", padx=6)
            box.pack_propagate(False)
            
            ctk.CTkLabel(
                box,
                text=label,
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color="#666666"
            ).pack(pady=(6, 2))
            
            ctk.CTkLabel(
                box,
                text=str(valor),
                font=ctk.CTkFont(size=15, weight="bold"),
                text_color=color
            ).pack(pady=(0, 5))
        
        # DISTRIBUCIÓN DE CAPTURAS (compacto)
        distribucion_frame = ctk.CTkFrame(resumen_frame, fg_color="transparent")
        distribucion_frame.pack(fill="x", padx=15, pady=(3, 5))
        
        # Frame para retenidas
        retenidas_box = ctk.CTkFrame(distribucion_frame, fg_color="#E8F5E9", corner_radius=8, 
                                     border_width=2, border_color="#4CAF50")
        retenidas_box.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        ret_content = ctk.CTkFrame(retenidas_box, fg_color="transparent")
        ret_content.pack(pady=6, padx=10)
        
        ctk.CTkLabel(
            ret_content,
            text="🎯 RETENIDA:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#2E7D32"
        ).pack(side="left", padx=(0, 8))
        
        ctk.CTkLabel(
            ret_content,
            text=f"{total_retenidas_ton:.2f} TON ({porc_retenidas:.1f}%)",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1B5E20"
        ).pack(side="left")
        
        # Frame para descartadas
        descartadas_box = ctk.CTkFrame(distribucion_frame, fg_color="#FFEBEE", corner_radius=8,
                                       border_width=2, border_color="#EF5350")
        descartadas_box.pack(side="left", fill="both", expand=True, padx=(5, 0))
        
        desc_content = ctk.CTkFrame(descartadas_box, fg_color="transparent")
        desc_content.pack(pady=6, padx=10)
        
        ctk.CTkLabel(
            desc_content,
            text="🗑️ DESCARTE:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#C62828"
        ).pack(side="left", padx=(0, 8))
        
        ctk.CTkLabel(
            desc_content,
            text=f"{total_descartadas_ton:.2f} TON ({porc_descartadas:.1f}%)",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#B71C1C"
        ).pack(side="left")
        
        # Scroll para las tablas con mejor diseño
        tabla_scroll = ctk.CTkScrollableFrame(
            resumen_frame,
            fg_color="white",
            corner_radius=10,
            border_width=2,
            border_color="#E0E0E0"
        )
        tabla_scroll.pack(fill="both", expand=True, padx=15, pady=10)
        
        # ============================================================
        # SECCIÓN 1: ESPECIES RETENIDAS
        # ============================================================
        if especies_retenidas:
            # Header de sección retenidas
            header_retenidas = ctk.CTkFrame(tabla_scroll, fg_color="#4CAF50", corner_radius=10)
            header_retenidas.pack(fill="x", pady=(10, 5))
            
            ctk.CTkLabel(
                header_retenidas,
                text="🎯 ESPECIES RETENIDAS",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color="white",
                anchor="w"
            ).pack(side="left", padx=20, pady=12)
            
            ctk.CTkLabel(
                header_retenidas,
                text=f"Total: {total_retenidas_ton:.2f} TON ({porc_retenidas:.1f}%)",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="white",
                anchor="e"
            ).pack(side="right", padx=20, pady=12)
            
            # Header de tabla retenidas
            header_tabla_ret = ctk.CTkFrame(tabla_scroll, fg_color="#2E7D32", corner_radius=8)
            header_tabla_ret.pack(fill="x", pady=(10, 5))
            
            headers_ret = [
                ("#", 50),
                ("ESPECIE", 400),
                ("TOTAL CAPTURADO", 150),
                ("% DEL RETENIDO", 120)
            ]
            
            for texto, ancho in headers_ret:
                ctk.CTkLabel(
                    header_tabla_ret,
                    text=texto,
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color="white",
                    width=ancho,
                    anchor="w" if texto == "ESPECIE" else "e"
                ).pack(side="left", padx=10, pady=12)
            
            # Ordenar especies retenidas por cantidad
            especies_retenidas_ord = sorted(especies_retenidas.items(), key=lambda x: x[1], reverse=True)
            
            # Mostrar especies retenidas
            for i, (especie, cantidad) in enumerate(especies_retenidas_ord, 1):
                porcentaje = (cantidad / total_retenidas_ton * 100) if total_retenidas_ton > 0 else 0
                fila_bg = "#E8F5E9" if i % 2 == 0 else "#F1F8E9"
                
                fila = ctk.CTkFrame(tabla_scroll, fg_color=fila_bg, corner_radius=5)
                fila.pack(fill="x", pady=2)
                
                # Número
                ctk.CTkLabel(
                    fila,
                    text=str(i),
                    font=ctk.CTkFont(size=11),
                    text_color="#666666",
                    width=50
                ).pack(side="left", padx=10, pady=8)
                
                # Especie
                ctk.CTkLabel(
                    fila,
                    text=especie,
                    font=ctk.CTkFont(size=11, weight="bold"),
                    text_color="#1B5E20",
                    width=400,
                    anchor="w"
                ).pack(side="left", padx=10, pady=8)
                
                # Cantidad
                ctk.CTkLabel(
                    fila,
                    text=f"{cantidad:.3f} TON",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    text_color="#2E7D32",
                    width=150,
                    anchor="e"
                ).pack(side="left", padx=10, pady=8)
                
                # Porcentaje
                ctk.CTkLabel(
                    fila,
                    text=f"{porcentaje:.1f}%",
                    font=ctk.CTkFont(size=11),
                    text_color="#388E3C",
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=10, pady=8)
        
        # ============================================================
        # SECCIÓN 2: ESPECIES DESCARTADAS
        # ============================================================
        if especies_descartadas:
            # Separador
            ctk.CTkFrame(tabla_scroll, fg_color="#E0E0E0", height=2).pack(fill="x", pady=20)
            
            # Header de sección descartadas
            header_descartadas = ctk.CTkFrame(tabla_scroll, fg_color="#EF5350", corner_radius=10)
            header_descartadas.pack(fill="x", pady=(10, 5))
            
            ctk.CTkLabel(
                header_descartadas,
                text="🗑️ ESPECIES DESCARTADAS",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color="white",
                anchor="w"
            ).pack(side="left", padx=20, pady=12)
            
            ctk.CTkLabel(
                header_descartadas,
                text=f"Total: {total_descartadas_ton:.2f} TON ({porc_descartadas:.1f}%)",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="white",
                anchor="e"
            ).pack(side="right", padx=20, pady=12)
            
            # Header de tabla descartadas
            header_tabla_desc = ctk.CTkFrame(tabla_scroll, fg_color="#C62828", corner_radius=8)
            header_tabla_desc.pack(fill="x", pady=(10, 5))
            
            headers_desc = [
                ("#", 50),
                ("ESPECIE", 300),
                ("TON", 120),
                ("UNIDADES", 120),
                ("% DEL DESCARTE", 120)
            ]
            
            for texto, ancho in headers_desc:
                ctk.CTkLabel(
                    header_tabla_desc,
                    text=texto,
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color="white",
                    width=ancho,
                    anchor="w" if texto == "ESPECIE" else "e"
                ).pack(side="left", padx=10, pady=12)
            
            # Ordenar especies descartadas por cantidad (ton primero, luego unidades)
            especies_descartadas_ord = sorted(especies_descartadas.items(), 
                                            key=lambda x: (x[1]['ton'], x[1]['unidades']), 
                                            reverse=True)
            
            # Mostrar especies descartadas
            for i, (especie, datos) in enumerate(especies_descartadas_ord, 1):
                cantidad_ton = datos['ton']
                cantidad_unidades = datos['unidades']
                porcentaje = (cantidad_ton / total_descartadas_ton * 100) if total_descartadas_ton > 0 else 0
                fila_bg = "#FFEBEE" if i % 2 == 0 else "#FFCDD2"
                
                fila = ctk.CTkFrame(tabla_scroll, fg_color=fila_bg, corner_radius=5)
                fila.pack(fill="x", pady=2)
                
                # Número
                ctk.CTkLabel(
                    fila,
                    text=str(i),
                    font=ctk.CTkFont(size=11),
                    text_color="#666666",
                    width=50
                ).pack(side="left", padx=10, pady=8)
                
                # Especie
                ctk.CTkLabel(
                    fila,
                    text=especie,
                    font=ctk.CTkFont(size=11, weight="bold"),
                    text_color="#B71C1C",
                    width=300,
                    anchor="w"
                ).pack(side="left", padx=10, pady=8)
                
                # TON
                ctk.CTkLabel(
                    fila,
                    text=f"{cantidad_ton:.3f}" if cantidad_ton > 0 else "-",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    text_color="#C62828",
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=10, pady=8)
                
                # UNIDADES
                ctk.CTkLabel(
                    fila,
                    text=f"{int(cantidad_unidades):,}" if cantidad_unidades > 0 else "-",
                    font=ctk.CTkFont(size=11),
                    text_color="#D32F2F",
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=10, pady=8)
                
                # Porcentaje
                ctk.CTkLabel(
                    fila,
                    text=f"{porcentaje:.1f}%",
                    font=ctk.CTkFont(size=11),
                    text_color="#E57373",
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=10, pady=8)
    
        # ============================================================
        # SECCIÓN 3: ESPECIES INCIDENTALES
        # ============================================================
        if especies_incidentales:
            # Separador
            ctk.CTkFrame(tabla_scroll, fg_color="#E0E0E0", height=2).pack(fill="x", pady=20)
            
            # Header de sección incidentales
            header_incidentales = ctk.CTkFrame(tabla_scroll, fg_color="#9C27B0", corner_radius=10)
            header_incidentales.pack(fill="x", pady=(10, 5))
            
            ctk.CTkLabel(
                header_incidentales,
                text="🦭 ESPECIES INCIDENTALES (Fauna Acompañante)",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color="white",
                anchor="w"
            ).pack(side="left", padx=20, pady=12)
            
            ctk.CTkLabel(
                header_incidentales,
                text=f"Total: {total_incidentales_unidades} individuo(s)",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="white",
                anchor="e"
            ).pack(side="right", padx=20, pady=12)
            
            # Header de tabla incidentales
            header_tabla_inc = ctk.CTkFrame(tabla_scroll, fg_color="#7B1FA2", corner_radius=8)
            header_tabla_inc.pack(fill="x", pady=(10, 5))
            
            headers_inc = [
                ("#", 50),
                ("ESPECIE", 400),
                ("INDIVIDUOS", 150),
                ("% DEL TOTAL", 120)
            ]
            
            for texto, ancho in headers_inc:
                ctk.CTkLabel(
                    header_tabla_inc,
                    text=texto,
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color="white",
                    width=ancho,
                    anchor="w" if texto == "ESPECIE" else "e"
                ).pack(side="left", padx=10, pady=12)
            
            # Ordenar incidentales por cantidad
            especies_incidentales_ord = sorted(especies_incidentales.items(), key=lambda x: x[1], reverse=True)
            
            for i, (especie, cantidad) in enumerate(especies_incidentales_ord, 1):
                porcentaje = (cantidad / total_incidentales_unidades * 100) if total_incidentales_unidades > 0 else 0
                fila_bg = "#F3E5F5" if i % 2 == 0 else "#EDE7F6"
                
                fila = ctk.CTkFrame(tabla_scroll, fg_color=fila_bg, corner_radius=5)
                fila.pack(fill="x", pady=2)
                
                ctk.CTkLabel(
                    fila, text=str(i),
                    font=ctk.CTkFont(size=11), text_color="#666666", width=50
                ).pack(side="left", padx=10, pady=8)
                
                ctk.CTkLabel(
                    fila, text=especie,
                    font=ctk.CTkFont(size=11, weight="bold"), text_color="#6A1B9A",
                    width=400, anchor="w"
                ).pack(side="left", padx=10, pady=8)
                
                ctk.CTkLabel(
                    fila, text=f"{int(cantidad)} ind.",
                    font=ctk.CTkFont(size=11, weight="bold"), text_color="#7B1FA2",
                    width=150, anchor="e"
                ).pack(side="left", padx=10, pady=8)
                
                ctk.CTkLabel(
                    fila, text=f"{porcentaje:.1f}%",
                    font=ctk.CTkFont(size=11), text_color="#9C27B0",
                    width=120, anchor="e"
                ).pack(side="left", padx=10, pady=8)
    
    def _crear_fila_especie(self, parent, numero, especie, cantidad, porcentaje):
        """Crea una fila en la tabla de especies"""
        fila_bg = "#FFFFFF" if numero % 2 == 0 else "#F0F0F0"
        fila = ctk.CTkFrame(parent, fg_color=fila_bg, corner_radius=5)
        fila.pack(fill="x", pady=2)
        
        # Número
        ctk.CTkLabel(
            fila,
            text=str(numero),
            font=ctk.CTkFont(size=11),
            text_color="#666666",
            width=50
        ).pack(side="left", padx=10, pady=8)
        
        # Especie
        ctk.CTkLabel(
            fila,
            text=especie,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#000000",
            width=400,
            anchor="w"
        ).pack(side="left", padx=10, pady=8)
        
        # Cantidad
        ctk.CTkLabel(
            fila,
            text=f"{cantidad:.3f} TON",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#0277BD",
            width=150,
            anchor="e"
        ).pack(side="left", padx=10, pady=8)
        
        # Porcentaje
        ctk.CTkLabel(
            fila,
            text=f"{porcentaje:.1f}%",
            font=ctk.CTkFont(size=11),
            text_color="#00A86B" if porcentaje >= 10 else "#666666",
            width=120,
            anchor="e"
        ).pack(side="left", padx=10, pady=8)
    
    def toggle_otras_especies(self, otras_especies_ord, total_general, especies_objetivo_ord):
        """Muestra u oculta el panel de otras especies"""
        if self.contenedor_otras.winfo_ismapped():
            # Ocultar
            self.contenedor_otras.pack_forget()
            self.btn_toggle_otras.configure(text=f"▶ MOSTRAR OTRAS ESPECIES ({len(otras_especies_ord)})")
        else:
            # Mostrar
            # Limpiar contenedor
            for widget in self.contenedor_otras.winfo_children():
                widget.destroy()
            
            ctk.CTkLabel(
                self.contenedor_otras,
                text="📦 OTRAS ESPECIES:",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#666666",
                anchor="w"
            ).pack(anchor="w", pady=(10, 5), padx=10)
            
            # La numeración continúa desde donde terminaron las especies objetivo
            num_inicial = len(especies_objetivo_ord) + 1
            
            # Mostrar otras especies
            for i, (especie, cantidad) in enumerate(otras_especies_ord, num_inicial):
                porcentaje = (cantidad / total_general * 100) if total_general > 0 else 0
                self._crear_fila_especie(self.contenedor_otras, i, especie, cantidad, porcentaje)
            
            self.contenedor_otras.pack(fill="x", pady=10)
            self.btn_toggle_otras.configure(text=f"▼ OCULTAR OTRAS ESPECIES ({len(otras_especies_ord)})")
    
    def mostrar_bitacoras_individuales(self, viajes):
        """Muestra las bitácoras en formato de tarjetas individuales con paginación"""
        print(f"\n{'='*60}")
        print(f"🎴 INICIANDO CREACIÓN DE TARJETAS INDIVIDUALES")
        print(f"{'='*60}")
        print(f"Total viajes recibidos: {len(viajes)}")
        
        # Inicializar variables de paginación si no existen
        if not hasattr(self, 'pagina_actual'):
            self.pagina_actual = 1
        if not hasattr(self, 'bitacoras_por_pagina'):
            self.bitacoras_por_pagina = 5
        if not hasattr(self, 'viajes_completos'):
            self.viajes_completos = viajes
        if not hasattr(self, 'viajes_filtrados'):
            self.viajes_filtrados = viajes
        
        print(f"Página actual: {self.pagina_actual}")
        print(f"Viajes filtrados: {len(self.viajes_filtrados)}")
        
        # Frame contenedor principal
        container = ctk.CTkFrame(self.contenedor_vistas, fg_color="#E8F4F8")
        container.pack(fill="both", expand=True)
        print("✅ Contenedor creado")
        
        # BARRA DE BÚSQUEDA
        search_frame = ctk.CTkFrame(container, fg_color="#FFFFFF", height=70, corner_radius=10)
        search_frame.pack(fill="x", padx=10, pady=10)
        search_frame.pack_propagate(False)
        
        ctk.CTkLabel(
            search_frame,
            text="🔍 Buscar:",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#000000"
        ).pack(side="left", padx=(20, 10), pady=20)
        
        self.entry_buscar = ctk.CTkEntry(
            search_frame,
            placeholder_text="ID de viaje, nave, capitán...",
            width=400,
            height=35,
            font=ctk.CTkFont(size=12),
            border_color="#05BFDB",
            border_width=2
        )
        self.entry_buscar.pack(side="left", padx=10, pady=20)
        
        # Bind para buscar al presionar Enter
        self.entry_buscar.bind("<Return>", lambda e: self.filtrar_bitacoras())
        
        ctk.CTkButton(
            search_frame,
            text="Buscar",
            width=100,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color="#05BFDB",
            hover_color="#057A8F",
            command=self.filtrar_bitacoras
        ).pack(side="left", padx=5, pady=20)
        
        ctk.CTkButton(
            search_frame,
            text="Limpiar",
            width=100,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color="#666666",
            hover_color="#888888",
            command=self.limpiar_busqueda
        ).pack(side="left", padx=5, pady=20)
        
        # ESTADÍSTICAS
        stats_text = f"Total: {len(self.viajes_filtrados)} bitácoras"
        if len(self.viajes_filtrados) != len(self.viajes_completos):
            stats_text += f" (de {len(self.viajes_completos)} totales)"
        
        ctk.CTkLabel(
            search_frame,
            text=stats_text,
            font=ctk.CTkFont(size=12),
            text_color="#666666"
        ).pack(side="left", padx=20, pady=20)
        
        # Calcular paginación
        total_paginas = max(1, (len(self.viajes_filtrados) + self.bitacoras_por_pagina - 1) // self.bitacoras_por_pagina)
        self.pagina_actual = min(self.pagina_actual, total_paginas)
        
        # SCROLL PARA LAS TARJETAS
        scroll_frame = ctk.CTkScrollableFrame(container, fg_color="#E8F4F8")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Calcular índices para la página actual
        inicio = (self.pagina_actual - 1) * self.bitacoras_por_pagina
        fin = min(inicio + self.bitacoras_por_pagina, len(self.viajes_filtrados))
        viajes_pagina = self.viajes_filtrados[inicio:fin]
        
        # Crear tarjetas para la página actual
        if viajes_pagina:
            print(f"\n📦 Creando {len(viajes_pagina)} tarjetas...")
            for i, viaje in enumerate(viajes_pagina, inicio + 1):
                print(f"  Tarjeta #{i}: {viaje.get('id_viaje', 'N/A')}")
                try:
                    self.crear_tarjeta_resultado(scroll_frame, viaje, i)
                    print(f"    ✅ Tarjeta #{i} creada")
                except Exception as e:
                    print(f"    ❌ ERROR en tarjeta #{i}: {e}")
                    import traceback
                    traceback.print_exc()
            print(f"✅ Todas las tarjetas creadas")
        else:
            ctk.CTkLabel(
                scroll_frame,
                text="No se encontraron bitácoras con los criterios de búsqueda",
                font=ctk.CTkFont(size=14),
                text_color="#666666"
            ).pack(pady=50)
        
        # PAGINACIÓN INFERIOR
        paginacion_bottom = ctk.CTkFrame(container, fg_color="#FFFFFF", height=50, corner_radius=10)
        paginacion_bottom.pack(fill="x", padx=10, pady=(10, 10))
        paginacion_bottom.pack_propagate(False)
        
        self._crear_controles_paginacion(paginacion_bottom, total_paginas)
    
    def _crear_controles_paginacion(self, parent, total_paginas):
        """Crea los controles de navegación de paginación"""
        # Botón Anterior
        btn_anterior = ctk.CTkButton(
            parent,
            text="← Anterior",
            width=120,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color="#05BFDB" if self.pagina_actual > 1 else "#CCCCCC",
            hover_color="#0492A8" if self.pagina_actual > 1 else "#CCCCCC",
            state="normal" if self.pagina_actual > 1 else "disabled",
            command=lambda: self.cambiar_pagina(self.pagina_actual - 1)
        )
        btn_anterior.pack(side="left", padx=20, pady=7)
        
        # Info de página
        ctk.CTkLabel(
            parent,
            text=f"Página {self.pagina_actual} de {total_paginas}",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#000000"
        ).pack(side="left", expand=True)
        
        # Botón Siguiente
        btn_siguiente = ctk.CTkButton(
            parent,
            text="Siguiente →",
            width=120,
            height=35,
            font=ctk.CTkFont(size=12),
            fg_color="#05BFDB" if self.pagina_actual < total_paginas else "#CCCCCC",
            hover_color="#0492A8" if self.pagina_actual < total_paginas else "#CCCCCC",
            state="normal" if self.pagina_actual < total_paginas else "disabled",
            command=lambda: self.cambiar_pagina(self.pagina_actual + 1)
        )
        btn_siguiente.pack(side="right", padx=20, pady=7)
    
    def cambiar_pagina(self, nueva_pagina):
        """Cambia a una nueva página de resultados"""
        self.pagina_actual = nueva_pagina
        self.actualizar_vista_individual()
    
    def filtrar_bitacoras(self):
        """Filtra las bitácoras según el texto de búsqueda"""
        texto_busqueda = self.entry_buscar.get().lower().strip()
        
        if not texto_busqueda:
            self.viajes_filtrados = self.viajes_completos
        else:
            self.viajes_filtrados = []
            for viaje in self.viajes_completos:
                # Buscar en ID, nave, capitán, puerto
                if (texto_busqueda in str(viaje.get('id_viaje', '')).lower() or
                    texto_busqueda in viaje.get('nave_nombre', '').lower() or
                    texto_busqueda in viaje.get('capitan', '').lower() or
                    texto_busqueda in viaje.get('puerto_zarpe', '').lower()):
                    self.viajes_filtrados.append(viaje)
        
        # Volver a página 1 después de filtrar
        self.pagina_actual = 1
        self.actualizar_vista_individual()
    
    def limpiar_busqueda(self):
        """Limpia el campo de búsqueda y restaura todos los resultados"""
        self.entry_buscar.delete(0, 'end')
        self.viajes_filtrados = self.viajes_completos
        self.pagina_actual = 1
        self.actualizar_vista_individual()
    
    def actualizar_vista_individual(self):
        """Reconstruye la vista individual con los filtros y paginación actuales"""
        # Limpiar contenedor de vistas
        for widget in self.contenedor_vistas.winfo_children():
            widget.destroy()
        
        # Recrear la vista
        self.mostrar_bitacoras_individuales(self.viajes_completos)
    
    def crear_tarjeta_resultado(self, parent, viaje, numero):
        """Crea una tarjeta de resultado en la ventana emergente"""
        # Obtener lances y calcular totales
        viaje_id = viaje.get('id_viaje', 'N/A')
        lances = self.firebase.obtener_lances_viaje(viaje_id)
        
        if not lances:
            print(f"⚠️  Viaje {viaje_id} no tiene lances - saltando tarjeta")
            return
        
        # FILTRAR SOLO EL LANCE CAPTURA TOTAL (lance 0)
        lance_captura_total = None
        for lance in lances:
            if lance.get('es_captura_total', False):
                lance_captura_total = lance
                break
        
        # Si no hay lance de captura total, usar el primero como fallback
        if not lance_captura_total:
            lance_captura_total = lances[0]
        
        total_camaron = 0
        total_merluza = 0
        especies_totales = {}
        
        # Procesar SOLO las especies del lance CAPTURA TOTAL
        especies = lance_captura_total.get('especies', [])
        for especie in especies:
            # Contar TODAS las especies (retenida Y descartada)
            nombre = especie.get('nombre', '')
            if not nombre:
                continue
            
            cantidad = especie.get('cantidad_ton', 0)
            
            # Acumular por especie
            if nombre not in especies_totales:
                especies_totales[nombre] = 0
            especies_totales[nombre] += cantidad
            
            # Totales específicos
            nombre_lower = nombre.lower()
            if 'camarón' in nombre_lower or 'camaron' in nombre_lower:
                total_camaron += cantidad
            elif 'merluza' in nombre_lower:
                total_merluza += cantidad
        
        # Identificar especies objetivo
        especies_objetivo_nombres = ['Camarón nailon', 'Langostino amarillo', 'Langostino colorado', 'Gamba', 'Camarón', 'Langostino']
        especies_objetivo = {}
        otras_especies = {}
        
        for especie, cantidad in especies_totales.items():
            es_objetivo = any(obj.lower() in especie.lower() for obj in especies_objetivo_nombres)
            if es_objetivo:
                especies_objetivo[especie] = cantidad
            else:
                otras_especies[especie] = cantidad
        
        # Ordenar por cantidad
        top_objetivo = sorted(especies_objetivo.items(), key=lambda x: x[1], reverse=True)[:3]
        top_otras = sorted(otras_especies.items(), key=lambda x: x[1], reverse=True)[:3]
        
        # Calcular ratio y alerta
        if total_camaron > 0:
            ratio = (total_merluza / total_camaron) * 100
            if ratio <= 10:
                alerta = "🟢 VERDE"
                alerta_color = "#00C853"
            elif ratio <= 20:
                alerta = "🟡 AMARILLO"
                alerta_color = "#FFB300"
            else:
                alerta = "🔴 ROJO"
                alerta_color = "#D32F2F"
        else:
            ratio = 0
            alerta = "⚪ N/A"
            alerta_color = "#999999"
        
        # Card principal con hover effects
        card = ctk.CTkFrame(
            parent,
            corner_radius=15,
            fg_color="white",
            border_width=3,
            border_color="#05BFDB"
        )
        card.pack(pady=15, padx=10, fill="x")
        
        # Efectos hover en la tarjeta
        def on_card_enter(e):
            try:
                card.configure(border_color="#0492A8", border_width=4)
            except (tk.TclError, Exception):
                pass
        
        def on_card_leave(e):
            try:
                card.configure(border_color="#05BFDB", border_width=3)
            except (tk.TclError, Exception):
                pass
        
        card.bind("<Enter>", on_card_enter)
        card.bind("<Leave>", on_card_leave)
        
        # SECCIÓN DE IDENTIFICACIÓN (Header destacado)
        id_section = ctk.CTkFrame(
            card,
            corner_radius=10,
            fg_color="#05BFDB",
            height=80
        )
        id_section.pack(fill="x", padx=5, pady=5)
        id_section.pack_propagate(False)
        
        # Número de viaje
        num_label = ctk.CTkLabel(
            id_section,
            text=f"BITÁCORA #{numero}",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="white"
        )
        num_label.pack(side="left", padx=20, pady=10)
        
        # ID/Folio principal
        id_viaje = viaje.get('id_viaje', 'N/A')
        folio_text = id_viaje if id_viaje != 'N/A' else viaje.get('folio_interno', 'N/A')
        
        ctk.CTkLabel(
            id_section,
            text=f"📋 {folio_text}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="white"
        ).pack(side="left", padx=10, pady=10)
        
        # Matrícula y señal
        id_info = ctk.CTkFrame(id_section, fg_color="transparent")
        id_info.pack(side="right", padx=20, pady=10)
        
        matricula = viaje.get('nave_matricula', 'N/A')
        senal = viaje.get('señal_llamada', 'N/A')
        
        ctk.CTkLabel(
            id_info,
            text=f"🏷️ Matrícula: {matricula}  |  📡 Señal: {senal}",
            font=ctk.CTkFont(size=11),
            text_color="white"
        ).pack()
        
        # Información del viaje
        info_frame = ctk.CTkFrame(card, fg_color="transparent")
        info_frame.pack(pady=15, padx=20, fill="x")
        
        # Obtener fechas formateadas
        fecha_zarpe_str = viaje.get('fecha_zarpe', 'N/A')
        fecha_recalada_str = viaje.get('fecha_recalada', 'N/A')
        
        try:
            if fecha_zarpe_str != 'N/A':
                from datetime import datetime
                fecha_zarpe_str = datetime.fromisoformat(fecha_zarpe_str).strftime('%d/%m/%Y %H:%M')
        except:
            pass
        
        try:
            if fecha_recalada_str != 'N/A':
                from datetime import datetime
                fecha_recalada_str = datetime.fromisoformat(fecha_recalada_str).strftime('%d/%m/%Y %H:%M')
        except:
            pass
        
        info_items = [
            ("🚀 Zarpe:", f"{fecha_zarpe_str} ({viaje.get('puerto_zarpe', 'N/A')})"),
            ("⚓ Recalada:", f"{fecha_recalada_str} ({viaje.get('puerto_recalada', 'N/A')})"),
            ("🚢 Nave:", viaje.get('nave_nombre', 'N/A')),
            ("👨‍✈️ Capitán:", viaje.get('capitan', 'DESCONOCIDO')),
            ("🏢 Armador:", viaje.get('armador', 'N/A'))
        ]
        
        for label, value in info_items:
            row = ctk.CTkFrame(info_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(
                row,
                text=label,
                font=ctk.CTkFont(size=12, weight="bold"),
                anchor="w",
                width=100,
                text_color="#000000"
            ).pack(side="left")
            
            ctk.CTkLabel(
                row,
                text=value,
                font=ctk.CTkFont(size=12),
                anchor="w",
                text_color="#1A1A1A"
            ).pack(side="left")
        
        # SEPARAR ESPECIES POR TIPO DE CAPTURA (usando solo lance CAPTURA TOTAL)
        especies_retenidas = {}
        especies_descartadas = {}  # Guardará {'nombre': {'ton': X, 'unidades': Y}}
        especies_incidentales = {}
        
        # Procesar SOLO las especies del lance CAPTURA TOTAL
        for especie in lance_captura_total.get('especies', []):
            nombre = especie.get('nombre', '')
            if not nombre:
                continue
            cantidad = especie.get('cantidad_ton', 0)
            cantidad_unidades = especie.get('cantidad_unidades', 0)
            tipo_captura = especie.get('tipo_captura', 'retenida')
            tipo_especie = especie.get('tipo_especie', 'otra')
            
            if tipo_captura == 'retenida':
                if nombre not in especies_retenidas:
                    especies_retenidas[nombre] = 0
                especies_retenidas[nombre] += cantidad
            elif tipo_captura == 'descartada':
                if nombre not in especies_descartadas:
                    especies_descartadas[nombre] = {'ton': 0, 'unidades': 0}
                especies_descartadas[nombre]['ton'] += cantidad
                especies_descartadas[nombre]['unidades'] += cantidad_unidades
            elif tipo_captura == 'incidental':
                if nombre not in especies_incidentales:
                    especies_incidentales[nombre] = 0
                especies_incidentales[nombre] += cantidad_unidades
        
        # Ordenar por cantidad
        top_retenidas = sorted(especies_retenidas.items(), key=lambda x: x[1], reverse=True)[:5]
        # Para descartadas, mostrar hasta 15 para incluir todas las especies en TON y unidades
        # Ordenar: primero por TON (mayor a menor), luego por unidades (mayor a menor)
        top_descartadas = sorted(especies_descartadas.items(), 
                                key=lambda x: (x[1]['ton'], x[1]['unidades']), 
                                reverse=True)[:15]
        
        # Calcular totales para el gráfico
        total_retenidas = sum(especies_retenidas.values())
        total_descartadas = sum(e['ton'] for e in especies_descartadas.values())
        total_descartadas_unidades = sum(e['unidades'] for e in especies_descartadas.values())
        
        # SECCIÓN DE INFORMACIÓN CLAVE
        resumen_frame = ctk.CTkFrame(card, fg_color="#E3F2FD", corner_radius=10, border_width=2, border_color="#2196F3")
        resumen_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        # Header con título y botón
        header_frame = ctk.CTkFrame(resumen_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=10, pady=(10, 5))
        
        ctk.CTkLabel(
            header_frame,
            text="📈 INFORMACIÓN CLAVE DEL VIAJE",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#1565C0"
        ).pack(side="left")
        
        # Frame para botones
        botones_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        botones_frame.pack(side="right")
        
        # Botón para generar mapa de calor
        ctk.CTkButton(
            botones_frame,
            text="🗺️ Mapa de Calor",
            command=lambda v=viaje: self.generar_mapa_calor([v], parent_window=self.resultado_window),
            fg_color="#4CAF50",
            hover_color="#388E3C",
            height=30,
            width=140,
            font=ctk.CTkFont(size=11, weight="bold")
        ).pack(side="right", padx=(5, 0))
        
        # Botón para generar gráfico
        ctk.CTkButton(
            botones_frame,
            text="📊 Gráfico",
            command=lambda er=especies_retenidas, ed=especies_descartadas, vid=viaje.get('id_viaje', 'N/A'): self.generar_grafico_especies(er, ed, vid, parent_window=self.resultado_window),
            fg_color="#1976D2",
            hover_color="#1565C0",
            height=30,
            width=140,
            font=ctk.CTkFont(size=11, weight="bold")
        ).pack(side="right", padx=(0, 5))
        
        # Grid de información clave
        resumen_grid = ctk.CTkFrame(resumen_frame, fg_color="transparent")
        resumen_grid.pack(pady=(0, 10), padx=20)
        
        # Calcular número de especies y determinar si hay unidades descartadas
        num_especies = len(especies_retenidas) + len(especies_descartadas) + len(especies_incidentales)
        num_lances = viaje.get('total_lances_declarados', 0)  # Obtener del viaje, no de len(lances)
        total_incidentales = sum(especies_incidentales.values())
        
        # Determinar el texto de captura descartada
        if total_descartadas_unidades > 0:
            descartada_texto = f"{total_descartadas:.2f} toneladas y {int(total_descartadas_unidades)} unidades"
        else:
            descartada_texto = f"{total_descartadas:.2f} toneladas"
        
        stats_items = [
            ("🎣 Total de Lances:", f"{num_lances}", "#1976D2"),
            ("🐟 Especies Diferentes:", f"{num_especies}", "#00796B"),
            ("✅ Captura Retenida:", f"{total_retenidas:.2f} toneladas", "#388E3C"),
            ("⚠️ Captura Descartada:", descartada_texto, "#F57C00")
        ]
        
        if total_incidentales > 0:
            stats_items.append(("🦭 Incidental:", f"{total_incidentales} individuo(s)", "#9C27B0"))
        
        for label, value, color in stats_items:
            row = ctk.CTkFrame(resumen_grid, fg_color="transparent")
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(
                row,
                text=label,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color="#000000",
                anchor="w"
            ).pack(side="left", padx=(0, 10))
            
            ctk.CTkLabel(
                row,
                text=value,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=color,
                anchor="e"
            ).pack(side="right")
        
        # SECCIÓN DE ESPECIES DETALLADAS
        especies_frame = ctk.CTkFrame(card, fg_color="transparent")
        especies_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        # ESPECIES RETENIDAS
        if top_retenidas:
            retenidas_frame = ctk.CTkFrame(especies_frame, fg_color="#E8F5E9", corner_radius=10, border_width=1, border_color="#4CAF50")
            retenidas_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
            
            ctk.CTkLabel(
                retenidas_frame,
                text="✅ ESPECIES RETENIDAS",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#2E7D32"
            ).pack(anchor="w", padx=10, pady=(10, 5))
            
            for especie, cantidad in top_retenidas:
                esp_row = ctk.CTkFrame(retenidas_frame, fg_color="transparent")
                esp_row.pack(fill="x", padx=10, pady=2)
                
                # Calcular porcentaje respecto al total retenido
                porc_retenido = (cantidad / total_retenidas * 100) if total_retenidas > 0 else 0
                
                ctk.CTkLabel(
                    esp_row,
                    text=f"• {especie}",
                    font=ctk.CTkFont(size=11),
                    text_color="#1B5E20",
                    anchor="w",
                    wraplength=150
                ).pack(side="left", fill="x", expand=True)
                
                ctk.CTkLabel(
                    esp_row,
                    text=f"{cantidad:.3f} TON ({porc_retenido:.1f}%)",
                    font=ctk.CTkFont(size=10, weight="bold"),
                    text_color="#388E3C"
                ).pack(side="right")
            
            # Padding inferior
            ctk.CTkFrame(retenidas_frame, fg_color="transparent", height=5).pack()
        
        # ESPECIES DESCARTADAS
        if top_descartadas:
            descartadas_frame = ctk.CTkFrame(especies_frame, fg_color="#FFF3E0", corner_radius=10, border_width=1, border_color="#FF9800")
            descartadas_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))
            
            # Header
            header_desc = ctk.CTkFrame(descartadas_frame, fg_color="transparent")
            header_desc.pack(fill="x", padx=10, pady=(10, 5))
            
            ctk.CTkLabel(
                header_desc,
                text="⚠️ ESPECIES DESCARTADAS",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#E65100"
            ).pack(side="left")
            
            ctk.CTkLabel(
                header_desc,
                text=f"({len(top_descartadas)})",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color="#F57C00"
            ).pack(side="left", padx=(5, 0))
            
            # Contenedor scrolleable si hay muchas especies
            if len(top_descartadas) > 6:
                especies_container = ctk.CTkScrollableFrame(
                    descartadas_frame, 
                    fg_color="transparent",
                    height=200
                )
                especies_container.pack(fill="both", expand=True, padx=5, pady=(0, 5))
            else:
                especies_container = descartadas_frame
            
            for especie, datos in top_descartadas:
                esp_row = ctk.CTkFrame(especies_container, fg_color="transparent")
                esp_row.pack(fill="x", padx=10, pady=2)
                
                ctk.CTkLabel(
                    esp_row,
                    text=f"• {especie}",
                    font=ctk.CTkFont(size=11),
                    text_color="#BF360C",
                    anchor="w",
                    wraplength=150
                ).pack(side="left", fill="x", expand=True)
                
                # Mostrar TON o unidades según corresponda con diferenciación visual
                ton = datos['ton']
                unidades = datos['unidades']
                
                # Calcular porcentaje respecto al total descartado (solo TON)
                if ton > 0:
                    porc_descartado = (ton / total_descartadas * 100) if total_descartadas > 0 else 0
                    cantidad_texto = f"{ton:.3f} TON ({porc_descartado:.1f}%)"
                    color_cantidad = "#F57C00"  # Naranja para TON
                else:
                    # Para unidades, calcular respecto al total de unidades
                    porc_unidades = (unidades / total_descartadas_unidades * 100) if total_descartadas_unidades > 0 else 0
                    cantidad_texto = f"⚠️ {int(unidades)} UNID ({porc_unidades:.1f}%)"
                    color_cantidad = "#9C27B0"  # Morado para UNIDADES (diferenciación visual)
                
                ctk.CTkLabel(
                    esp_row,
                    text=cantidad_texto,
                    font=ctk.CTkFont(size=10, weight="bold"),
                    text_color=color_cantidad
                ).pack(side="right")
            
            # Padding inferior
            ctk.CTkFrame(descartadas_frame, fg_color="transparent", height=5).pack()
        
        # ESPECIES INCIDENTALES
        if especies_incidentales:
            incidentales_frame = ctk.CTkFrame(especies_frame, fg_color="#F3E5F5", corner_radius=10, border_width=1, border_color="#9C27B0")
            incidentales_frame.pack(fill="x", padx=0, pady=(10, 0))
            
            inc_header = ctk.CTkFrame(incidentales_frame, fg_color="transparent")
            inc_header.pack(fill="x", padx=10, pady=(10, 5))
            
            ctk.CTkLabel(
                inc_header,
                text="🦭 ESPECIES INCIDENTALES",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#6A1B9A"
            ).pack(side="left")
            
            total_inc = sum(especies_incidentales.values())
            ctk.CTkLabel(
                inc_header,
                text=f"({total_inc} ind.)",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color="#9C27B0"
            ).pack(side="left", padx=(5, 0))
            
            for especie, cantidad in sorted(especies_incidentales.items(), key=lambda x: x[1], reverse=True):
                esp_row = ctk.CTkFrame(incidentales_frame, fg_color="transparent")
                esp_row.pack(fill="x", padx=10, pady=2)
                
                ctk.CTkLabel(
                    esp_row,
                    text=f"• {especie}",
                    font=ctk.CTkFont(size=11),
                    text_color="#6A1B9A",
                    anchor="w",
                    wraplength=200
                ).pack(side="left", fill="x", expand=True)
                
                ctk.CTkLabel(
                    esp_row,
                    text=f"🦭 {int(cantidad)} ind.",
                    font=ctk.CTkFont(size=10, weight="bold"),
                    text_color="#9C27B0"
                ).pack(side="right")
            
            ctk.CTkFrame(incidentales_frame, fg_color="transparent", height=5).pack()
    
    def generar_grafico_especies(self, especies_retenidas, especies_descartadas, id_viaje, parent_window=None):
        """Genera un gráfico de torta con todas las especies y tabla de detalles"""
        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib.patches as mpatches
            import tkinter as tk
            
            # Usar parent_window si está disponible, sino self.root
            parent = parent_window if parent_window else self.root
            
            # Crear ventana nueva
            ventana = tk.Toplevel(parent)
            ventana.title(f"Gráfico de Capturas - {id_viaje}")
            # Adaptar tamaño a la pantalla
            scr_w = ventana.winfo_screenwidth()
            scr_h = ventana.winfo_screenheight()
            g_w = min(1400, scr_w - 40)
            g_h = min(800, scr_h - 100)
            ventana.geometry(f"{g_w}x{g_h}+{max(10, (scr_w - g_w) // 2)}+{max(10, (scr_h - g_h) // 2 - 30)}")
            
            # Configurar para que aparezca al frente
            ventana.transient(parent)
            ventana.lift()
            ventana.focus_force()
            ventana.attributes('-topmost', True)
            ventana.after(500, lambda: ventana.attributes('-topmost', False))
            
            # Función para cerrar correctamente
            def cerrar_ventana():
                plt.close('all')
                ventana.destroy()
            
            ventana.protocol("WM_DELETE_WINDOW", cerrar_ventana)
            
            # Preparar datos (SOLO TONELADAS, no unidades)
            especies_data = []
            
            # Agregar especies retenidas
            for nombre, toneladas in especies_retenidas.items():
                if toneladas > 0:  # Solo especies con toneladas
                    especies_data.append({
                        'nombre': nombre,
                        'toneladas': toneladas,
                        'tipo': 'retenida',
                        'color_base': 'green'
                    })
            
            # Agregar especies descartadas (SOLO con toneladas)
            for nombre, datos in especies_descartadas.items():
                if datos['ton'] > 0:  # Solo especies con toneladas (ignorar unidades)
                    especies_data.append({
                        'nombre': nombre,
                        'toneladas': datos['ton'],
                        'tipo': 'descartada',
                        'color_base': 'red'
                    })
            
            # Ordenar por cantidad
            especies_data.sort(key=lambda x: x['toneladas'], reverse=True)
            
            # Calcular total
            total_toneladas = sum(e['toneladas'] for e in especies_data)
            
            # Generar paleta de colores profesional
            # Verdes para retenidas (del más claro al más oscuro)
            colores_verdes = ['#00ff00', '#00e600', '#00cc00', '#00b300', '#009900', 
                             '#008000', '#006600', '#004d00', '#003300', '#001a00']
            # Rojos/naranjas para descartadas
            colores_rojos = ['#ff0000', '#ff3300', '#ff6600', '#ff9900', '#ffcc00',
                            '#cc0000', '#990000', '#660000', '#ff4444', '#ff6666']
            
            # Asignar colores
            idx_verde = 0
            idx_rojo = 0
            colores_finales = []
            
            for especie in especies_data:
                if especie['tipo'] == 'retenida':
                    colores_finales.append(colores_verdes[idx_verde % len(colores_verdes)])
                    idx_verde += 1
                else:
                    colores_finales.append(colores_rojos[idx_rojo % len(colores_rojos)])
                    idx_rojo += 1
            
            # Crear figura con un subplot para el gráfico y espacio para tabla
            fig = plt.figure(figsize=(16, 9))
            gs = fig.add_gridspec(1, 2, width_ratios=[1, 1.3], wspace=0.25)
            ax_pie = fig.add_subplot(gs[0])
            ax_table = fig.add_subplot(gs[1])
            
            fig.suptitle(f'Análisis de Capturas - {id_viaje}', fontsize=18, fontweight='bold', y=0.98)
            
            # Gráfico de torta LIMPIO con mejor separación visual
            if especies_data:
                sizes = [e['toneladas'] for e in especies_data]
                
                # Pequeño explode para todas las secciones (mejor separación visual)
                explode_vals = [0.05] * len(especies_data)
                
                # Pie chart limpio con mejor separación
                wedges, texts = ax_pie.pie(
                    sizes, 
                    labels=None,  # Sin labels
                    colors=colores_finales,
                    autopct=None,  # Sin porcentajes internos
                    explode=explode_vals,  # Separación ligera entre secciones
                    shadow=False,
                    startangle=90,
                    wedgeprops={'edgecolor': 'white', 'linewidth': 3}  # Bordes más gruesos para distinguir
                )
                
                ax_pie.set_title('Distribución de Especies\n(Solo Toneladas)', 
                               fontsize=14, fontweight='bold', pad=15)
                
                # AGREGAR TOOLTIPS INTERACTIVOS al pasar el mouse
                # Crear annotation que se mostrará al hover
                annot = ax_pie.annotate("", xy=(0,0), xytext=(20,20),
                                       textcoords="offset points",
                                       bbox=dict(boxstyle="round,pad=0.8", fc="#05BFDB", ec="#0492A8", lw=2, alpha=0.95),
                                       arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.3", 
                                                      color="#0492A8", lw=2),
                                       fontsize=11, fontweight='bold', color='white',
                                       visible=False, zorder=100)
                
                def on_hover(event):
                    """Mostrar información al pasar el mouse sobre una sección"""
                    if event.inaxes == ax_pie:
                        for i, wedge in enumerate(wedges):
                            cont, ind = wedge.contains(event)
                            if cont:
                                # Mostrar información de la especie
                                nombre = especies_data[i]['nombre']
                                toneladas = especies_data[i]['toneladas']
                                porcentaje = (toneladas / total_toneladas * 100) if total_toneladas > 0 else 0
                                tipo = 'RETENIDA' if especies_data[i]['tipo'] == 'retenida' else 'DESCARTE'
                                
                                # Calcular porcentaje respecto a su categoría
                                if especies_data[i]['tipo'] == 'retenida':
                                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
                                    texto_cat = "% del Retenido"
                                else:
                                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
                                    texto_cat = "% del Descarte"
                                
                                porc_categoria = (toneladas / total_categoria * 100) if total_categoria > 0 else 0
                                
                                # Actualizar texto del tooltip con info completa
                                annot.set_text(f"{nombre}\n{tipo}\n{toneladas:.3f} TON\n{porcentaje:.1f}% del Total\n{porc_categoria:.1f}% {texto_cat}")
                                annot.xy = (event.xdata, event.ydata)
                                annot.set_visible(True)
                                fig.canvas.draw_idle()
                                return
                        
                        # Si no está sobre ningún wedge, ocultar tooltip
                        if annot.get_visible():
                            annot.set_visible(False)
                            fig.canvas.draw_idle()
                
                # Conectar evento de movimiento del mouse
                fig.canvas.mpl_connect("motion_notify_event", on_hover)
            
            # TABLA PROFESIONAL DE DETALLES
            ax_table.axis('off')
            ax_table.set_xlim(0, 10)
            ax_table.set_ylim(0, 10)
            
            # Título de la tabla con fondo
            from matplotlib.patches import Rectangle
            rect_title = Rectangle((0.2, 9.2), 9.6, 0.7, 
                                   facecolor='#05BFDB', edgecolor='#0492A8', linewidth=2)
            ax_table.add_patch(rect_title)
            ax_table.text(5, 9.55, 'DETALLE DE ESPECIES POR TONELADAS', 
                         ha='center', va='center', fontsize=13, fontweight='bold', color='white')
            
            # Encabezados con fondo gris
            y_pos = 8.7
            rect_header = Rectangle((0.2, y_pos - 0.15), 9.6, 0.35, 
                                    facecolor='#e0e0e0', edgecolor='gray', linewidth=1)
            ax_table.add_patch(rect_header)
            
            ax_table.text(0.8, y_pos, '█', ha='center', fontsize=14, fontweight='bold')
            ax_table.text(2.2, y_pos, 'Especie', ha='left', fontsize=11, fontweight='bold')
            ax_table.text(5.8, y_pos, 'Tipo', ha='center', fontsize=11, fontweight='bold')
            ax_table.text(7.2, y_pos, 'TON', ha='right', fontsize=11, fontweight='bold')
            ax_table.text(8.5, y_pos, '% Total', ha='right', fontsize=11, fontweight='bold')
            ax_table.text(9.6, y_pos, '% Cat.', ha='right', fontsize=11, fontweight='bold')
            
            # Línea separadora
            y_pos -= 0.35
            ax_table.plot([0.2, 9.8], [y_pos, y_pos], 'k-', linewidth=2)
            
            # Datos de especies con alternancia de fondo
            y_pos -= 0.45
            max_especies_mostrar = 18  # Máximo que cabe en pantalla
            
            for i, especie in enumerate(especies_data[:max_especies_mostrar]):
                # Fondo alternado para mejor lectura
                if i % 2 == 0:
                    rect_bg = Rectangle((0.2, y_pos - 0.18), 9.6, 0.35, 
                                       facecolor='#f5f5f5', edgecolor='none')
                    ax_table.add_patch(rect_bg)
                
                porcentaje = (especie['toneladas'] / total_toneladas * 100) if total_toneladas > 0 else 0
                
                # Calcular porcentaje respecto a su categoría
                if especie['tipo'] == 'retenida':
                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
                else:
                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
                
                porc_categoria = (especie['toneladas'] / total_categoria * 100) if total_categoria > 0 else 0
                
                # Cuadrado de color más grande
                ax_table.text(0.8, y_pos, '█', ha='center', fontsize=16, 
                            color=colores_finales[i])
                
                # Nombre de especie (acortado para dar espacio a nueva columna)
                nombre = especie['nombre'][:25]
                ax_table.text(2.2, y_pos, nombre, ha='left', fontsize=9,
                            weight='bold' if i < 5 else 'normal')
                
                # Tipo (Retenida/Descartada)
                tipo_texto = 'RETENIDA' if especie['tipo'] == 'retenida' else 'DESCARTE'
                tipo_color = '#00a000' if especie['tipo'] == 'retenida' else '#ff3300'
                ax_table.text(5.8, y_pos, tipo_texto, ha='center', fontsize=8,
                            color=tipo_color, weight='bold',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor=tipo_color, 
                                     alpha=0.2, edgecolor=tipo_color, linewidth=1))
                
                # Toneladas
                ax_table.text(7.2, y_pos, f"{especie['toneladas']:.3f}", 
                            ha='right', fontsize=9, weight='bold')
                
                # Porcentaje del total
                ax_table.text(8.5, y_pos, f"{porcentaje:.1f}%", 
                            ha='right', fontsize=9, weight='bold',
                            color='#0277bd')
                
                # Porcentaje de su categoría (NUEVA COLUMNA)
                ax_table.text(9.6, y_pos, f"{porc_categoria:.1f}%", 
                            ha='right', fontsize=9, weight='bold',
                            color=tipo_color)
                
                y_pos -= 0.4
            
            # Si hay más especies, indicarlo
            if len(especies_data) > max_especies_mostrar:
                ax_table.text(5, y_pos, f'...y {len(especies_data) - max_especies_mostrar} especies más', 
                            ha='center', fontsize=9, style='italic', color='gray')
                y_pos -= 0.4
            
            # SECCIÓN DE TOTALES con fondo destacado
            y_pos -= 0.3
            ax_table.plot([0.2, 9.8], [y_pos, y_pos], 'k-', linewidth=2.5)
            y_pos -= 0.5
            
            total_retenidas = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
            total_descartadas = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
            
            # Total retenidas con fondo verde claro
            rect_ret = Rectangle((0.2, y_pos - 0.2), 9.6, 0.4, 
                                facecolor='#e8f5e9', edgecolor='#00a000', linewidth=1.5)
            ax_table.add_patch(rect_ret)
            ax_table.text(2.5, y_pos, 'TOTAL RETENIDAS:', ha='left', fontsize=11, 
                        weight='bold', color='#00a000')
            ax_table.text(8, y_pos, f"{total_retenidas:.3f}", ha='right', 
                        fontsize=11, weight='bold', color='#00a000')
            porc_ret = (total_retenidas / total_toneladas * 100) if total_toneladas > 0 else 0
            ax_table.text(9.5, y_pos, f"{porc_ret:.1f}%", ha='right', 
                        fontsize=11, weight='bold', color='#00a000')
            
            y_pos -= 0.5
            
            # Total descartadas con fondo rojo claro
            rect_desc = Rectangle((0.2, y_pos - 0.2), 9.6, 0.4, 
                                  facecolor='#ffebee', edgecolor='#ff3300', linewidth=1.5)
            ax_table.add_patch(rect_desc)
            ax_table.text(2.5, y_pos, 'TOTAL DESCARTADAS:', ha='left', fontsize=11, 
                        weight='bold', color='#ff3300')
            ax_table.text(8, y_pos, f"{total_descartadas:.3f}", ha='right', 
                        fontsize=11, weight='bold', color='#ff3300')
            porc_desc = (total_descartadas / total_toneladas * 100) if total_toneladas > 0 else 0
            ax_table.text(9.5, y_pos, f"{porc_desc:.1f}%", ha='right', 
                        fontsize=11, weight='bold', color='#ff3300')
            
            y_pos -= 0.6
            
            # Total general con fondo azul
            rect_total = Rectangle((0.2, y_pos - 0.25), 9.6, 0.5, 
                                   facecolor='#e3f2fd', edgecolor='#0277bd', linewidth=2)
            ax_table.add_patch(rect_total)
            ax_table.text(2.5, y_pos, 'TOTAL GENERAL:', ha='left', fontsize=12, 
                        weight='bold', color='#0277bd')
            ax_table.text(8, y_pos, f"{total_toneladas:.3f} TON", ha='right', 
                        fontsize=12, weight='bold', color='#0277bd')
            ax_table.text(9.5, y_pos, "100%", ha='right', 
                        fontsize=12, weight='bold', color='#0277bd')
            
            # Suprimir warning de tight_layout
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                plt.tight_layout()
            
            # Integrar en ventana Tkinter
            canvas = FigureCanvasTkAgg(fig, master=ventana)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Frame para botones
            btn_frame = tk.Frame(ventana, bg='white')
            btn_frame.pack(fill=tk.X, padx=10, pady=10)
            
            # Botón cerrar
            btn_cerrar = tk.Button(
                btn_frame,
                text="✖ Cerrar Gráfico",
                command=cerrar_ventana,
                bg="#D32F2F",
                fg="white",
                font=("Arial", 12, "bold"),
                padx=30,
                pady=10,
                relief=tk.RAISED,
                cursor="hand2"
            )
            btn_cerrar.pack(side=tk.RIGHT)
            
        except ImportError as e:
            from CTkMessagebox import CTkMessagebox
            CTkMessagebox(
                title="Librería no disponible",
                message=f"No se pudo cargar matplotlib:\n{e}",
                icon="warning"
            )
        except Exception as e:
            print(f"Error generando gráfico: {e}")
            import traceback
            traceback.print_exc()
    
    def exportar_resumen_excel(self, viajes, especies_retenidas, especies_descartadas,
                               total_retenidas_ton, total_descartadas_ton, total_general,
                               total_lances, total_especies, porc_retenidas, porc_descartadas,
                               especies_incidentales=None, total_incidentales_unidades=0):
        """Exporta el resumen total de capturas a un archivo Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Pedir ubicación para guardar
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar Resumen Excel",
            initialfile=f"Resumen_Capturas_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not archivo:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumen Capturas"
            
            # Estilos
            titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            bold_font = Font(name='Calibri', size=11, bold=True)
            normal_font = Font(name='Calibri', size=11)
            small_font = Font(name='Calibri', size=10, italic=True, color='888888')
            
            titulo_fill = PatternFill(start_color='0A4D68', end_color='0A4D68', fill_type='solid')
            header_ret_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            header_desc_fill = PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid')
            light_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            light_red = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            info_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            right = Alignment(horizontal='right', vertical='center')
            
            # Obtener rango de fechas de los filtros
            try:
                fecha_desde = self.fecha_desde.get_date()
                fecha_hasta = self.fecha_hasta.get_date()
                rango_str = f"{fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
            except:
                rango_str = "Sin filtro de fechas"
                fecha_desde = None
                fecha_hasta = None
            
            # Obtener años únicos
            años_set = set()
            for v in viajes:
                fz = v.get('fecha_zarpe', '')
                if fz:
                    try:
                        años_set.add(datetime.fromisoformat(fz).year)
                    except:
                        pass
            años_ordenados = sorted(años_set)
            
            # Anchos de columna
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            
            row = 1
            
            # ===== TÍTULO =====
            ws.merge_cells('A1:E1')
            cell = ws.cell(row=1, column=1, value='RESUMEN TOTAL DE CAPTURAS - BENTOS MSC')
            cell.font = titulo_font
            cell.fill = titulo_fill
            cell.alignment = center
            ws.row_dimensions[1].height = 35
            
            # Subtítulo con rango de fechas
            row = 2
            ws.merge_cells('A2:E2')
            cell = ws.cell(row=2, column=1, value=f'Período: {rango_str}')
            cell.font = Font(name='Calibri', size=12, bold=True, color='0A4D68')
            cell.fill = PatternFill(start_color='B2EBF2', end_color='B2EBF2', fill_type='solid')
            cell.alignment = center
            ws.row_dimensions[2].height = 25
            
            # ===== ESTADÍSTICAS GENERALES =====
            row = 4
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value='ESTADÍSTICAS GENERALES')
            cell.font = header_font
            cell.fill = titulo_fill
            cell.alignment = center
            
            stats_data = [
                ('Bitácoras analizadas', len(viajes)),
                ('Lances totales', total_lances),
                ('Especies diferentes', total_especies),
                ('Captura total (TON)', f'{total_general:.3f}'),
                ('Captura retenida (TON)', f'{total_retenidas_ton:.3f} ({porc_retenidas:.1f}%)'),
                ('Descarte (TON)', f'{total_descartadas_ton:.3f} ({porc_descartadas:.1f}%)'),
            ]
            
            for label, valor in stats_data:
                row += 1
                c1 = ws.cell(row=row, column=2, value=label)
                c1.font = bold_font
                c1.fill = info_fill
                c1.border = thin_border
                c1.alignment = left
                c2 = ws.cell(row=row, column=3, value=str(valor))
                c2.font = normal_font
                c2.fill = info_fill
                c2.border = thin_border
                c2.alignment = center
            
            # ===== LISTADO DE BITÁCORAS =====
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value='DETALLE DE BITÁCORAS')
            cell.font = header_font
            cell.fill = titulo_fill
            cell.alignment = center
            
            row += 1
            for col_idx, header_text in enumerate(['#', 'Folio', 'Nave', 'Capitán', 'Fecha Zarpe'], 1):
                c = ws.cell(row=row, column=col_idx, value=header_text)
                c.font = header_font
                c.fill = PatternFill(start_color='1A5F7A', end_color='1A5F7A', fill_type='solid')
                c.border = thin_border
                c.alignment = center
            
            for i, viaje in enumerate(viajes, 1):
                row += 1
                fz = viaje.get('fecha_zarpe', '')
                try:
                    fecha_fmt = datetime.fromisoformat(fz).strftime('%d/%m/%Y')
                except:
                    fecha_fmt = fz
                
                vals = [i, viaje.get('id_viaje', 'N/A'), viaje.get('nave_nombre', 'N/A'),
                        viaje.get('capitan', 'N/A'), fecha_fmt]
                bg = light_green if i % 2 == 0 else PatternFill()
                for col_idx, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.font = normal_font
                    c.border = thin_border
                    c.alignment = center if col_idx != 4 else left
                    if i % 2 == 0:
                        c.fill = bg
            
            # ===== ESPECIES RETENIDAS =====
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f'ESPECIES RETENIDAS — Total: {total_retenidas_ton:.3f} TON ({porc_retenidas:.1f}%)')
            cell.font = header_font
            cell.fill = header_ret_fill
            cell.alignment = center
            
            row += 1
            for col_idx, header_text in enumerate(['#', 'Especie', 'Captura (TON)', '% del Retenido'], 1):
                c = ws.cell(row=row, column=col_idx, value=header_text)
                c.font = header_font
                c.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
                c.border = thin_border
                c.alignment = center
            
            especies_ret_ord = sorted(especies_retenidas.items(), key=lambda x: x[1], reverse=True)
            for i, (especie, cantidad) in enumerate(especies_ret_ord, 1):
                row += 1
                porc = (cantidad / total_retenidas_ton * 100) if total_retenidas_ton > 0 else 0
                vals = [i, especie, f'{cantidad:.3f}', f'{porc:.1f}%']
                for col_idx, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.font = normal_font
                    c.border = thin_border
                    c.alignment = left if col_idx == 2 else center
                    if i % 2 == 0:
                        c.fill = light_green
            
            # ===== ESPECIES DESCARTADAS =====
            if especies_descartadas:
                row += 2
                ws.merge_cells(f'A{row}:E{row}')
                cell = ws.cell(row=row, column=1, value=f'ESPECIES DESCARTADAS — Total: {total_descartadas_ton:.3f} TON ({porc_descartadas:.1f}%)')
                cell.font = header_font
                cell.fill = header_desc_fill
                cell.alignment = center
                
                row += 1
                for col_idx, header_text in enumerate(['#', 'Especie', 'TON', 'Unidades', '% del Descarte'], 1):
                    c = ws.cell(row=row, column=col_idx, value=header_text)
                    c.font = header_font
                    c.fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
                    c.border = thin_border
                    c.alignment = center
                
                especies_desc_ord = sorted(especies_descartadas.items(),
                                          key=lambda x: (x[1]['ton'], x[1]['unidades']), reverse=True)
                for i, (especie, datos) in enumerate(especies_desc_ord, 1):
                    row += 1
                    cant_ton = datos['ton']
                    cant_uni = datos['unidades']
                    porc = (cant_ton / total_descartadas_ton * 100) if total_descartadas_ton > 0 else 0
                    vals = [i, especie,
                            f'{cant_ton:.3f}' if cant_ton > 0 else '-',
                            int(cant_uni) if cant_uni > 0 else '-',
                            f'{porc:.1f}%' if cant_ton > 0 else '-']
                    for col_idx, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col_idx, value=val)
                        c.font = normal_font
                        c.border = thin_border
                        c.alignment = left if col_idx == 2 else center
                        if i % 2 == 0:
                            c.fill = light_red
            
            # ===== ESPECIES INCIDENTALES =====
            if especies_incidentales:
                row += 2
                ws.merge_cells(f'A{row}:E{row}')
                cell = ws.cell(row=row, column=1, value=f'ESPECIES INCIDENTALES (FAUNA ACOMPAÑANTE) — Total: {int(total_incidentales_unidades)} individuo(s)')
                cell.font = header_font
                header_inc_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
                cell.fill = header_inc_fill
                cell.alignment = center
                
                row += 1
                for col_idx, header_text in enumerate(['#', 'Especie', 'Individuos', '% del Total'], 1):
                    c = ws.cell(row=row, column=col_idx, value=header_text)
                    c.font = header_font
                    c.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
                    c.border = thin_border
                    c.alignment = center
                
                light_purple = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
                especies_inc_ord = sorted(especies_incidentales.items(), key=lambda x: x[1], reverse=True)
                for i, (especie, cantidad) in enumerate(especies_inc_ord, 1):
                    row += 1
                    porc = (cantidad / total_incidentales_unidades * 100) if total_incidentales_unidades > 0 else 0
                    vals = [i, especie, int(cantidad), f'{porc:.1f}%']
                    for col_idx, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col_idx, value=val)
                        c.font = normal_font
                        c.border = thin_border
                        c.alignment = left if col_idx == 2 else center
                        if i % 2 == 0:
                            c.fill = light_purple
            
            # ===== PIE DE PÁGINA CON AÑO =====
            row += 3
            ws.merge_cells(f'A{row}:E{row}')
            if años_ordenados:
                if len(años_ordenados) == 1:
                    año_texto = f"Año: {años_ordenados[0]}"
                else:
                    año_texto = f"Años: {años_ordenados[0]} - {años_ordenados[-1]}"
            else:
                año_texto = ""
            
            cell = ws.cell(row=row, column=1, value=f'Generado por BENTOS MSC — {datetime.now().strftime("%d/%m/%Y %H:%M")} — {año_texto}')
            cell.font = small_font
            cell.alignment = center
            
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f'Período filtrado: {rango_str} — Pesquera Quintero S.A.')
            cell.font = small_font
            cell.alignment = center
            
            # Guardar
            wb.save(archivo)
            
            CTkMessagebox(
                title="✅ Excel Exportado",
                message=f"Resumen exportado correctamente a:\n{archivo}",
                icon="check",
                option_1="OK"
            )
            
            # Abrir el archivo
            import subprocess, platform
            try:
                if platform.system() == 'Darwin':  # macOS
                    subprocess.Popen(['open', archivo])
                elif platform.system() == 'Windows':
                    os.startfile(archivo)
                else:
                    subprocess.Popen(['xdg-open', archivo])
            except Exception:
                pass
            
        except Exception as e:
            print(f"Error exportando Excel: {e}")
            CTkMessagebox(
                title="❌ Error",
                message=f"Error al exportar a Excel:\n{str(e)}",
                icon="cancel",
                option_1="OK"
            )
    
    def generar_grafico_resumen_total(self, viajes):
        """Genera un gráfico del resumen total de todas las bitácoras"""
        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.patches import Rectangle
            import tkinter as tk
            
            # Calcular datos agregados
            especies_retenidas = {}
            especies_descartadas = {}
            
            for viaje in viajes:
                viaje_id = viaje.get('id_viaje')
                lances = self.firebase.obtener_lances_viaje(viaje_id)
                
                # FILTRAR SOLO EL LANCE CAPTURA TOTAL (lance 0)
                lance_captura_total = None
                for lance in lances:
                    if lance.get('es_captura_total', False):
                        lance_captura_total = lance
                        break
                
                # Si no hay lance de captura total, usar el primero como fallback
                if not lance_captura_total and lances:
                    lance_captura_total = lances[0]
                
                # Procesar SOLO las especies del lance CAPTURA TOTAL
                if lance_captura_total:
                    for especie in lance_captura_total.get('especies', []):
                        nombre = especie.get('nombre', '')
                        if not nombre:
                            continue
                        cantidad_ton = especie.get('cantidad_ton', 0)
                        cantidad_unidades = especie.get('cantidad_unidades', 0)
                        tipo_captura = especie.get('tipo_captura', 'retenida')
                        
                        if tipo_captura == 'retenida':
                            if nombre not in especies_retenidas:
                                especies_retenidas[nombre] = 0
                            especies_retenidas[nombre] += cantidad_ton
                        elif tipo_captura == 'descartada':
                            if nombre not in especies_descartadas:
                                especies_descartadas[nombre] = {'ton': 0, 'unidades': 0}
                            especies_descartadas[nombre]['ton'] += cantidad_ton
                            especies_descartadas[nombre]['unidades'] += cantidad_unidades
            
            # Preparar datos (SOLO TONELADAS, no unidades)
            especies_data = []
            
            # Agregar especies retenidas
            for nombre, toneladas in especies_retenidas.items():
                if toneladas > 0:
                    especies_data.append({
                        'nombre': nombre,
                        'toneladas': toneladas,
                        'tipo': 'retenida',
                        'color_base': 'green'
                    })
            
            # Agregar especies descartadas (SOLO con toneladas)
            for nombre, datos in especies_descartadas.items():
                if datos['ton'] > 0:
                    especies_data.append({
                        'nombre': nombre,
                        'toneladas': datos['ton'],
                        'tipo': 'descartada',
                        'color_base': 'red'
                    })
            
            # Ordenar por cantidad
            especies_data.sort(key=lambda x: x['toneladas'], reverse=True)
            
            # Calcular total
            total_toneladas = sum(e['toneladas'] for e in especies_data)
            
            # Generar paleta de colores profesional
            colores_verdes = ['#00ff00', '#00e600', '#00cc00', '#00b300', '#009900', 
                             '#008000', '#006600', '#004d00', '#003300', '#001a00']
            colores_rojos = ['#ff0000', '#ff3300', '#ff6600', '#ff9900', '#ffcc00',
                            '#cc0000', '#990000', '#660000', '#ff4444', '#ff6666']
            
            # Asignar colores
            idx_verde = 0
            idx_rojo = 0
            colores_finales = []
            
            for especie in especies_data:
                if especie['tipo'] == 'retenida':
                    colores_finales.append(colores_verdes[idx_verde % len(colores_verdes)])
                    idx_verde += 1
                else:
                    colores_finales.append(colores_rojos[idx_rojo % len(colores_rojos)])
                    idx_rojo += 1
            
            # Crear ventana
            parent = self.root
            if hasattr(self, 'resultado_window'):
                try:
                    if self.resultado_window.winfo_exists():
                        parent = self.resultado_window
                except Exception:
                    pass
            ventana = tk.Toplevel(parent)
            ventana.title(f"Resumen Total - {len(viajes)} Bitácoras")
            # Adaptar tamaño a la pantalla
            scr_w = ventana.winfo_screenwidth()
            scr_h = ventana.winfo_screenheight()
            g_w = min(1400, scr_w - 40)
            g_h = min(800, scr_h - 100)
            ventana.geometry(f"{g_w}x{g_h}+{max(10, (scr_w - g_w) // 2)}+{max(10, (scr_h - g_h) // 2 - 30)}")
            
            ventana.transient(parent)
            ventana.lift()
            ventana.focus_force()
            ventana.attributes('-topmost', True)
            ventana.after(500, lambda: ventana.attributes('-topmost', False))
            
            def cerrar_ventana():
                plt.close('all')
                ventana.destroy()
            
            ventana.protocol("WM_DELETE_WINDOW", cerrar_ventana)
            
            # Crear figura
            fig = plt.figure(figsize=(16, 9))
            gs = fig.add_gridspec(1, 2, width_ratios=[1, 1.3], wspace=0.25)
            ax_pie = fig.add_subplot(gs[0])
            ax_table = fig.add_subplot(gs[1])
            
            fig.suptitle(f'Resumen Total - {len(viajes)} Bitácoras', fontsize=18, fontweight='bold', y=0.98)
            
            # Gráfico de torta LIMPIO con mejor separación visual
            if especies_data:
                sizes = [e['toneladas'] for e in especies_data]
                
                # Pequeño explode para todas las secciones (mejor separación visual)
                explode_vals = [0.05] * len(especies_data)
                
                wedges, texts = ax_pie.pie(
                    sizes, 
                    labels=None,
                    colors=colores_finales,
                    autopct=None,
                    explode=explode_vals,  # Separación ligera
                    shadow=False,
                    startangle=90,
                    wedgeprops={'edgecolor': 'white', 'linewidth': 3}  # Bordes más gruesos
                )
                
                ax_pie.set_title('Distribución de Especies\n(Solo Toneladas)', 
                               fontsize=14, fontweight='bold', pad=15)
                
                # AGREGAR TOOLTIPS INTERACTIVOS
                # Crear annotation que se mostrará al hover
                annot = ax_pie.annotate("", xy=(0,0), xytext=(20,20),
                                       textcoords="offset points",
                                       bbox=dict(boxstyle="round,pad=0.8", fc="#05BFDB", ec="#0492A8", lw=2, alpha=0.95),
                                       arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.3", 
                                                      color="#0492A8", lw=2),
                                       fontsize=11, fontweight='bold', color='white',
                                       visible=False, zorder=100)
                
                def on_hover(event):
                    if event.inaxes == ax_pie:
                        for i, wedge in enumerate(wedges):
                            cont, ind = wedge.contains(event)
                            if cont:
                                # Mostrar información de la especie
                                nombre = especies_data[i]['nombre']
                                toneladas = especies_data[i]['toneladas']
                                porcentaje = (toneladas / total_toneladas * 100) if total_toneladas > 0 else 0
                                tipo = 'RETENIDA' if especies_data[i]['tipo'] == 'retenida' else 'DESCARTE'
                                
                                # Calcular porcentaje respecto a su categoría
                                if especies_data[i]['tipo'] == 'retenida':
                                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
                                    texto_cat = "% del Retenido"
                                else:
                                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
                                    texto_cat = "% del Descarte"
                                
                                porc_categoria = (toneladas / total_categoria * 100) if total_categoria > 0 else 0
                                
                                # Actualizar texto del tooltip con info completa
                                annot.set_text(f"{nombre}\n{tipo}\n{toneladas:.3f} TON\n{porcentaje:.1f}% del Total\n{porc_categoria:.1f}% {texto_cat}")
                                annot.xy = (event.xdata, event.ydata)
                                annot.set_visible(True)
                                fig.canvas.draw_idle()
                                return
                        
                        # Si no está sobre ningún wedge, ocultar
                        if annot.get_visible():
                            annot.set_visible(False)
                            fig.canvas.draw_idle()
                
                # Conectar evento
                fig.canvas.mpl_connect("motion_notify_event", on_hover)
            
            # TABLA PROFESIONAL (igual que en individual)
            ax_table.axis('off')
            ax_table.set_xlim(0, 10)
            ax_table.set_ylim(0, 10)
            
            # Título
            rect_title = Rectangle((0.2, 9.2), 9.6, 0.7, 
                                   facecolor='#05BFDB', edgecolor='#0492A8', linewidth=2)
            ax_table.add_patch(rect_title)
            ax_table.text(5, 9.55, 'DETALLE DE ESPECIES POR TONELADAS', 
                         ha='center', va='center', fontsize=13, fontweight='bold', color='white')
            
            # Encabezados
            y_pos = 8.7
            rect_header = Rectangle((0.2, y_pos - 0.15), 9.6, 0.35, 
                                    facecolor='#e0e0e0', edgecolor='gray', linewidth=1)
            ax_table.add_patch(rect_header)
            
            ax_table.text(0.8, y_pos, '█', ha='center', fontsize=14, fontweight='bold')
            ax_table.text(2.2, y_pos, 'Especie', ha='left', fontsize=11, fontweight='bold')
            ax_table.text(5.8, y_pos, 'Tipo', ha='center', fontsize=11, fontweight='bold')
            ax_table.text(7.2, y_pos, 'TON', ha='right', fontsize=11, fontweight='bold')
            ax_table.text(8.5, y_pos, '% Total', ha='right', fontsize=11, fontweight='bold')
            ax_table.text(9.6, y_pos, '% Cat.', ha='right', fontsize=11, fontweight='bold')
            
            y_pos -= 0.35
            ax_table.plot([0.2, 9.8], [y_pos, y_pos], 'k-', linewidth=2)
            
            # Datos
            y_pos -= 0.45
            max_especies = 18
            
            for i, especie in enumerate(especies_data[:max_especies]):
                if i % 2 == 0:
                    rect_bg = Rectangle((0.2, y_pos - 0.18), 9.6, 0.35, 
                                       facecolor='#f5f5f5', edgecolor='none')
                    ax_table.add_patch(rect_bg)
                
                porcentaje = (especie['toneladas'] / total_toneladas * 100) if total_toneladas > 0 else 0
                
                # Calcular porcentaje respecto a su categoría
                if especie['tipo'] == 'retenida':
                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
                else:
                    total_categoria = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
                
                porc_categoria = (especie['toneladas'] / total_categoria * 100) if total_categoria > 0 else 0
                
                ax_table.text(0.8, y_pos, '█', ha='center', fontsize=16, 
                            color=colores_finales[i])
                
                nombre = especie['nombre'][:25]
                ax_table.text(2.2, y_pos, nombre, ha='left', fontsize=9,
                            weight='bold' if i < 5 else 'normal')
                
                tipo_texto = 'RETENIDA' if especie['tipo'] == 'retenida' else 'DESCARTE'
                tipo_color = '#00a000' if especie['tipo'] == 'retenida' else '#ff3300'
                ax_table.text(5.8, y_pos, tipo_texto, ha='center', fontsize=8,
                            color=tipo_color, weight='bold',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor=tipo_color, 
                                     alpha=0.2, edgecolor=tipo_color, linewidth=1))
                
                ax_table.text(7.2, y_pos, f"{especie['toneladas']:.3f}", 
                            ha='right', fontsize=9, weight='bold')
                
                ax_table.text(8.5, y_pos, f"{porcentaje:.1f}%", 
                            ha='right', fontsize=9, weight='bold', color='#0277bd')
                
                # Porcentaje de su categoría (NUEVA COLUMNA)
                ax_table.text(9.6, y_pos, f"{porc_categoria:.1f}%", 
                            ha='right', fontsize=9, weight='bold',
                            color=tipo_color)
                
                y_pos -= 0.4
            
            if len(especies_data) > max_especies:
                ax_table.text(5, y_pos, f'...y {len(especies_data) - max_especies} especies más', 
                            ha='center', fontsize=9, style='italic', color='gray')
                y_pos -= 0.4
            
            # Totales
            y_pos -= 0.3
            ax_table.plot([0.2, 9.8], [y_pos, y_pos], 'k-', linewidth=2.5)
            y_pos -= 0.5
            
            total_retenidas = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'retenida')
            total_descartadas = sum(e['toneladas'] for e in especies_data if e['tipo'] == 'descartada')
            
            rect_ret = Rectangle((0.2, y_pos - 0.2), 9.6, 0.4, 
                                facecolor='#e8f5e9', edgecolor='#00a000', linewidth=1.5)
            ax_table.add_patch(rect_ret)
            ax_table.text(2.5, y_pos, 'TOTAL RETENIDAS:', ha='left', fontsize=11, 
                        weight='bold', color='#00a000')
            ax_table.text(8, y_pos, f"{total_retenidas:.3f}", ha='right', 
                        fontsize=11, weight='bold', color='#00a000')
            porc_ret = (total_retenidas / total_toneladas * 100) if total_toneladas > 0 else 0
            ax_table.text(9.5, y_pos, f"{porc_ret:.1f}%", ha='right', 
                        fontsize=11, weight='bold', color='#00a000')
            
            y_pos -= 0.5
            
            rect_desc = Rectangle((0.2, y_pos - 0.2), 9.6, 0.4, 
                                  facecolor='#ffebee', edgecolor='#ff3300', linewidth=1.5)
            ax_table.add_patch(rect_desc)
            ax_table.text(2.5, y_pos, 'TOTAL DESCARTADAS:', ha='left', fontsize=11, 
                        weight='bold', color='#ff3300')
            ax_table.text(8, y_pos, f"{total_descartadas:.3f}", ha='right', 
                        fontsize=11, weight='bold', color='#ff3300')
            porc_desc = (total_descartadas / total_toneladas * 100) if total_toneladas > 0 else 0
            ax_table.text(9.5, y_pos, f"{porc_desc:.1f}%", ha='right', 
                        fontsize=11, weight='bold', color='#ff3300')
            
            y_pos -= 0.6
            
            rect_total = Rectangle((0.2, y_pos - 0.25), 9.6, 0.5, 
                                   facecolor='#e3f2fd', edgecolor='#0277bd', linewidth=2)
            ax_table.add_patch(rect_total)
            ax_table.text(2.5, y_pos, 'TOTAL GENERAL:', ha='left', fontsize=12, 
                        weight='bold', color='#0277bd')
            ax_table.text(8, y_pos, f"{total_toneladas:.3f} TON", ha='right', 
                        fontsize=12, weight='bold', color='#0277bd')
            ax_table.text(9.5, y_pos, "100%", ha='right', 
                        fontsize=12, weight='bold', color='#0277bd')
            
            plt.tight_layout()
            
            # Integrar en ventana
            canvas = FigureCanvasTkAgg(fig, master=ventana)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Botón cerrar
            btn_frame = tk.Frame(ventana, bg='white')
            btn_frame.pack(fill=tk.X, padx=10, pady=10)
            
            btn_cerrar = tk.Button(
                btn_frame,
                text="✖ Cerrar Gráfico",
                command=cerrar_ventana,
                bg="#D32F2F",
                fg="white",
                font=("Arial", 12, "bold"),
                padx=30,
                pady=10,
                relief=tk.RAISED,
                cursor="hand2"
            )
            btn_cerrar.pack(side=tk.RIGHT)
            
        except ImportError as e:
            from CTkMessagebox import CTkMessagebox
            CTkMessagebox(
                title="Librería no disponible",
                message=f"No se pudo cargar matplotlib:\n{e}",
                icon="warning"
            )
        except Exception as e:
            print(f"Error generando gráfico resumen: {e}")
            import traceback
            traceback.print_exc()
    
    def generar_mapa_calor(self, viajes, parent_window=None):
        """Genera un mapa de calor profesional con las ubicaciones de los lances"""
        try:
            import folium
            from folium.plugins import HeatMap, MiniMap
            import webbrowser
            import tempfile
            import os
            
            # Recolectar datos de lances
            lances_data = []
            
            for viaje in viajes:
                viaje_id = viaje.get('id_viaje')
                nave_nombre = viaje.get('nave_nombre', 'N/A')
                lances = self.firebase.obtener_lances_viaje(viaje_id)
                
                # Filtrar solo lances individuales (excluir lance 0 CAPTURA TOTAL)
                for lance in lances:
                    if lance.get('es_captura_total', False):
                        continue
                    
                    # Obtener coordenadas
                    lat_inicio = lance.get('latitud_inicio')
                    lon_inicio = lance.get('longitud_inicio')
                    lat_fin = lance.get('latitud_fin')
                    lon_fin = lance.get('longitud_fin')
                    
                    if lat_inicio is None or lon_inicio is None:
                        continue
                    
                    # Separar especies por tipo
                    especies_retenidas = []
                    especies_descartadas_ton = []
                    especies_descartadas_unidades = {}
                    especies_incidentales = {}  # {nombre: unidades}
                    total_objetivo = 0
                    total_retenida = 0
                    total_descarte = 0
                    total_incidental = 0
                    
                    for especie in lance.get('especies', []):
                        nombre = especie.get('nombre', '')
                        cantidad_ton = especie.get('cantidad_ton', 0)
                        cantidad_unidades = especie.get('cantidad_unidades', 0)
                        tipo_captura = especie.get('tipo_captura', 'retenida')
                        
                        if tipo_captura == 'retenida':
                            if cantidad_ton > 0:
                                especies_retenidas.append({
                                    'nombre': nombre,
                                    'ton': cantidad_ton
                                })
                                total_retenida += cantidad_ton
                                # Verificar si es objetivo
                                es_objetivo = any(obj.lower() in nombre.lower() 
                                                for obj in ['camarón', 'camaron', 'langostino', 'gamba'])
                                if es_objetivo:
                                    total_objetivo += cantidad_ton
                        
                        elif tipo_captura == 'descartada':
                            # NO mezclar toneladas y unidades - priorizar toneladas
                            if cantidad_ton > 0:
                                total_descarte += cantidad_ton
                                especies_descartadas_ton.append({
                                    'nombre': nombre,
                                    'ton': cantidad_ton
                                })
                            elif cantidad_unidades > 0:
                                # Solo mostrar unidades si NO hay toneladas
                                especies_descartadas_unidades[nombre] = cantidad_unidades
                        
                        elif tipo_captura == 'incidental':
                            if cantidad_unidades > 0:
                                especies_incidentales[nombre] = especies_incidentales.get(nombre, 0) + cantidad_unidades
                                total_incidental += cantidad_unidades
                    
                    # Incluir TODOS los lances (incluso sin capturas)
                    lances_data.append({
                        'viaje_id': viaje_id,
                        'nave': nave_nombre,
                        'num_lance': lance.get('numero_lance'),
                        'lat': lat_inicio,
                        'lon': lon_inicio,
                        'lat_fin': lat_fin,
                        'lon_fin': lon_fin,
                        'total_objetivo': total_objetivo,
                        'total_retenida': total_retenida,
                        'total_descarte': total_descarte,
                        'total_incidental': total_incidental,
                        'especies_retenidas': especies_retenidas,
                        'especies_descartadas_ton': especies_descartadas_ton,
                        'especies_descartadas_unidades': especies_descartadas_unidades,
                        'especies_incidentales': especies_incidentales,
                        'arte_pesca': lance.get('arte_pesca', 'N/A'),
                        'fecha_inicio': lance.get('fecha_inicio', 'N/A'),
                        'observaciones': lance.get('observaciones', ''),
                        'sin_capturas': total_retenida == 0 and total_descarte == 0 and total_incidental == 0
                    })
            
            if not lances_data:
                CTkMessagebox(
                    title="Sin datos de ubicación",
                    message="No se encontraron lances con coordenadas GPS válidas.",
                    icon="warning"
                )
                return
            
            # Calcular centro y límites del mapa
            lat_centro = sum(l['lat'] for l in lances_data) / len(lances_data)
            lon_centro = sum(l['lon'] for l in lances_data) / len(lances_data)
            
            lat_min = min(l['lat'] for l in lances_data)
            lat_max = max(l['lat'] for l in lances_data)
            lon_min = min(l['lon'] for l in lances_data)
            lon_max = max(l['lon'] for l in lances_data)
            
            # Crear mapa con estilo profesional CLARO
            mapa = folium.Map(
                location=[lat_centro, lon_centro],
                zoom_start=9,
                tiles='CartoDB Voyager',  # Tema CLARO profesional
                control_scale=True  # Agregar escala
            )
            
            # Agregar capas de tiles adicionales CLARAS
            folium.TileLayer('OpenStreetMap', name='OpenStreetMap').add_to(mapa)
            folium.TileLayer('CartoDB positron', name='Claro').add_to(mapa)
            folium.TileLayer(
                tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Topo_Map/MapServer/tile/{z}/{y}/{x}',
                attr='Esri',
                name='World Topo Map',
                overlay=False,
                control=True
            ).add_to(mapa)
            
            # Preparar datos para heat maps con mejor contraste
            coords_objetivo = [[l['lat'], l['lon'], l['total_objetivo'] * 8] 
                              for l in lances_data if l['total_objetivo'] > 0]
            coords_descarte = [[l['lat'], l['lon'], l['total_descarte'] * 8] 
                              for l in lances_data if l['total_descarte'] > 0]
            
            # Heat map para especies objetivo (gradiente verde brillante)
            if coords_objetivo:
                HeatMap(
                    coords_objetivo,
                    name='🎯 Heat Map: Captura Objetivo',
                    gradient={
                        0.0: '#00ff00',   # Verde brillante
                        0.3: '#00cc00',   # Verde medio
                        0.6: '#009900',   # Verde oscuro
                        1.0: '#006600'    # Verde muy oscuro
                    },
                    min_opacity=0.4,
                    max_opacity=0.7,
                    radius=22,            # Aumentado para mejor visualización al zoom
                    blur=18,              # Ajustado para balance
                    overlay=True,
                    control=True,
                    show=True
                ).add_to(mapa)
            
            # Heat map para descarte (MISMO TAMAÑO que objetivo para escalar igual)
            if coords_descarte:
                HeatMap(
                    coords_descarte,
                    name='🗑️ Heat Map: Descarte',
                    gradient={
                        0.0: '#ffff00',   # Amarillo brillante
                        0.3: '#ffaa00',   # Naranja claro
                        0.6: '#ff4400',   # Rojo-naranja intenso
                        1.0: '#ff0000'    # Rojo puro (más visible)
                    },
                    min_opacity=0.5,      # Mayor visibilidad
                    max_opacity=0.8,      # Mayor visibilidad
                    radius=22,            # MISMO que objetivo para escalar igual al zoom
                    blur=18,              # MISMO que objetivo
                    overlay=True,
                    control=True,
                    show=True
                ).add_to(mapa)
            
            # Agregar marcadores detallados con círculos proporcionales
            for lance_data in lances_data:
                lat = lance_data['lat']
                lon = lance_data['lon']
                total_objetivo = lance_data['total_objetivo']
                total_retenida = lance_data['total_retenida']
                total_descarte = lance_data['total_descarte']
                total_captura = total_retenida + total_descarte
                
                # Determinar color según tipo predominante (mejor contraste)
                if lance_data.get('sin_capturas'):
                    color_marker = 'gray'
                    color_circulo = '#888888'  # Gris
                    color_borde = '#555555'
                elif total_retenida > total_descarte:
                    color_marker = 'green'
                    color_circulo = '#00cc00'  # Verde brillante
                    color_borde = '#006600'
                elif total_descarte > 0:
                    color_marker = 'red'
                    color_circulo = '#ff3300'  # Rojo brillante
                    color_borde = '#990000'
                else:
                    color_marker = 'gray'
                    color_circulo = '#888888'
                    color_borde = '#555555'
                
                # Formatear fecha
                fecha_str = lance_data['fecha_inicio']
                if fecha_str and fecha_str != 'N/A':
                    try:
                        from datetime import datetime
                        fecha_obj = datetime.fromisoformat(fecha_str)
                        fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
                    except:
                        pass
                
                # Crear popup PROFESIONAL con toda la información
                popup_html = f"""
                <div style="font-family: 'Segoe UI', Arial, sans-serif; width: 350px; max-height: 450px; overflow-y: auto;">
                    <div style="background: linear-gradient(135deg, #05BFDB 0%, #0492A8 100%); 
                                color: white; padding: 12px; margin: -10px -10px 10px -10px; 
                                border-radius: 5px 5px 0 0;">
                        <h3 style="margin: 0; font-size: 16px; font-weight: 600;">
                            ⚓ LANCE #{lance_data['num_lance']}
                        </h3>
                        <p style="margin: 5px 0 0 0; font-size: 11px; opacity: 0.9;">
                            {lance_data['viaje_id']} | {lance_data['nave']}
                        </p>
                    </div>
                    
                    <div style="background: #f8f9fa; padding: 8px; border-radius: 4px; margin-bottom: 8px;">
                        <p style="margin: 3px 0; font-size: 11px; color: #666;">
                            <b>📅 Fecha:</b> {fecha_str}
                        </p>
                        <p style="margin: 3px 0; font-size: 11px; color: #666;">
                            <b>🎣 Arte:</b> {lance_data['arte_pesca']}
                        </p>
                        <p style="margin: 3px 0; font-size: 11px; color: #666;">
                            <b>📍 Inicio:</b> {lat:.5f}°S, {-abs(lon):.5f}°W
                        </p>
                """
                
                if lance_data['lat_fin'] and lance_data['lon_fin']:
                    popup_html += f"""
                        <p style="margin: 3px 0; font-size: 11px; color: #666;">
                            <b>📍 Fin:</b> {lance_data['lat_fin']:.5f}°S, {-abs(lance_data['lon_fin']):.5f}°W
                        </p>
                    """
                
                popup_html += "</div>"
                
                # CAPTURA RETENIDA
                popup_html += f"""
                    <div style="border-left: 4px solid #00cc00; padding-left: 8px; margin: 10px 0;">
                        <p style="margin: 5px 0; font-weight: bold; color: #00cc00; font-size: 13px;">
                            🎯 CAPTURA RETENIDA: {total_retenida:.3f} TON
                        </p>
                """
                
                if lance_data['especies_retenidas']:
                    popup_html += "<table style='width: 100%; font-size: 11px; margin-top: 5px; border-collapse: collapse;'>"
                    popup_html += "<tr style='background: #e8f5e9; font-weight: bold;'>"
                    popup_html += "<td style='padding: 4px; border-bottom: 1px solid #ddd;'>Especie</td>"
                    popup_html += "<td style='padding: 4px; text-align: right; border-bottom: 1px solid #ddd;'>TON</td>"
                    popup_html += "<td style='padding: 4px; text-align: right; border-bottom: 1px solid #ddd;'>%</td>"
                    popup_html += "</tr>"
                    
                    for esp in lance_data['especies_retenidas']:
                        # Calcular porcentaje respecto al total de captura retenida
                        porcentaje_retenida = (esp['ton'] / total_retenida * 100) if total_retenida > 0 else 0
                        
                        popup_html += f"""
                            <tr style='border-bottom: 1px solid #eee;'>
                                <td style='padding: 3px;'>{esp['nombre']}</td>
                                <td style='padding: 3px; text-align: right; font-weight: bold; color: #00cc00;'>{esp['ton']:.3f}</td>
                                <td style='padding: 3px; text-align: right; color: #006600; font-size: 10px;'>{porcentaje_retenida:.1f}%</td>
                            </tr>
                        """
                    popup_html += "</table>"
                    
                    # Agregar subtotal con porcentaje del total general
                    porcentaje_retenida_total = (total_retenida / total_captura * 100) if total_captura > 0 else 0
                    popup_html += f"<p style='margin: 5px 0; font-size: 10px; color: #00cc00; text-align: right; font-style: italic;'><b>{porcentaje_retenida_total:.1f}% del total del lance</b></p>"
                else:
                    if lance_data.get('sin_capturas'):
                        obs_text = lance_data.get('observaciones', '')
                        popup_html += f"<p style='font-size: 11px; color: #888; margin: 5px 0;'><b>LANCE DECLARADO SIN CAPTURAS</b></p>"
                        if obs_text:
                            popup_html += f"<p style='font-size: 10px; color: #666; margin: 3px 0;'><i>📝 {obs_text}</i></p>"
                    else:
                        popup_html += "<p style='font-size: 10px; color: #999; margin: 5px 0;'><i>Sin captura retenida</i></p>"
                
                popup_html += "</div>"
                
                # DESCARTE (mostrar toneladas O unidades, NO ambas)
                popup_html += f"""
                    <div style="border-left: 4px solid #ff3300; padding-left: 8px; margin: 10px 0;">
                        <p style="margin: 5px 0; font-weight: bold; color: #ff3300; font-size: 13px;">
                            🗑️ DESCARTE: {total_descarte:.3f} TON
                        </p>
                """
                
                if lance_data['especies_descartadas_ton'] or lance_data['especies_descartadas_unidades']:
                    popup_html += "<table style='width: 100%; font-size: 11px; margin-top: 5px; border-collapse: collapse;'>"
                    popup_html += "<tr style='background: #ffebee; font-weight: bold;'>"
                    popup_html += "<td style='padding: 4px; border-bottom: 1px solid #ddd;'>Especie</td>"
                    popup_html += "<td style='padding: 4px; text-align: right; border-bottom: 1px solid #ddd;'>Cantidad</td>"
                    popup_html += "<td style='padding: 4px; text-align: right; border-bottom: 1px solid #ddd;'>%</td>"
                    popup_html += "</tr>"
                    
                    # Primero mostrar especies con TONELADAS
                    especies_ton_nombres = set()
                    for esp in lance_data['especies_descartadas_ton']:
                        nombre = esp['nombre']
                        toneladas = esp['ton']
                        especies_ton_nombres.add(nombre)
                        
                        # Calcular porcentaje respecto al total de descarte
                        porcentaje_descarte = (toneladas / total_descarte * 100) if total_descarte > 0 else 0
                        # Calcular porcentaje respecto al total general del lance
                        porcentaje_total = (toneladas / total_captura * 100) if total_captura > 0 else 0
                        
                        popup_html += f"""
                            <tr style='border-bottom: 1px solid #eee;'>
                                <td style='padding: 3px; font-size: 10px;'>{nombre}</td>
                                <td style='padding: 3px; text-align: right; font-weight: bold; color: #ff3300;'>{toneladas:.3f} TON</td>
                                <td style='padding: 3px; text-align: right; color: #cc0000; font-size: 10px;'>{porcentaje_descarte:.1f}%</td>
                            </tr>
                        """
                    
                    # Luego mostrar especies SOLO con UNIDADES (que NO tengan toneladas)
                    for nombre, unidades in lance_data['especies_descartadas_unidades'].items():
                        if nombre not in especies_ton_nombres:  # Solo si NO está en toneladas
                            popup_html += f"""
                                <tr style='border-bottom: 1px solid #eee;'>
                                    <td style='padding: 3px; font-size: 10px;'>{nombre}</td>
                                    <td style='padding: 3px; text-align: right; color: #666; font-size: 10px;'>{int(unidades)} unidades</td>
                                    <td style='padding: 3px; text-align: right; color: #999; font-size: 9px;'>-</td>
                                </tr>
                            """
                    
                    popup_html += "</table>"
                    
                    # Agregar subtotal con porcentaje del total general
                    if total_descarte > 0:
                        porcentaje_descarte_total = (total_descarte / total_captura * 100) if total_captura > 0 else 0
                        popup_html += f"<p style='margin: 5px 0; font-size: 10px; color: #ff3300; text-align: right; font-style: italic;'><b>{porcentaje_descarte_total:.1f}% del total del lance</b></p>"
                else:
                    popup_html += "<p style='font-size: 10px; color: #999; margin: 5px 0;'><i>Sin descarte</i></p>"
                
                popup_html += "</div>"
                
                # INCIDENTAL (fauna acompañante)
                total_incidental = lance_data.get('total_incidental', 0)
                if total_incidental > 0:
                    popup_html += f"""
                    <div style="border-left: 4px solid #9C27B0; padding-left: 8px; margin: 10px 0;">
                        <p style="margin: 5px 0; font-weight: bold; color: #9C27B0; font-size: 13px;">
                            🦭 INCIDENTAL: {total_incidental} individuo(s)
                        </p>
                    """
                    popup_html += "<table style='width: 100%; font-size: 11px; margin-top: 5px; border-collapse: collapse;'>"
                    popup_html += "<tr style='background: #f3e5f5; font-weight: bold;'>"
                    popup_html += "<td style='padding: 4px; border-bottom: 1px solid #ddd;'>Especie</td>"
                    popup_html += "<td style='padding: 4px; text-align: right; border-bottom: 1px solid #ddd;'>Individuos</td>"
                    popup_html += "</tr>"
                    
                    for nombre_inc, unidades_inc in lance_data['especies_incidentales'].items():
                        popup_html += f"""
                            <tr style='border-bottom: 1px solid #eee;'>
                                <td style='padding: 3px; font-size: 10px;'>{nombre_inc}</td>
                                <td style='padding: 3px; text-align: right; font-weight: bold; color: #7B1FA2;'>{int(unidades_inc)} ind.</td>
                            </tr>
                        """
                    
                    popup_html += "</table></div>"
                
                # Total con distribución de porcentajes
                porcentaje_retenida_final = (total_retenida / total_captura * 100) if total_captura > 0 else 0
                porcentaje_descarte_final = (total_descarte / total_captura * 100) if total_captura > 0 else 0
                
                if lance_data.get('sin_capturas'):
                    popup_html += f"""
                    <div style="background: #f5f5f5; padding: 10px; border-radius: 4px; margin-top: 10px;">
                        <p style="margin: 0; font-weight: bold; color: #888; font-size: 12px; text-align: center;">
                            ⚠️ LANCE SIN CAPTURAS
                        </p>
                    </div>
                </div>
                """
                else:
                    popup_html += f"""
                    <div style="background: #e3f2fd; padding: 10px; border-radius: 4px; margin-top: 10px;">
                        <p style="margin: 0 0 6px 0; font-weight: bold; color: #0277bd; font-size: 12px; text-align: center;">
                            ⚖️ TOTAL CAPTURA: {total_captura:.3f} TON
                        </p>
                        <div style="display: flex; justify-content: space-around; font-size: 10px; margin-top: 6px;">
                            <span style="color: #00cc00;">
                                🎯 <b>{porcentaje_retenida_final:.1f}%</b> Retenido
                            </span>
                            <span style="color: #666;">|</span>
                            <span style="color: #ff3300;">
                                🗑️ <b>{porcentaje_descarte_final:.1f}%</b> Descarte
                            </span>
                        </div>
                    </div>
                </div>
                """
                
                # Círculo proporcional mejorado con tooltip profesional
                radio_metros = min(max(total_captura * 400, 150), 2000)
                
                # Tooltip más informativo con porcentajes
                porcentaje_retenida_tooltip = (total_retenida / total_captura * 100) if total_captura > 0 else 0
                porcentaje_descarte_tooltip = (total_descarte / total_captura * 100) if total_captura > 0 else 0
                
                if lance_data.get('sin_capturas'):
                    obs_tooltip = lance_data.get('observaciones', '')
                    tooltip_html = f"""
                    <div style='font-family: Arial; text-align: center;'>
                        <b style='font-size: 13px; color: #888;'>🎣 LANCE #{lance_data['num_lance']}</b><br>
                        <span style='font-size: 11px;'>{lance_data['viaje_id']}</span><br>
                        <hr style='margin: 4px 0; border: none; border-top: 1px solid #ddd;'>
                        <span style='font-size: 10px; color: #888;'>
                            ⚠️ <b>Sin capturas</b><br>
                            {'📝 ' + obs_tooltip[:50] + '...' if obs_tooltip else ''}
                        </span>
                    </div>
                    """
                else:
                    incidental_tooltip_line = ""
                    if total_incidental > 0:
                        incidental_tooltip_line = f"🦭 Incidental: <b style='color: #9C27B0;'>{int(total_incidental)} ind.</b><br>"
                    tooltip_html = f"""
                    <div style='font-family: Arial; text-align: center;'>
                        <b style='font-size: 13px; color: {color_circulo};'>🎣 LANCE #{lance_data['num_lance']}</b><br>
                        <span style='font-size: 11px;'>{lance_data['viaje_id']}</span><br>
                        <hr style='margin: 4px 0; border: none; border-top: 1px solid #ddd;'>
                        <span style='font-size: 10px;'>
                            🎯 Retenido: <b style='color: #00cc00;'>{total_retenida:.2f}T ({porcentaje_retenida_tooltip:.1f}%)</b><br>
                            🗑️ Descarte: <b style='color: #ff3300;'>{total_descarte:.3f}T ({porcentaje_descarte_tooltip:.1f}%)</b><br>
                            {incidental_tooltip_line}
                            ⚖️ <b>Total: {total_captura:.2f}T</b>
                        </span>
                    </div>
                    """
                
                folium.Circle(
                    location=[lat, lon],
                    radius=radio_metros,
                    color=color_borde,
                    fill=True,
                    fillColor=color_circulo,
                    fillOpacity=0.5,
                    weight=3,
                    popup=folium.Popup(popup_html, max_width=370),
                    tooltip=folium.Tooltip(tooltip_html, sticky=True)
                ).add_to(mapa)
                
                # Marcador central mejorado con tooltip profesional
                icon_html = f"""
                <div style="background-color: {color_circulo}; 
                            border: 3px solid {color_borde}; 
                            border-radius: 50%; 
                            width: 26px; 
                            height: 26px; 
                            display: flex; 
                            align-items: center; 
                            justify-content: center;
                            font-weight: bold;
                            color: white;
                            font-size: 12px;
                            box-shadow: 0 3px 10px rgba(0,0,0,0.5);
                            cursor: pointer;
                            transition: transform 0.2s;">
                    {lance_data['num_lance']}
                </div>
                """
                
                # Tooltip del marcador igual que del círculo
                tooltip_marcador = tooltip_html
                
                folium.Marker(
                    location=[lat, lon],
                    popup=folium.Popup(popup_html, max_width=370),
                    icon=folium.DivIcon(html=icon_html),
                    tooltip=folium.Tooltip(tooltip_marcador, sticky=True)
                ).add_to(mapa)
            
            # Calcular rangos de coordenadas para el panel
            latitudes = [l['lat'] for l in lances_data]
            longitudes = [l['lon'] for l in lances_data]
            lat_min, lat_max = min(latitudes), max(latitudes)
            lon_min, lon_max = min(longitudes), max(longitudes)
            
            # Agregar control de capas
            folium.LayerControl().add_to(mapa)
            
            # Agregar mini-mapa para navegación profesional
            from folium.plugins import MiniMap, Fullscreen
            minimap = MiniMap(
                toggle_display=True,
                tile_layer='CartoDB Voyager',
                width=150,
                height=150,
                zoom_level_offset=-5,
                position='bottomleft'
            )
            mapa.add_child(minimap)
            
            # Agregar botón de pantalla completa para visualización profesional
            Fullscreen(
                position='topleft',
                title='Pantalla Completa',
                title_cancel='Salir de Pantalla Completa',
                force_separate_button=True
            ).add_to(mapa)
            
            # Agregar leyenda profesional personalizada
            leyenda_html = f'''
            <div style="position: fixed; 
                        top: 10px; 
                        right: 10px; 
                        width: 220px; 
                        background-color: white; 
                        border: 3px solid #05BFDB;
                        border-radius: 8px;
                        padding: 12px;
                        font-family: 'Segoe UI', Arial, sans-serif;
                        box-shadow: 0 4px 12px rgba(0,0,0,0.25);
                        z-index: 9999;">
                
                <h4 style="margin: 0 0 12px 0; 
                           padding-bottom: 8px; 
                           border-bottom: 2px solid #05BFDB; 
                           color: #0492A8;
                           font-size: 14px;
                           text-align: center;">
                    📊 MAPA DE CALOR - PESCA
                </h4>
                
                <div style="margin: 10px 0;">
                    <p style="margin: 0 0 8px 0; font-weight: bold; font-size: 12px; color: #333;">
                        🎯 Captura Objetivo:
                    </p>
                    <div style="background: linear-gradient(to right, #00ff00, #00cc00, #009900, #006600); 
                                height: 20px; 
                                border-radius: 4px; 
                                border: 1px solid #006600;
                                margin-bottom: 3px;">
                    </div>
                    <p style="margin: 0; font-size: 9px; color: #666; display: flex; justify-content: space-between;">
                        <span>Baja</span><span>Alta</span>
                    </p>
                </div>
                
                <div style="margin: 10px 0;">
                    <p style="margin: 0 0 8px 0; font-weight: bold; font-size: 12px; color: #333;">
                        🗑️ Descarte:
                    </p>
                    <div style="background: linear-gradient(to right, #ffff00, #ffaa00, #ff4400, #ff0000); 
                                height: 20px; 
                                border-radius: 4px; 
                                border: 1px solid #ff0000;
                                margin-bottom: 3px;">
                    </div>
                    <p style="margin: 0; font-size: 9px; color: #666; display: flex; justify-content: space-between;">
                        <span>Baja</span><span>Alta</span>
                    </p>
                </div>
                
                <hr style="margin: 12px 0; border: none; border-top: 1px solid #ddd;">
                
                <div style="font-size: 11px; color: #666;">
                    <p style="margin: 5px 0;">
                        <span style="display: inline-block; 
                                     width: 16px; 
                                     height: 16px; 
                                     background-color: #00cc00; 
                                     border: 2px solid #006600; 
                                     border-radius: 50%; 
                                     vertical-align: middle; 
                                     margin-right: 6px;">
                        </span>
                        Predomina objetivo
                    </p>
                    <p style="margin: 5px 0;">
                        <span style="display: inline-block; 
                                     width: 16px; 
                                     height: 16px; 
                                     background-color: #ff3300; 
                                     border: 2px solid #990000; 
                                     border-radius: 50%; 
                                     vertical-align: middle; 
                                     margin-right: 6px;">
                        </span>
                        Predomina descarte
                    </p>
                    <p style="margin: 5px 0;">
                        <span style="display: inline-block; 
                                     width: 16px; 
                                     height: 16px; 
                                     background-color: #9C27B0; 
                                     border: 2px solid #6A1B9A; 
                                     border-radius: 50%; 
                                     vertical-align: middle; 
                                     margin-right: 6px;">
                        </span>
                        Captura incidental
                    </p>
                    <p style="margin: 5px 0;">
                        <span style="display: inline-block; 
                                     width: 16px; 
                                     height: 16px; 
                                     background-color: #888888; 
                                     border: 2px solid #555555; 
                                     border-radius: 50%; 
                                     vertical-align: middle; 
                                     margin-right: 6px;">
                        </span>
                        Sin capturas
                    </p>
                </div>
                
                <hr style="margin: 12px 0; border: none; border-top: 1px solid #ddd;">
                
                <p style="margin: 5px 0; font-size: 10px; color: #666; text-align: center;">
                    <b>Total Lances:</b> {len(lances_data)}<br>
                    <b>Viajes:</b> {len(viajes)}<br>
                    <b>Capturas:</b> {sum(l['total_retenida'] + l['total_descarte'] for l in lances_data):.2f} TON<br>
                    <b>Incidentales:</b> {int(sum(l.get('total_incidental', 0) for l in lances_data))} ind.
                </p>
                
                <hr style="margin: 8px 0; border: none; border-top: 1px solid #ddd;">
                
                <p style="margin: 5px 0; font-size: 9px; color: #666; text-align: center; line-height: 1.4;">
                    <i>• Círculo = Captura total<br>
                    • Click en marcador para detalles<br>
                    • Use capas para filtrar</i>
                </p>
                
                <p style="margin: 8px 0 0 0; 
                          font-size: 8px; 
                          color: #999; 
                          text-align: center; 
                          font-style: italic;
                          padding-top: 6px;
                          border-top: 1px solid #eee;">
                    🎣 Sistema BENTOS TI<br>
                    Análisis Pesquero Profesional
                </p>
            </div>
            '''
            mapa.get_root().html.add_child(folium.Element(leyenda_html))
            
            # Guardar mapa en archivo temporal y abrirlo en navegador
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.html', mode='w', encoding='utf-8')
            temp_file.close()
            self._temp_files.append(temp_file.name)
            
            mapa.save(temp_file.name)
            
            # Abrir en navegador
            webbrowser.open('file://' + os.path.abspath(temp_file.name))
            
            # No mostrar mensaje de confirmación
            
        except ImportError:
            CTkMessagebox(
                title="Librería no disponible",
                message="Para generar mapas de calor necesitas instalar folium:\npip install folium",
                icon="warning"
            )
        except Exception as e:
            print(f"Error generando mapa de calor: {e}")
            import traceback
            traceback.print_exc()
            CTkMessagebox(
                title="Error",
                message=f"Error al generar mapa de calor:\n{str(e)}",
                icon="cancel"
            )
    
    def select_multiple_pdfs(self):
        """Selecciona múltiples PDFs para procesar"""
        filenames = filedialog.askopenfilenames(
            title="Seleccionar Bitácoras PDF (hasta 5)",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if filenames:
            # Limitar a 5 archivos
            if len(filenames) > 5:
                CTkMessagebox(
                    title="⚠️ Límite de archivos",
                    message=f"Solo puedes seleccionar hasta 5 archivos.\nSe procesarán los primeros 5.",
                    icon="warning",
                    option_1="OK"
                )
                filenames = filenames[:5]
            
            self.archivos_pendientes = list(filenames)
            self.mostrar_archivos_seleccionados()
    
    def drop_files(self, event):
        """Maneja el evento de drag & drop con efecto visual"""
        # Efecto visual de recepción
        try:
            self.drop_frame.configure(border_color="#4CAF50", fg_color=("#D4EDDA", "#1A4A2A"))
            self.root.after(600, lambda: self.drop_frame.configure(
                border_color="#05BFDB", fg_color=("#E8F4F8", "#1A3A4A")))
        except (tk.TclError, Exception):
            pass
        try:
            # Obtener archivos del evento - manejar diferentes formatos
            data = event.data
            
            # Parsear paths correctamente (tkinterdnd2 usa {} para paths con espacios)
            files = []
            if isinstance(data, str):
                data = data.strip()
                i = 0
                while i < len(data):
                    if data[i] == '{':
                        # Path entre llaves (contiene espacios)
                        end = data.index('}', i)
                        files.append(data[i+1:end])
                        i = end + 2  # saltar } y espacio
                    elif data[i] == ' ':
                        i += 1
                    else:
                        # Path sin espacios, termina en siguiente espacio
                        end = data.find(' ', i)
                        if end == -1:
                            files.append(data[i:])
                            break
                        else:
                            files.append(data[i:end])
                            i = end + 1
            else:
                files = [str(data)]
            
            # Filtrar solo PDFs y limpiar paths
            pdf_files = []
            for f in files:
                # Quitar comillas si las tiene
                f = f.strip('"').strip("'")
                if f.lower().endswith('.pdf') and os.path.exists(f):
                    pdf_files.append(f)
            
            if not pdf_files:
                CTkMessagebox(
                    title="⚠️ Formato incorrecto",
                    message="Solo se aceptan archivos PDF válidos",
                    icon="warning",
                    option_1="OK"
                )
                return
            
            # Limitar a 5 archivos
            if len(pdf_files) > 5:
                CTkMessagebox(
                    title="⚠️ Límite de archivos",
                    message=f"Solo puedes procesar hasta 5 archivos a la vez.\nSe procesarán los primeros 5.",
                    icon="warning",
                    option_1="OK"
                )
                pdf_files = pdf_files[:5]
            
            self.archivos_pendientes = pdf_files
            self.mostrar_archivos_seleccionados()
            
            print(f"✓ {len(pdf_files)} archivo(s) agregado(s) por drag & drop")
            
        except Exception as e:
            print(f"Error en drag & drop: {e}")
            import traceback
            traceback.print_exc()
    
    def mostrar_archivos_seleccionados(self):
        """Muestra la lista de archivos seleccionados"""
        # Limpiar lista
        for widget in self.files_list_frame.winfo_children():
            widget.destroy()
        
        if not self.archivos_pendientes:
            self.files_list_frame.pack_forget()
            return
        
        # Mostrar frame de lista
        self.files_list_frame.pack(pady=15, fill="x")
        
        # Título
        ctk.CTkLabel(
            self.files_list_frame,
            text=f"📋 Archivos seleccionados ({len(self.archivos_pendientes)}):",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(anchor="w", padx=10, pady=5)
        
        # Lista de archivos
        for i, archivo in enumerate(self.archivos_pendientes, 1):
            nombre = os.path.basename(archivo)
            
            file_frame = ctk.CTkFrame(self.files_list_frame, fg_color=("#FFFFFF", "#1A1A1A"))
            file_frame.pack(fill="x", padx=10, pady=3)
            
            ctk.CTkLabel(
                file_frame,
                text=f"{i}. {nombre}",
                font=ctk.CTkFont(size=11),
                anchor="w"
            ).pack(side="left", padx=10, pady=5, fill="x", expand=True)
            
            # Botón para remover
            ctk.CTkButton(
                file_frame,
                text="✕",
                width=30,
                height=25,
                fg_color="#D32F2F",
                hover_color="#B71C1C",
                command=lambda idx=i-1: self.remover_archivo(idx)
            ).pack(side="right", padx=5)
        
        # Botón para procesar todos
        ctk.CTkButton(
            self.files_list_frame,
            text=f"✅ Procesar {len(self.archivos_pendientes)} archivo(s)",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40,
            fg_color="#4CAF50",
            hover_color="#388E3C",
            command=self.procesar_archivos_multiples
        ).pack(pady=10)
    
    def remover_archivo(self, index):
        """Remueve un archivo de la lista"""
        if 0 <= index < len(self.archivos_pendientes):
            self.archivos_pendientes.pop(index)
            self.mostrar_archivos_seleccionados()
    
    def procesar_archivos_multiples(self):
        """Procesa múltiples archivos con confirmación previa"""
        if not self.archivos_pendientes:
            return
        
        # Verificar conexión a internet antes de procesar
        if self._sin_internet_alerta("subir bitácoras a la nube"):
            return
        
        # Procesar TODOS los archivos para obtener información
        try:
            # Mostrar loading
            self.upload_btn.configure(state="disabled", text="⏳ Analizando...")
            self.root.update()
            
            # Parsear todos los PDFs y guardar resultados
            todos_resultados = []
            for i, archivo in enumerate(self.archivos_pendientes):
                nombre = os.path.basename(archivo)
                self.upload_btn.configure(text=f"⏳ Analizando {i+1}/{len(self.archivos_pendientes)}...")
                self.root.update()
                
                try:
                    with BitacoraParser(archivo) as parser:
                        resultado = parser.parsear_completo()
                    self.resultados_parseados[archivo] = resultado
                    todos_resultados.append((archivo, resultado, None))
                except Exception as e:
                    todos_resultados.append((archivo, None, str(e)))
            
            # Mostrar diálogo de confirmación con TODOS los resultados
            self.mostrar_confirmacion_subida(todos_resultados, self.archivos_pendientes)
            
        except Exception as e:
            CTkMessagebox(
                title="❌ Error",
                message=f"Error al analizar los archivos:\n\n{str(e)}",
                icon="cancel",
                option_1="OK"
            )
        finally:
            self.upload_btn.configure(state="normal", text="📁 Seleccionar Archivos PDF")
    
    def mostrar_confirmacion_subida(self, todos_resultados, archivos):
        """Muestra diálogo de confirmación antes de subir con info de TODOS los PDFs"""
        # Crear ventana de confirmación
        confirm_window = ctk.CTkToplevel(self.root)
        confirm_window.title("Confirmar Subida de Bitácoras")
        confirm_window.attributes('-alpha', 0.0)
        
        # Ajustar tamaño según pantalla disponible (dejar margen para barra de tareas)
        screen_w = confirm_window.winfo_screenwidth()
        screen_h = confirm_window.winfo_screenheight()
        win_w = min(780, screen_w - 40)
        win_h = min(screen_h - 100, 750)  # Margen de 100px para barra de tareas
        
        confirm_window.geometry(f"{win_w}x{win_h}")
        confirm_window.transient(self.root)
        confirm_window.grab_set()
        
        # Centrar ventana (arriba del centro para evitar barra de tareas)
        confirm_window.update_idletasks()
        x = (screen_w - win_w) // 2
        y = max(10, (screen_h - win_h) // 2 - 40)
        confirm_window.geometry(f"+{x}+{y}")
        
        # Fade-in
        self._fade_in_widget(confirm_window)
        
        # Frame principal con layout grid para que botones siempre sean visibles
        confirm_window.grid_rowconfigure(2, weight=1)  # El scrollable frame se expande
        confirm_window.grid_columnconfigure(0, weight=1)
        
        # Título
        header_frame = ctk.CTkFrame(confirm_window, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(10, 0))
        
        ctk.CTkLabel(
            header_frame,
            text="⚠️ CONFIRMAR INFORMACIÓN",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="#FF9800"
        ).pack(pady=(5, 2))
        
        ctk.CTkLabel(
            header_frame,
            text=f"Estás a punto de subir {len(archivos)} bitácora(s) al sistema",
            font=ctk.CTkFont(size=13),
            text_color=("#666666", "#AAAAAA")
        ).pack(pady=(0, 5))
        
        # Scroll para información - se expande con la ventana
        info_frame = ctk.CTkScrollableFrame(
            confirm_window,
            fg_color=("#F0F0F0", "#2B2B2B"),
            corner_radius=10
        )
        info_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=5)
        
        # Verificar duplicados antes de mostrar
        duplicados = set()
        for archivo, resultado, error in todos_resultados:
            if error or not resultado:
                continue
            folio = resultado['viaje'].get('id_viaje')
            if folio and self.firebase.existe_viaje(str(folio)):
                duplicados.add(str(folio))
        
        # Banner de alerta si hay duplicados
        if duplicados:
            alert_frame = ctk.CTkFrame(confirm_window, fg_color=("#FFF3E0", "#3B2A00"), corner_radius=8, border_width=2, border_color="#FF9800")
            alert_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(5, 0))
            
            folios_dup = ", ".join(sorted(duplicados))
            ctk.CTkLabel(
                alert_frame,
                text=f"⚠️ Bitácora(s) ya existente(s): {folios_dup}",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color=("#E65100", "#FFB74D")
            ).pack(anchor="w", padx=12, pady=(8, 2))
            
            ctk.CTkLabel(
                alert_frame,
                text="Los datos anteriores serán reemplazados al confirmar la subida.",
                font=ctk.CTkFont(size=11),
                text_color=("#BF360C", "#FFCC80")
            ).pack(anchor="w", padx=12, pady=(0, 8))
        
        # Mostrar info de CADA archivo
        for file_idx, (archivo, resultado, error) in enumerate(todos_resultados):
            nombre_archivo = os.path.basename(archivo)
            
            # Separador entre bitácoras
            if file_idx > 0:
                sep = ctk.CTkFrame(info_frame, height=2, fg_color=("#05BFDB", "#0A4D68"))
                sep.pack(fill="x", padx=10, pady=(15, 5))
            
            if error:
                # Archivo con error
                ctk.CTkLabel(
                    info_frame,
                    text=f"❌ {nombre_archivo} - Error: {error}",
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color="#D32F2F"
                ).pack(anchor="w", padx=10, pady=5)
                continue
            
            viaje = resultado['viaje']
            lances = resultado['lances']
            
            # Filtrar lances
            lances_individuales = [l for l in lances if l.get('numero_lance', -1) != 0]
            num_lances_reales = len(lances_individuales)
            
            # Buscar CAPTURA TOTAL
            lance_captura_total = next((l for l in lances if l.get('numero_lance') == 0), None)
            
            # Título de la bitácora
            folio = viaje.get('id_viaje', 'N/A')
            nave = viaje.get('nave_nombre', 'N/A')
            es_duplicado = str(folio) in duplicados
            
            titulo_bg = ("#FF9800", "#CC7A00") if es_duplicado else ("#05BFDB", "#0A4D68")
            titulo_frame = ctk.CTkFrame(info_frame, fg_color=titulo_bg, corner_radius=6)
            titulo_frame.pack(fill="x", padx=10, pady=(8, 5))
            
            titulo_text = f"📄 Bitácora {file_idx + 1}/{len(todos_resultados)}: {nombre_archivo}"
            if es_duplicado:
                titulo_text = f"🔄 {titulo_text} (YA EXISTE - se reemplazará)"
            
            ctk.CTkLabel(
                titulo_frame,
                text=titulo_text,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="white"
            ).pack(anchor="w", padx=10, pady=6)
            
            # Detalles en formato compacto (sin Total Camarón, ya se muestra en la tabla de abajo)
            detalles = [
                ("📋 Folio/ID", folio),
                ("🚢 Nave", nave),
                ("👨‍✈️ Capitán", viaje.get('capitan', 'N/A')),
                ("🏢 Armador", viaje.get('armador', 'N/A')),
                ("🎣 Total Lances", num_lances_reales),
            ]
            
            for label, valor in detalles:
                detail_frame = ctk.CTkFrame(info_frame, fg_color=("#FFFFFF", "#1A1A1A"))
                detail_frame.pack(fill="x", padx=10, pady=2)
                
                ctk.CTkLabel(
                    detail_frame, text=label,
                    font=ctk.CTkFont(size=11, weight="bold"),
                    anchor="w", width=140
                ).pack(side="left", padx=8, pady=5)
                
                ctk.CTkLabel(
                    detail_frame, text=str(valor),
                    font=ctk.CTkFont(size=11), anchor="w"
                ).pack(side="left", padx=8, pady=5)
            
            # Tabla de CAPTURA TOTAL
            if lance_captura_total and lance_captura_total.get('especies'):
                ctk.CTkLabel(
                    info_frame,
                    text="📊 CAPTURA TOTAL:",
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color=("#0A4D68", "#05BFDB")
                ).pack(anchor="w", padx=10, pady=(8, 3))
                
                # Header
                header_ct = ctk.CTkFrame(info_frame, fg_color=("#05BFDB", "#0A4D68"))
                header_ct.pack(fill="x", padx=10, pady=(0, 1))
                
                ctk.CTkLabel(header_ct, text="Especie", font=ctk.CTkFont(size=10, weight="bold"),
                            text_color="white", anchor="w", width=240).pack(side="left", padx=6, pady=4)
                ctk.CTkLabel(header_ct, text="Tipo", font=ctk.CTkFont(size=10, weight="bold"),
                            text_color="white", anchor="center", width=85).pack(side="left", padx=4, pady=4)
                ctk.CTkLabel(header_ct, text="Cantidad", font=ctk.CTkFont(size=10, weight="bold"),
                            text_color="white", anchor="e", width=95).pack(side="right", padx=6, pady=4)
                
                # Filas
                especies_ct = lance_captura_total.get('especies', [])
                especies_ct_sorted = sorted(especies_ct, key=lambda e: (0 if e.get('tipo_captura') == 'retenida' else 1, e.get('nombre', '')))
                
                for esp in especies_ct_sorted:
                    tipo = esp.get('tipo_captura', 'retenida')
                    nombre_esp = esp.get('nombre', 'N/A')
                    cant_ton = esp.get('cantidad_ton', 0)
                    cant_uni = esp.get('cantidad_unidades', 0)
                    
                    bg_color = ("#E8F5E9", "#1B3B1B") if tipo == 'retenida' else ("#FFEBEE", "#3B1B1B")
                    tipo_label = "🎯 Ret." if tipo == 'retenida' else "🗑️ Desc."
                    tipo_color = ("#2E7D32", "#66BB6A") if tipo == 'retenida' else ("#C62828", "#EF5350")
                    
                    cant_str = f"{cant_ton:.3f} TON" if cant_ton > 0 else (f"{int(cant_uni)} unid." if cant_uni > 0 else "0")
                    
                    row_frame = ctk.CTkFrame(info_frame, fg_color=bg_color)
                    row_frame.pack(fill="x", padx=10, pady=1)
                    
                    ctk.CTkLabel(row_frame, text=nombre_esp, font=ctk.CTkFont(size=10),
                                anchor="w", width=240).pack(side="left", padx=6, pady=3)
                    ctk.CTkLabel(row_frame, text=tipo_label, font=ctk.CTkFont(size=9),
                                anchor="center", width=85, text_color=tipo_color).pack(side="left", padx=4, pady=3)
                    ctk.CTkLabel(row_frame, text=cant_str, font=ctk.CTkFont(size=10, weight="bold"),
                                anchor="e", width=95).pack(side="right", padx=6, pady=3)
        
        # Frame inferior fijo (comentarios + botones) - NO se mueve con scroll
        bottom_frame = ctk.CTkFrame(confirm_window, fg_color="transparent")
        bottom_frame.grid(row=3, column=0, sticky="ew", padx=20, pady=(5, 10))
        
        # Comentarios individuales por archivo
        comment_entries = {}  # {archivo: CTkTextbox}
        
        # Si hay varios archivos, mostrar un comentario por cada uno
        archivos_validos = [a for a, r, e in todos_resultados if not e]
        
        if len(archivos_validos) > 1:
            ctk.CTkLabel(
                bottom_frame,
                text="💬 Comentarios por bitácora:",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=("#0A4D68", "#05BFDB")
            ).pack(anchor="w", pady=(0, 4))
            
            comments_scroll = ctk.CTkScrollableFrame(
                bottom_frame,
                height=min(120, 50 * len(archivos_validos)),
                fg_color=("#F5F5F5", "#1A1A1A"),
                corner_radius=8
            )
            comments_scroll.pack(fill="x", pady=(0, 8))
            
            for archivo, resultado, error in todos_resultados:
                if error:
                    continue
                nombre_arch = os.path.basename(archivo)
                folio_arch = resultado['viaje'].get('id_viaje', 'N/A')
                
                row_frame = ctk.CTkFrame(comments_scroll, fg_color="transparent")
                row_frame.pack(fill="x", pady=2)
                
                ctk.CTkLabel(
                    row_frame,
                    text=f"📄 {folio_arch}:",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    width=100, anchor="w"
                ).pack(side="left", padx=(5, 5))
                
                entry = ctk.CTkTextbox(
                    row_frame,
                    height=35,
                    fg_color=("#FFFFFF", "#2B2B2B"),
                    border_width=2,
                    border_color="#05BFDB",
                    font=ctk.CTkFont(size=11)
                )
                entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
                comment_entries[archivo] = entry
        else:
            # Un solo archivo: comentario más grande
            comment_row = ctk.CTkFrame(bottom_frame, fg_color="transparent")
            comment_row.pack(fill="x", pady=(0, 8))
            
            ctk.CTkLabel(
                comment_row,
                text="💬 Comentario:",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=("#0A4D68", "#05BFDB")
            ).pack(anchor="w", pady=(0, 4))
            
            single_comment = ctk.CTkTextbox(
                comment_row,
                height=60,
                fg_color=("#FFFFFF", "#2B2B2B"),
                border_width=2,
                border_color="#05BFDB",
                font=ctk.CTkFont(size=12)
            )
            single_comment.pack(fill="x")
            if archivos_validos:
                comment_entries[archivos_validos[0]] = single_comment
        
        # Botones siempre visibles
        buttons_frame = ctk.CTkFrame(bottom_frame, fg_color="transparent")
        buttons_frame.pack(fill="x")
        
        def confirmar_y_subir():
            # Recoger comentarios individuales
            comentarios = {}
            for archivo, entry in comment_entries.items():
                texto = entry.get("1.0", "end").strip()
                if texto:
                    comentarios[archivo] = texto
            confirm_window.destroy()
            self.ejecutar_subida_multiple(archivos, comentarios if comentarios else None)
        
        def cancelar():
            confirm_window.destroy()
            self.archivos_pendientes = []
            self.mostrar_archivos_seleccionados()
        
        ctk.CTkButton(
            buttons_frame,
            text="✅  SÍ, SUBIR AL SISTEMA",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=45,
            fg_color="#4CAF50",
            hover_color="#388E3C",
            command=confirmar_y_subir
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        ctk.CTkButton(
            buttons_frame,
            text="❌  NO, CANCELAR",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=45,
            fg_color="#D32F2F",
            hover_color="#B71C1C",
            command=cancelar
        ).pack(side="left", padx=5, fill="x", expand=True)
    
    def ejecutar_subida_multiple(self, archivos, comentarios=None):
        """Ejecuta la subida de múltiples archivos en un hilo secundario
        
        Args:
            archivos: Lista de rutas de archivos PDF
            comentarios: dict {archivo: comentario} con comentarios individuales por archivo, o None
        """
        self._actualizar_estado(f"Subiendo {len(archivos)} archivo(s) a la nube...", "📤")
        # Ocultar lista de archivos
        self.files_list_frame.pack_forget()
        
        # Mostrar progress
        self.upload_progress.pack(pady=10)
        self.upload_status.pack()
        self.upload_btn.configure(state="disabled", text="⏳ Procesando...")
        
        def _subir_en_hilo():
            total_archivos = len(archivos)
            exitosos = 0
            fallidos = 0
            reemplazados = 0
            
            for i, archivo in enumerate(archivos, 1):
                nombre = os.path.basename(archivo)
                
                try:
                    # Actualizar UI desde hilo principal
                    self.root.after(0, lambda n=nombre, idx=i, total=total_archivos: (
                        self.upload_status.configure(
                            text=f"📄 Procesando {idx}/{total}: {n}",
                            text_color="#05BFDB"
                        ),
                        self.upload_progress.set(idx / total),
                        self._actualizar_estado(f"Procesando {idx}/{total}: {n}", "📄")
                    ))
                    
                    # Usar resultado cacheado si existe, sino parsear de nuevo
                    if archivo in self.resultados_parseados:
                        resultado = self.resultados_parseados[archivo]
                        print(f"✓ Usando resultado cacheado para {nombre}")
                    else:
                        with BitacoraParser(archivo) as parser:
                            resultado = parser.parsear_completo()
                    
                    if not resultado['viaje'].get('id_viaje'):
                        raise ValueError("No se pudo extraer el Folio del PDF")
                    
                    folio = resultado['viaje'].get('id_viaje', 'N/A')
                    nave = resultado['viaje'].get('nave_nombre', 'N/A')
                    es_reemplazo = self.firebase.existe_viaje(str(folio))
                    
                    exito = self.firebase.guardar_viaje_completo(resultado)
                    
                    if exito:
                        exitosos += 1
                        if es_reemplazo:
                            reemplazados += 1
                        
                        comentario = comentarios.get(archivo) if comentarios else None
                        
                        # Guardar comentario en Firebase para que otros equipos lo vean
                        if comentario:
                            try:
                                self.firebase.db.collection('viajes').document(str(folio)).update({
                                    'comentario': comentario
                                })
                            except Exception as e_com:
                                print(f"⚠️ No se pudo guardar comentario: {e_com}")
                        
                        # NO agregar notificación local — el polling la detectará igual
                        # (así es consistente con lo que ven los otros equipos)
                    else:
                        fallidos += 1
                        
                except Exception as e:
                    print(f"Error procesando {nombre}: {e}")
                    fallidos += 1
            
            # Finalizar en hilo principal
            self.root.after(0, lambda: self._finalizar_subida(exitosos, fallidos, reemplazados))
        
        threading.Thread(target=_subir_en_hilo, daemon=True).start()
    
    def _finalizar_subida(self, exitosos, fallidos, reemplazados):
        """Finaliza el proceso de subida (llamar desde hilo principal)"""
        # Limpiar caché de resultados
        self.resultados_parseados = {}
        self.archivos_pendientes = []
        
        try:
            self.upload_progress.pack_forget()
            self.upload_status.pack_forget()
            self.upload_btn.configure(state="normal", text="📁 Seleccionar Archivos PDF")
        except (tk.TclError, Exception):
            pass
        
        # Actualizar estadísticas (ya es no-bloqueante)
        self.update_stats()
        self._actualizar_estado(
            f"Subida completada — {exitosos} exitosa(s), {fallidos} fallida(s)", 
            "✅" if fallidos == 0 else "⚠️"
        )
        
        # Mensaje final
        if exitosos > 0:
            msg = f"Subida completada:\n\n✅ Exitosas: {exitosos}\n"
            if reemplazados > 0:
                msg += f"🔄 Reemplazadas: {reemplazados}\n"
            if fallidos > 0:
                msg += f"❌ Fallidas: {fallidos}\n"
            if reemplazados > 0:
                msg += f"\n⚠️ {reemplazados} bitácora(s) ya existían y fueron reemplazadas."
            
            CTkMessagebox(
                title="✅ Proceso Completado",
                message=msg,
                icon="check" if fallidos == 0 and reemplazados == 0 else "warning",
                option_1="OK"
            )
        else:
            CTkMessagebox(
                title="❌ Error",
                message=f"No se pudo procesar ningún archivo.\n\n"
                        f"❌ Fallidas: {fallidos}",
                icon="cancel",
                option_1="OK"
            )
    
    def verificar_notificaciones(self):
        """Verifica periódicamente si hay nuevas bitácoras subidas desde otros equipos.
        Las queries a Firebase se ejecutan en un hilo secundario para no bloquear la GUI."""
        if self._app_closing:
            return
        
        def _verificar_en_hilo():
            """Consulta Firebase en hilo secundario"""
            try:
                if not (self.firebase.db and self.internet_conectado):
                    return
                
                ids_actuales = self.firebase.obtener_ids_viajes()
                
                if not self._viajes_inicializados:
                    # Primera vez: detectar cambios desde la última ejecución
                    if hasattr(self, '_tiene_estado_previo') and self._tiene_estado_previo:
                        nuevos = ids_actuales - self._viajes_previos
                        eliminados = self._viajes_previos - ids_actuales
                        
                        # Generar notificaciones de cambios mientras estaba apagado
                        notifs_previas = []
                        for id_viaje in nuevos:
                            info = self.firebase.obtener_info_viaje(id_viaje)
                            if info:
                                nave = info.get('nave_nombre', 'N/A')
                                folio = info.get('id_viaje', id_viaje)
                                comentario = info.get('comentario', None)
                                notifs_previas.append((f"📥 Bitácora subida (mientras estabas offline): #{folio} - {nave}", folio, comentario))
                            else:
                                notifs_previas.append((f"📥 Bitácora subida (offline): #{id_viaje}", id_viaje, None))
                        
                        for id_viaje in eliminados:
                            notifs_previas.append((f"🗑️ Bitácora eliminada (mientras estabas offline): #{id_viaje}", id_viaje, None))
                        
                        # Enviar notificaciones previas al hilo principal
                        if notifs_previas and not self._app_closing:
                            self.root.after(0, lambda: self._procesar_notifs_remotas(notifs_previas))
                    
                    # Inicializar con estado actual
                    self._viajes_conocidos = ids_actuales
                    self._viajes_inicializados = True
                    return
                
                # Detección normal de cambios en tiempo real
                nuevos = ids_actuales - self._viajes_conocidos
                eliminados = self._viajes_conocidos - ids_actuales
                
                if not nuevos and not eliminados:
                    return
                
                # Obtener info de viajes nuevos (en este hilo background)
                notifs_nuevas = []
                for id_viaje in nuevos:
                    info = self.firebase.obtener_info_viaje(id_viaje)
                    if info:
                        nave = info.get('nave_nombre', 'N/A')
                        folio = info.get('id_viaje', id_viaje)
                        comentario = info.get('comentario', None)
                        notifs_nuevas.append((f"📥 Nueva bitácora subida: #{folio} - {nave}", folio, comentario))
                    else:
                        notifs_nuevas.append((f"📥 Nueva bitácora: #{id_viaje}", id_viaje, None))
                
                for id_viaje in eliminados:
                    notifs_nuevas.append((f"🗑️ Bitácora eliminada: #{id_viaje}", id_viaje, None))
                
                self._viajes_conocidos = ids_actuales
                
                # Enviar resultados al hilo principal
                if not self._app_closing:
                    self.root.after(0, lambda: self._procesar_notifs_remotas(notifs_nuevas))
                    
            except Exception as e:
                print(f"⚠️ Error verificando nuevas bitácoras: {e}")
        
        # Lanzar consulta en hilo secundario
        threading.Thread(target=_verificar_en_hilo, daemon=True).start()
        
        # Actualizar badge con lo que ya tenemos (sin bloquear)
        self._actualizar_badge_notificaciones()
        
        # Programar siguiente verificación (solo si no hay otro timer activo)
        if not hasattr(self, '_notif_timer_id') or self._notif_timer_id is None:
            self._notif_timer_id = self.root.after(30000, self._notif_timer_tick)
    
    def _notif_timer_tick(self):
        """Callback del timer periódico de notificaciones"""
        self._notif_timer_id = None  # Limpiar para permitir reprogramar
        if not self._app_closing:
            self.verificar_notificaciones()
    
    def _procesar_notifs_remotas(self, notifs_nuevas):
        """Procesa notificaciones descubiertas en el hilo background (llamar desde hilo principal)"""
        for notif in notifs_nuevas:
            mensaje, folio = notif[0], notif[1]
            comentario = notif[2] if len(notif) > 2 else None
            self.notificaciones.append({
                'mensaje': mensaje,
                'folio': folio,
                'comentario': comentario,
                'fecha': datetime.now(),
                'leida': False
            })
        self._actualizar_badge_notificaciones()
        self.update_stats()
    
    def _actualizar_badge_notificaciones(self):
        """Actualiza el badge visual de notificaciones (llamar desde hilo principal)"""
        try:
            no_leidas = sum(1 for n in self.notificaciones if not n.get('leida', False))
            if no_leidas > 0:
                self.notif_badge.configure(text=str(no_leidas))
                self.notif_badge.place(relx=0.75, rely=0.1, anchor="center")
            else:
                self.notif_badge.place_forget()
        except (tk.TclError, Exception):
            pass
    
    def agregar_notificacion(self, mensaje, folio=None, comentario=None):
        """Agrega una notificación a la lista y actualiza el badge"""
        self.notificaciones.append({
            'mensaje': mensaje,
            'folio': folio,
            'comentario': comentario,
            'fecha': datetime.now(),
            'leida': False
        })
        self._actualizar_badge_notificaciones()
    
    def mostrar_notificaciones(self):
        """Muestra panel de notificaciones"""
        # Marcar todas las notificaciones como leídas
        for notif in self.notificaciones:
            notif['leida'] = True
        
        # Actualizar badge inmediatamente
        try:
            self.notif_badge.place_forget()
        except (tk.TclError, Exception):
            pass
        
        # Crear ventana de notificaciones
        notif_window = ctk.CTkToplevel(self.root)
        notif_window.title("Centro de Notificaciones")
        notif_window.attributes('-alpha', 0.0)
        scr_w = notif_window.winfo_screenwidth()
        scr_h = notif_window.winfo_screenheight()
        n_w = min(600, scr_w - 40)
        n_h = min(650, scr_h - 100)
        notif_window.geometry(f"{n_w}x{n_h}")
        notif_window.transient(self.root)
        
        # Centrar ventana
        notif_window.update_idletasks()
        x = max(10, (scr_w - n_w) // 2)
        y = max(10, (scr_h - n_h) // 2 - 30)
        notif_window.geometry(f"+{x}+{y}")
        
        # Fade-in
        self._fade_in_widget(notif_window)
        
        # Centrar ventana
        notif_window.update_idletasks()
        x = max(10, (scr_w - n_w) // 2)
        y = max(10, (scr_h - n_h) // 2 - 30)
        notif_window.geometry(f"+{x}+{y}")
        
        # Título
        title_frame = ctk.CTkFrame(notif_window, fg_color="transparent")
        title_frame.pack(pady=20, fill="x")
        
        ctk.CTkLabel(
            title_frame,
            text="🔔 Centro de Notificaciones",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack()
        
        if self.notificaciones:
            ctk.CTkLabel(
                title_frame,
                text=f"{len(self.notificaciones)} notificación(es) nueva(s)",
                font=ctk.CTkFont(size=12),
                text_color=("#666666", "#AAAAAA")
            ).pack(pady=(5, 0))
        
        # Scroll de notificaciones
        notif_scroll = ctk.CTkScrollableFrame(
            notif_window,
            width=550,
            height=450,
            fg_color=("#F0F0F0", "#2B2B2B")
        )
        notif_scroll.pack(pady=10, padx=20, fill="both", expand=True)
        
        if not self.notificaciones:
            ctk.CTkLabel(
                notif_scroll,
                text="No hay notificaciones nuevas",
                font=ctk.CTkFont(size=14),
                text_color=("#999999", "#666666")
            ).pack(pady=50)
        else:
            for notif in reversed(self.notificaciones):
                notif_frame = ctk.CTkFrame(
                    notif_scroll, 
                    fg_color=("#FFFFFF", "#1A1A1A"),
                    border_width=2,
                    border_color=("#05BFDB", "#05BFDB")
                )
                notif_frame.pack(fill="x", padx=10, pady=5)
                
                # Contenedor principal
                content_frame = ctk.CTkFrame(notif_frame, fg_color="transparent")
                content_frame.pack(fill="both", expand=True, padx=10, pady=10)
                
                # Mensaje principal
                ctk.CTkLabel(
                    content_frame,
                    text=notif['mensaje'],
                    font=ctk.CTkFont(size=13, weight="bold"),
                    anchor="w",
                    text_color=("#0A4D68", "#05BFDB")
                ).pack(anchor="w")
                
                # Fila de detalles (folio, comentario, hora)
                details_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
                details_frame.pack(anchor="w", pady=(5, 0))
                
                # Folio si existe
                if notif.get('folio'):
                    ctk.CTkLabel(
                        details_frame,
                        text=f"📋 Folio: {notif['folio']}",
                        font=ctk.CTkFont(size=11),
                        text_color=("#666666", "#AAAAAA")
                    ).pack(side="left", padx=(0, 15))
                
                # Indicador de comentario
                if notif.get('comentario'):
                    ctk.CTkLabel(
                        details_frame,
                        text="💬 Con comentario",
                        font=ctk.CTkFont(size=11),
                        text_color=("#4CAF50", "#66BB6A")
                    ).pack(side="left", padx=(0, 15))
                else:
                    ctk.CTkLabel(
                        details_frame,
                        text="Sin comentario",
                        font=ctk.CTkFont(size=11),
                        text_color=("#999999", "#666666")
                    ).pack(side="left", padx=(0, 15))
                
                # Hora
                ctk.CTkLabel(
                    details_frame,
                    text=f"🕐 {notif['fecha'].strftime('%H:%M')}",
                    font=ctk.CTkFont(size=11),
                    text_color=("#666666", "#AAAAAA")
                ).pack(side="left")
                
                # Mostrar comentario si existe
                if notif.get('comentario'):
                    comment_frame = ctk.CTkFrame(
                        content_frame,
                        fg_color=("#F0F8FF", "#1A3A4A"),
                        corner_radius=5
                    )
                    comment_frame.pack(fill="x", pady=(8, 0))
                    
                    ctk.CTkLabel(
                        comment_frame,
                        text=f"💬 \"{notif['comentario']}\"",
                        font=ctk.CTkFont(size=11),
                        anchor="w",
                        wraplength=400,
                        text_color=("#0A4D68", "#88D4E8")
                    ).pack(padx=8, pady=6, anchor="w")
        
        # Botón para limpiar todas las notificaciones
        if self.notificaciones:
            btn_frame = ctk.CTkFrame(notif_window, fg_color="transparent")
            btn_frame.pack(pady=15)
            
            ctk.CTkButton(
                btn_frame,
                text="🗑️ Limpiar Todas las Notificaciones",
                font=ctk.CTkFont(size=14, weight="bold"),
                width=300,
                height=40,
                fg_color="#FF5722",
                hover_color="#E64A19",
                command=lambda: self.limpiar_notificaciones(notif_window)
            ).pack()
    
    def limpiar_notificaciones(self, window):
        """Limpia todas las notificaciones"""
        self.notificaciones = []
        self._actualizar_badge_notificaciones()
        window.destroy()
    
    def abrir_panel_administrador(self):
        """Abre el panel de administrador con selección múltiple y eliminación masiva"""
        # Verificar conexión a internet
        if self._sin_internet_alerta("acceder al panel de administrador"):
            return
        
        # Crear ventana de administrador
        admin_window = ctk.CTkToplevel(self.root)
        admin_window.title("Administrar Datos")
        admin_window.attributes('-alpha', 0.0)
        scr_w = admin_window.winfo_screenwidth()
        scr_h = admin_window.winfo_screenheight()
        a_w = min(900, scr_w - 40)
        a_h = min(700, scr_h - 100)
        admin_window.geometry(f"{a_w}x{a_h}+{max(10, (scr_w - a_w) // 2)}+{max(10, (scr_h - a_h) // 2 - 30)}")
        admin_window.transient(self.root)
        
        # Fade-in
        self._fade_in_widget(admin_window)
        
        # Estado de selección
        checkboxes_var = {}  # {id_viaje: BooleanVar}
        checkbox_widgets = {}  # {id_viaje: checkbox_widget}
        
        # Título
        header = ctk.CTkFrame(admin_window, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(20, 5))
        
        ctk.CTkLabel(
            header,
            text="⚙️ Administrar Datos",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(side="left")
        
        # Barra de acciones (selección masiva)
        action_bar = ctk.CTkFrame(admin_window, fg_color=("#E0E8ED", "#111D2E"), corner_radius=10)
        action_bar.pack(fill="x", padx=30, pady=(10, 5))
        
        select_all_var = ctk.BooleanVar(value=False)
        
        def toggle_select_all():
            val = select_all_var.get()
            for vid, var in checkboxes_var.items():
                var.set(val)
            actualizar_contador()
        
        select_all_cb = ctk.CTkCheckBox(
            action_bar,
            text="Seleccionar todo",
            variable=select_all_var,
            font=ctk.CTkFont(size=12),
            command=toggle_select_all,
            fg_color="#05BFDB",
            hover_color="#057A8F",
            border_color="#3A6A7A"
        )
        select_all_cb.pack(side="left", padx=15, pady=10)
        
        sel_count_label = ctk.CTkLabel(
            action_bar,
            text="0 seleccionadas",
            font=ctk.CTkFont(size=12),
            text_color=("#666666", "#AAAAAA")
        )
        sel_count_label.pack(side="left", padx=10)
        
        def actualizar_contador():
            n = sum(1 for v in checkboxes_var.values() if v.get())
            sel_count_label.configure(text=f"{n} seleccionada(s)")
            btn_eliminar_sel.configure(
                state="normal" if n > 0 else "disabled",
                fg_color="#D32F2F" if n > 0 else "#555555"
            )
        
        btn_eliminar_sel = ctk.CTkButton(
            action_bar,
            text="🗑️ Eliminar seleccionadas",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=35,
            fg_color="#555555",
            hover_color="#B71C1C",
            state="disabled",
            command=lambda: eliminar_seleccionadas()
        )
        btn_eliminar_sel.pack(side="right", padx=15, pady=10)
        
        # Frame de búsqueda
        search_frame = ctk.CTkFrame(admin_window, fg_color="transparent")
        search_frame.pack(pady=(10, 5), padx=30, fill="x")
        
        ctk.CTkLabel(
            search_frame,
            text="🔍",
            font=ctk.CTkFont(size=14)
        ).pack(side="left", padx=(0, 5))
        
        search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="Buscar por ID, nave o capitán...",
            height=36,
            corner_radius=8,
            border_color="#1B3A4B"
        )
        search_entry.pack(side="left", padx=5, fill="x", expand=True)
        
        ctk.CTkButton(
            search_frame,
            text="Buscar",
            width=90,
            height=36,
            corner_radius=8,
            fg_color="#05BFDB",
            hover_color="#057A8F",
            command=lambda: cargar_bitacoras_admin(viajes_scroll, search_entry.get().strip())
        ).pack(side="left", padx=5)
        
        # Scroll de bitácoras
        viajes_scroll = ctk.CTkScrollableFrame(
            admin_window,
            fg_color=("#E8EEF2", "#161E2E")
        )
        viajes_scroll.pack(pady=5, padx=30, fill="both", expand=True)
        
        def cargar_bitacoras_admin(scroll_frame, filtro=""):
            # Limpiar
            for widget in scroll_frame.winfo_children():
                widget.destroy()
            checkboxes_var.clear()
            checkbox_widgets.clear()
            select_all_var.set(False)
            actualizar_contador()
            
            # Mostrar loading
            loading_label = ctk.CTkLabel(
                scroll_frame,
                text="⏳ Cargando bitácoras...",
                font=ctk.CTkFont(size=14),
                text_color=("#999999", "#666666")
            )
            loading_label.pack(pady=50)
            
            def _cargar():
                viajes = self.firebase.listar_viajes(limite=100)
                self.root.after(0, lambda: _render_bitacoras_admin(scroll_frame, viajes, filtro))
            
            threading.Thread(target=_cargar, daemon=True).start()
        
        def _render_bitacoras_admin(scroll_frame, viajes, filtro):
            # Limpiar loading
            for widget in scroll_frame.winfo_children():
                widget.destroy()
            
            if filtro:
                filtro_lower = filtro.lower()
                viajes = [
                    v for v in viajes 
                    if filtro_lower in str(v.get('id_viaje', '')).lower() or
                       filtro_lower in v.get('nave_nombre', '').lower() or
                       filtro_lower in v.get('capitan', '').lower()
                ]
            
            if not viajes:
                ctk.CTkLabel(
                    scroll_frame,
                    text="No se encontraron bitácoras",
                    font=ctk.CTkFont(size=14),
                    text_color=("#999999", "#666666")
                ).pack(pady=50)
                return
            
            for viaje in viajes:
                vid = viaje.get('id_viaje', 'N/A')
                
                row = ctk.CTkFrame(scroll_frame, fg_color=("#FFFFFF", "#1A1A2A"), corner_radius=8)
                row.pack(fill="x", padx=8, pady=3)
                
                # Checkbox
                var = ctk.BooleanVar(value=False)
                checkboxes_var[vid] = var
                
                cb = ctk.CTkCheckBox(
                    row, text="",
                    variable=var,
                    width=24,
                    fg_color="#05BFDB",
                    hover_color="#057A8F",
                    border_color="#3A6A7A",
                    command=actualizar_contador
                )
                cb.pack(side="left", padx=(12, 4), pady=10)
                checkbox_widgets[vid] = cb
                
                # Info
                info = ctk.CTkFrame(row, fg_color="transparent")
                info.pack(side="left", fill="x", expand=True, padx=6, pady=8)
                
                ctk.CTkLabel(
                    info,
                    text=f"📋 {vid}",
                    font=ctk.CTkFont(size=13, weight="bold"),
                    anchor="w"
                ).pack(anchor="w")
                
                detalles = f"Nave: {viaje.get('nave_nombre', 'N/A')}  •  Capitán: {viaje.get('capitan', 'N/A')}"
                fecha_proc = viaje.get('fecha_procesamiento', '')
                if fecha_proc:
                    try:
                        dt = datetime.fromisoformat(fecha_proc)
                        detalles += f"  •  {dt.strftime('%d/%m/%Y %H:%M')}"
                    except Exception:
                        pass
                
                ctk.CTkLabel(
                    info,
                    text=detalles,
                    font=ctk.CTkFont(size=11),
                    text_color=("#666666", "#8A8A9A"),
                    anchor="w"
                ).pack(anchor="w")
                
                # Botón eliminar individual
                ctk.CTkButton(
                    row,
                    text="🗑️",
                    width=36,
                    height=36,
                    corner_radius=8,
                    fg_color="transparent",
                    hover_color="#D32F2F",
                    text_color=("#D32F2F", "#FF5252"),
                    font=ctk.CTkFont(size=14),
                    command=lambda v=vid: confirmar_eliminar_uno(v)
                ).pack(side="right", padx=10)
        
        def confirmar_eliminar_uno(v_id):
            resultado = CTkMessagebox(
                title="⚠️ Confirmar Eliminación",
                message=f"¿Eliminar la bitácora {v_id}?\n\nEsta acción NO se puede deshacer.",
                icon="warning",
                option_1="Cancelar",
                option_2="Eliminar"
            )
            if resultado.get() == "Eliminar":
                def _eliminar():
                    try:
                        ok = self.firebase.eliminar_viaje(v_id)
                        self.root.after(0, lambda: _post_eliminar_uno(ok, v_id))
                    except Exception as e:
                        self.root.after(0, lambda: CTkMessagebox(
                            title="❌ Error", message=f"Error: {str(e)}", icon="cancel", option_1="OK"))
                
                threading.Thread(target=_eliminar, daemon=True).start()
        
        def _post_eliminar_uno(ok, v_id):
            if ok:
                CTkMessagebox(title="✅ Éxito", message=f"Bitácora {v_id} eliminada", icon="check", option_1="OK")
            else:
                CTkMessagebox(title="❌ Error", message=f"No se pudo eliminar {v_id}", icon="cancel", option_1="OK")
            cargar_bitacoras_admin(viajes_scroll, search_entry.get().strip())
            self.update_stats()
        
        def eliminar_seleccionadas():
            seleccionados = [vid for vid, var in checkboxes_var.items() if var.get()]
            if not seleccionados:
                return
            
            n = len(seleccionados)
            resultado = CTkMessagebox(
                title="⚠️ Eliminación Masiva",
                message=f"¿Eliminar {n} bitácora(s) seleccionada(s)?\n\nEsta acción NO se puede deshacer.",
                icon="warning",
                option_1="Cancelar",
                option_2=f"Eliminar {n}"
            )
            if resultado.get() == f"Eliminar {n}":
                def _eliminar_lote():
                    errores = []
                    for vid in seleccionados:
                        try:
                            self.firebase.eliminar_viaje(vid)
                        except Exception as e:
                            errores.append(f"{vid}: {str(e)}")
                    self.root.after(0, lambda: _post_eliminar_lote(n, errores))
                
                threading.Thread(target=_eliminar_lote, daemon=True).start()
        
        def _post_eliminar_lote(n, errores):
            exitos = n - len(errores)
            if errores:
                CTkMessagebox(
                    title="⚠️ Resultado",
                    message=f"{exitos} eliminada(s) correctamente.\n{len(errores)} error(es):\n" + "\n".join(errores[:5]),
                    icon="warning",
                    option_1="OK"
                )
            else:
                CTkMessagebox(
                    title="✅ Éxito",
                    message=f"{exitos} bitácora(s) eliminada(s) correctamente",
                    icon="check",
                    option_1="OK"
                )
            cargar_bitacoras_admin(viajes_scroll, search_entry.get().strip())
            self.update_stats()
        
        # Cargar bitácoras inicialmente
        cargar_bitacoras_admin(viajes_scroll)
        
        # Advertencia inferior
        ctk.CTkLabel(
            admin_window,
            text="⚠️ Las eliminaciones son permanentes e irreversibles",
            font=ctk.CTkFont(size=11),
            text_color="#FF5252"
        ).pack(pady=(5, 12))
    
    def abrir_reporte_bug(self):
        """Abre una ventana para que el usuario reporte un bug"""
        bug_window = ctk.CTkToplevel(self.root)
        bug_window.title("Reportar un Problema")
        bug_window.attributes('-alpha', 0.0)
        scr_w = bug_window.winfo_screenwidth()
        scr_h = bug_window.winfo_screenheight()
        w, h = 520, 480
        bug_window.geometry(f"{w}x{h}+{(scr_w - w) // 2}+{(scr_h - h) // 2}")
        bug_window.resizable(False, False)
        bug_window.transient(self.root)
        
        self._fade_in_widget(bug_window)
        
        # Header
        ctk.CTkLabel(
            bug_window,
            text="🐛 Reportar un Problema",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=("#0A4D68", "#05BFDB")
        ).pack(pady=(22, 5))
        
        ctk.CTkLabel(
            bug_window,
            text="Describe el problema y lo revisaremos lo antes posible",
            font=ctk.CTkFont(size=12),
            text_color=("#666666", "#AAAAAA")
        ).pack(pady=(0, 15))
        
        form = ctk.CTkFrame(bug_window, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=30)
        
        # Título del bug
        ctk.CTkLabel(
            form, text="Título (breve)",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=("#333333", "#CCCCCC")
        ).pack(anchor="w")
        
        titulo_entry = ctk.CTkEntry(
            form, height=38,
            placeholder_text="Ej: Error al subir un PDF",
            border_color="#1B3A4B",
            corner_radius=8
        )
        titulo_entry.pack(fill="x", pady=(3, 12))
        
        # Sección donde ocurrió
        ctk.CTkLabel(
            form, text="¿En qué sección ocurrió?",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=("#333333", "#CCCCCC")
        ).pack(anchor="w")
        
        seccion_combo = ctk.CTkComboBox(
            form,
            values=["Subir Bitácoras", "Buscar Datos", "Administrar Datos", "Login", "Otro"],
            height=38,
            border_color="#1B3A4B",
            corner_radius=8
        )
        seccion_combo.pack(fill="x", pady=(3, 12))
        seccion_combo.set("Subir Bitácoras")
        
        # Descripción
        ctk.CTkLabel(
            form, text="Descripción detallada",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=("#333333", "#CCCCCC")
        ).pack(anchor="w")
        
        desc_textbox = ctk.CTkTextbox(
            form, height=120,
            border_color="#1B3A4B",
            corner_radius=8,
            font=ctk.CTkFont(size=12)
        )
        desc_textbox.pack(fill="x", pady=(3, 15))
        
        # Estado de envío
        status_label = ctk.CTkLabel(
            form, text="",
            font=ctk.CTkFont(size=12)
        )
        status_label.pack()
        
        def enviar_reporte():
            titulo = titulo_entry.get().strip()
            descripcion = desc_textbox.get("1.0", "end").strip()
            seccion = seccion_combo.get()
            
            if not titulo:
                status_label.configure(text="⚠️ Escribe un título", text_color="#FF9800")
                return
            if not descripcion:
                status_label.configure(text="⚠️ Describe el problema", text_color="#FF9800")
                return
            
            if not self.firebase.db:
                status_label.configure(text="❌ Sin conexión a la nube", text_color="#FF5252")
                return
            
            ok = self.firebase.guardar_reporte_bug(titulo, descripcion, seccion)
            if ok:
                status_label.configure(text="✅ Reporte enviado. ¡Gracias!", text_color="#4CAF50")
                # Deshabilitar para evitar duplicados
                enviar_btn.configure(state="disabled", fg_color="#555555")
                titulo_entry.configure(state="disabled")
                desc_textbox.configure(state="disabled")
                bug_window.after(2500, bug_window.destroy)
            else:
                status_label.configure(text="❌ Error al enviar el reporte", text_color="#FF5252")
        
        enviar_btn = ctk.CTkButton(
            form,
            text="📨 Enviar Reporte",
            height=42,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="#05BFDB",
            hover_color="#057A8F",
            corner_radius=8,
            command=enviar_reporte
        )
        enviar_btn.pack(fill="x", pady=(5, 0))
    
    def run(self):
        """Ejecuta la aplicación"""
        self.root.mainloop()


def main():
    """Función principal: Splash → Login → Aplicación"""
    
    # Verificar si hay una actualización pendiente (fallback si el .bat falló)
    if aplicar_actualizacion_pendiente():
        print("🔄 Actualización pendiente aplicada, reiniciando...")
        os._exit(0)  # La función ya lanzó el nuevo exe
    
    def abrir_app():
        app = BentosApp()
        app.run()
    
    def abrir_login():
        login = LoginScreen(on_success=abrir_app)
        login.show()
    
    # Mostrar splash screen primero, luego login
    splash = SplashScreen(abrir_login)
    splash.show()


if __name__ == "__main__":
    main()
